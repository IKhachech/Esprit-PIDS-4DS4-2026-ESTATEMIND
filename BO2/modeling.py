"""
modeling.py — Post-cleaning modeling and feature engineering.

KEY CHANGES vs previous version:
  - MARKET_CONTEXT : computed from real BCT + INS data (external_data.py)
  - NEGO_BASE       : computed from real INS immobilier glissements
  - market_tension  : enriched with score_attractivite + population from signaux
  - image_quality_score : enriched with nb_immo_direct from signaux
  - temp_high_season : derived from actual listing volume seasonality
  - prix_transaction_estimated : uses real nego rates, computed for all rows (prix NaN dropped in step 5)
"""

import os, re, json, warnings
import numpy as np
import pandas as pd
from datetime import datetime

warnings.filterwarnings('ignore')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from mappings import (
    GOUVERNORAT_ENC, GOUVERNORAT_DEC, TYPE_BIEN_DEC, NLP_KW,
    CYCLE_MARCHE_ENC, CYCLE_MARCHE_DEC, RESEGMENT_MAP,
    encode_cycle_marche, encode_gouvernorat, encode_type_bien,
)
from cleaning import _gov_names, _gov_lats, _gov_lons, _KM_PER_LAT, _KM_PER_LON


def section(t): print("\n" + "=" * 65 + f"\n   {t}\n" + "=" * 65)
def log(m):     print(f"  {m}")


# ================================================================
# DISTRIBUTION VISUALIZATION
# ================================================================

def plot_distributions(groupes_clean: dict[str, pd.DataFrame]) -> None:
    section("ETAPE 5c — VISUALISATION DISTRIBUTIONS & OUTLIERS (IQR)")
    os.makedirs('plots', exist_ok=True)
    COLS_VIZ = ['prix', 'surface_m2', 'chambres', 'pieces']

    for groupe, dg in groupes_clean.items():
        cols_avail = [c for c in COLS_VIZ if c in dg.columns and dg[c].notna().sum() > 10]
        if not cols_avail: continue

        n_cols = len(cols_avail)
        fig, axes = plt.subplots(3, n_cols, figsize=(5 * n_cols, 14))
        if n_cols == 1: axes = axes.reshape(3, 1)
        fig.suptitle(f'[{groupe}] — Distributions & Outliers', fontsize=13, fontweight='bold')

        for i, col in enumerate(cols_avail):
            data = dg[col].dropna()

            ax = axes[0][i]
            ax.hist(data, bins=50, color='steelblue', edgecolor='white', alpha=0.85)
            ax.set_title(f'{col} — Histogramme', fontsize=10)
            ax.set_ylabel('Fréquence')
            ax.axvline(data.mean(), color='red', linestyle='--', linewidth=1.2,
                       label=f'Moy={data.mean():,.0f}')
            ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
            ax.tick_params(axis='x', rotation=30)
            ax.legend(fontsize=7)

            ax = axes[1][i]
            ax.boxplot(data, vert=True, patch_artist=True, widths=0.5,
                       boxprops=dict(facecolor='#AED6F1', color='navy'),
                       medianprops=dict(color='darkred', linewidth=2),
                       flierprops=dict(marker='o', markersize=3, markerfacecolor='red', alpha=0.4),
                       whiskerprops=dict(linestyle='--', color='gray'))
            Q1, Q3 = data.quantile(0.25), data.quantile(0.75)
            IQR = Q3 - Q1
            low_b, high_b = Q1 - 1.5 * IQR, Q3 + 1.5 * IQR
            n_out = ((data < low_b) | (data > high_b)).sum()
            ax.axhline(low_b,  color='orange', linestyle=':', linewidth=1.2)
            ax.axhline(high_b, color='orange', linestyle=':', linewidth=1.2)
            ax.set_title(f'{col} — Boxplot | {n_out} outliers', fontsize=9)
            ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))

            ax = axes[2][i]
            if col != 'prix' and 'prix' in dg.columns:
                both = dg[[col, 'prix']].dropna()
                ax.scatter(both[col], both['prix'], alpha=0.25, s=10, color='teal')
                if len(both) > 10:
                    try:
                        z  = np.polyfit(both[col], both['prix'], 1)
                        xr = np.linspace(both[col].min(), both[col].max(), 100)
                        ax.plot(xr, np.poly1d(z)(xr), 'r--', linewidth=1.2)
                    except: pass
                ax.set_xlabel(col); ax.set_ylabel('prix (TND)')
                ax.set_title(f'{col} vs prix', fontsize=9)
            else:
                ax.set_visible(False)

        plt.tight_layout()
        fp = f'plots/{groupe.lower()}_analyse.png'
        plt.savefig(fp, dpi=120, bbox_inches='tight')
        plt.close()
        log(f"  Plot : {fp}")


# ================================================================
# MARKET FEATURES — enriched with real signaux data
# ================================================================

def compute_market_features(
    df: pd.DataFrame,
    satellite_dir: str,
    gov_features: dict,
) -> pd.DataFrame:
    """
    Calcule market_tension, score_attractivite et features de marché.

    IMPORTANT: utilise map() sur l'index original — jamais merge() qui reset l'index
    et casse tous les join() ultérieurs (cycle_marche, etc.).

    Poids market_tension :
      35% tension locative  (depuis tes annonces)
      30% liquidité marché  (depuis tes annonces)
      25% score_attractivite (Google Maps signaux)
      10% nb_immo_direct_norm (profondeur marché)
    """
    section("ETAPE 7 — DYNAMIQUE MARCHE : market_tension")

    # ── Clé normalisée par ville (lowercase) ─────────────────────
    df['_ville_key'] = df['ville'].fillna('inconnu').str.lower().str.strip()

    # ── Calcul par ville → dict {ville_key: valeur} ──────────────
    # On construit des dicts puis on map() — JAMAIS merge() pour
    # conserver l'index original du df.

    # Liquidité = volume annonces par ville
    vol_series   = df.groupby('_ville_key')['prix'].count()
    max_vol      = vol_series.max()
    liq_dict     = (vol_series / max_vol).round(4).to_dict()

    # Tension locative = nb locations / nb total par ville
    loc_series   = df[df['type_transaction'] == 1].groupby('_ville_key').size()
    tot_series   = df.groupby('_ville_key').size()
    tension_dict = (loc_series / tot_series).fillna(0).round(4).to_dict()

    # Attractivité et immo_direct depuis signaux (via gouvernorat)
    from mappings import normalize_gouvernorat, GOUVERNORAT_DEC
    gov_attr = {gov: feat.get('score_attractivite', 0.5) for gov, feat in gov_features.items()} if gov_features else {}
    gov_immo = {gov: feat.get('nb_immo_direct_norm', 0.0) for gov, feat in gov_features.items()} if gov_features else {}

    def _ville_to_attr(vk):
        gov = normalize_gouvernorat(vk)
        return gov_attr.get(gov, 0.5)

    def _ville_to_immo(vk):
        gov = normalize_gouvernorat(vk)
        return gov_immo.get(gov, 0.0)

    # Build per-ville tension dict
    unique_villes = df['_ville_key'].unique()
    attr_dict  = {v: _ville_to_attr(v)  for v in unique_villes}
    immo_dict  = {v: _ville_to_immo(v)  for v in unique_villes}

    tension_loc_v = df['_ville_key'].map(tension_dict).fillna(0.0)
    liq_v         = df['_ville_key'].map(liq_dict).fillna(0.1)
    attr_v        = df['_ville_key'].map(attr_dict).fillna(0.5)
    immo_v        = df['_ville_key'].map(immo_dict).fillna(0.0)

    # ── Assign columns directly (no merge → index preserved) ─────
    df['market_rental_tension']  = tension_loc_v.values
    df['market_liquidity_score'] = liq_v.values
    df['market_tension'] = (
        tension_loc_v * 0.35 + liq_v * 0.30 + attr_v * 0.25 + immo_v * 0.10
    ).clip(0, 1).round(4).values

    if not gov_features:
        # Fallback: population.csv
        pop_density: dict = {}
        pop_path = f'{satellite_dir}/tunisia_population_20260215_133916.csv'
        if os.path.exists(pop_path):
            df_pop = pd.read_csv(pop_path)
            DSCORE = {'very_high': 5, 'high': 4, 'medium': 3, 'low': 2, 'very_low': 1}
            df_pop['dscore'] = df_pop['density_category'].str.lower().map(DSCORE).fillna(3)
            pop_density = {k.lower().strip(): v
                           for k, v in df_pop.groupby('city')['dscore'].mean().items()}
        pop_v = df['_ville_key'].map(pop_density).fillna(3.0)
        df['market_tension'] = (
            tension_loc_v * 0.40 + liq_v * 0.35 + (pop_v / 5.0) * 0.25
        ).clip(0, 1).round(4).values
        log("  market_tension calculé avec population.csv (fallback)")
    else:
        log("  market_tension calculé avec score_attractivite signaux")

    # ── score_attractivite par gouvernorat (via gouvernorat int) ──
    # Utilise la colonne gouvernorat (int encodé) — toujours présente
    gov_score_map = {gov: feat.get('score_attractivite', 0.5)
                     for gov, feat in gov_features.items()} if gov_features else {}
    gov_pop_map   = {gov: feat.get('population', 0.0)
                     for gov, feat in gov_features.items()} if gov_features else {}
    gov_immo_map  = {gov: feat.get('nb_immo_direct', 0.0)
                     for gov, feat in gov_features.items()} if gov_features else {}

    # gouvernorat int → gouvernorat string → score
    df['score_attractivite'] = (
        df['gouvernorat']
        .map(GOUVERNORAT_DEC)                    # int → string
        .map(gov_score_map)                      # string → score
        .fillna(0.5)
    )
    df['population_gouvernorat'] = (
        df['gouvernorat'].map(GOUVERNORAT_DEC).map(gov_pop_map).fillna(0.0)
    )
    df['nb_immo_direct'] = (
        df['gouvernorat'].map(GOUVERNORAT_DEC).map(gov_immo_map).fillna(0.0)
    )

    df = df.drop(columns=['_ville_key'], errors='ignore')
    log(f"  market_tension      : median={df['market_tension'].median():.3f}")
    log(f"  score_attractivite  : median={df['score_attractivite'].median():.3f}")
    log(f"  NaN market_tension  : {df['market_tension'].isna().sum()}")
    log(f"  NaN score_attractivite : {df['score_attractivite'].isna().sum()}")
    return df


# ================================================================
# TEMPORAL FEATURES — cycle_marche from real BCT+INS market_context
# ================================================================

def compute_temporal_features(df: pd.DataFrame, market_context: dict) -> pd.DataFrame:
    """
    Uses market_context computed from real BCT + INS data.
    temp_high_season is derived from actual listing volume seasonality
    in the dataset itself (replaces hardcoded months 3,4,5,9,10,11).
    """
    section("ETAPE 8 — FEATURES TEMPORELLES")

    NOW = datetime.now()

    # Derive high-season months from actual listing distribution
    if 'date_publication' in df.columns:
        dates = pd.to_datetime(df['date_publication'], errors='coerce').dropna()
        if len(dates) > 100:
            monthly_counts = dates.dt.month.value_counts()
            median_count   = monthly_counts.median()
            high_season_months = set(monthly_counts[monthly_counts > median_count].index.tolist())
            log(f"  Haute saison calculée depuis données : mois {sorted(high_season_months)}")
        else:
            high_season_months = {3, 4, 5, 9, 10, 11}  # fallback
            log(f"  Haute saison fallback (peu de données) : {sorted(high_season_months)}")
    else:
        high_season_months = {3, 4, 5, 9, 10, 11}

    def get_temporal_features(date_val):
        dt = NOW
        if not pd.isna(date_val):
            try:
                dt = pd.to_datetime(date_val)
                if dt > NOW or dt.year < 2020: dt = NOW
            except: dt = NOW
        year, month  = dt.year, dt.month
        quarter      = (month - 1) // 3 + 1
        ctx          = market_context.get((year, quarter), (7.0, 'stabilization'))
        return {
            'temp_days_on_market': max(0, (NOW - dt).days),
            'temp_high_season':    1 if month in high_season_months else 0,
            '_cycle_marche_str':   ctx[1],
            'taux_directeur':      ctx[0],
        }

    log("Calcul features temporelles...")
    # Compute features row by row and assign directly — no join()/merge()
    # to preserve the original index (avoids NaN from index mismatch)
    results = [get_temporal_features(row.get('date_publication'))
               for _, row in df.iterrows()]

    df['temp_days_on_market'] = [r['temp_days_on_market'] for r in results]
    df['temp_high_season']    = [r['temp_high_season']    for r in results]
    df['taux_directeur']      = [r['taux_directeur']      for r in results]
    _cycle_strs               = [r['_cycle_marche_str']   for r in results]

    log(f"  cycle_marche : {pd.Series(_cycle_strs).value_counts().to_dict()}")
    df['cycle_marche'] = [encode_cycle_marche(c) for c in _cycle_strs]
    log(f"  cycle_marche encodé : {df['cycle_marche'].value_counts().sort_index().to_dict()}")
    log(f"  NaN cycle_marche    : {df['cycle_marche'].isna().sum()}")
    return df


# ================================================================
# VISION FEATURES — image quality enriched by signaux immo_direct
# ================================================================

def compute_image_features(row, gov_features: dict) -> dict:
    """
    image_quality_score now uses nb_immo_direct (real data) instead of
    hardcoded per-source quality scores.
    Source quality is still used but as a secondary signal.
    """
    url    = str(row.get('image_url', '') or '')
    source = str(row.get('source', '') or '').lower()
    prix   = row.get('prix')
    surf   = row.get('surface_m2')
    liq    = row.get('market_liquidity_score', 0.0)

    f_url_valid = 1.0 if url.startswith('http') else 0.0

    # Source quality: use nb_immo_direct from signaux if available
    gov_str = str(row.get('_gouvernorat_str', '') or '')
    gov_f   = gov_features.get(gov_str, {}) if gov_features else {}
    nb_immo_norm = float(gov_f.get('nb_immo_direct_norm', 0.0) or 0.0)

    # Fallback per-source quality (secondary signal, not primary)
    SOURCE_QUALITY = {
        'century21': 0.95, 'homeintunisia': 0.90, 'mubawab': 0.85,
        'tayara': 0.70, 'tunisie annonces': 0.65,
        'mubawab partial': 0.75, 'facebook marketplace': 0.50, 'bnb': 0.75,
    }
    f_source = next((v for k, v in SOURCE_QUALITY.items() if k in source), 0.40)
    # Blend: 60% from real data (signaux), 40% from source heuristic
    f_source_quality = (nb_immo_norm * 0.60 + f_source * 0.40) if gov_features else f_source

    ext = url.lower().split('?')[0]
    f_ext_quality = (
        1.0 if any(ext.endswith(e) for e in ['.jpg', '.jpeg']) else
        0.8 if ext.endswith('.png')  else
        0.6 if ext.endswith('.webp') else
        0.3 if url.startswith('http') else 0.0
    )

    filled = sum(1 for c in ['prix', 'surface_m2', 'gouvernorat', 'description',
                              'image_url', 'date_publication', 'type_bien']
                 if not pd.isna(row.get(c)))
    f_completeness = round(filled / 7.0, 4)

    url_lower = url.lower()
    f_resolution_proxy = (
        1.0 if any(x in url_lower for x in ['large', 'hd', 'high', '1200', '800', 'full']) else
        0.6 if any(x in url_lower for x in ['medium', 'med', '600', '400']) else
        0.3 if any(x in url_lower for x in ['thumb', 'small', 'tiny', '150', '200']) else
        0.5 if url.startswith('http') else 0.0
    )

    f_multi_image = (
        1.0 if re.search(r'[_\-/](\d{1,2})[_\-/\.]', url) else
        0.5 if url.startswith('http') else 0.0
    )

    f_price_ok   = 1.0 if (not pd.isna(prix) and not pd.isna(surf)) else (0.5 if not pd.isna(prix) else 0.0)
    f_market_liq = round(float(liq) if not pd.isna(liq) else 0.3, 4)

    score = (
        f_url_valid        * 0.20 +
        f_source_quality   * 0.20 +
        f_ext_quality      * 0.10 +
        f_completeness     * 0.20 +
        f_resolution_proxy * 0.10 +
        f_multi_image      * 0.05 +
        f_price_ok         * 0.10 +
        f_market_liq       * 0.05
    )

    return {
        'img_f1_url_valid':        round(f_url_valid, 4),
        'img_f2_source_quality':   round(f_source_quality, 4),
        'img_f3_ext_quality':      round(f_ext_quality, 4),
        'img_f4_completeness':     round(f_completeness, 4),
        'img_f5_resolution_proxy': round(f_resolution_proxy, 4),
        'img_f6_multi_image':      round(f_multi_image, 4),
        'img_f7_price_ok':         round(f_price_ok, 4),
        'img_f8_market_liq':       round(f_market_liq, 4),
        'image_quality_score':     round(float(np.clip(score, 0.0, 1.0)), 4),
    }


def run_vision_features(df: pd.DataFrame, gov_features: dict) -> tuple[pd.DataFrame, list[str]]:
    section("ETAPE 9 — VISION : IMAGE EMBEDDING & QUALITY SCORE")
    log("Calcul image features (vecteur 8-dim + score)...")
    img_list = [compute_image_features(row, gov_features) for _, row in df.iterrows()]
    df_img   = pd.DataFrame(img_list, index=df.index)
    IMG_EMBED_COLS = [c for c in df_img.columns if c.startswith('img_f')]
    for col in df_img.columns:
        df[col] = df_img[col]
    log(f"  image_quality_score : mean={df['image_quality_score'].mean():.3f}")
    return df, IMG_EMBED_COLS


# ================================================================
# NLP TUNISIEN IMMOBILIER — 3 couches
#
# Couche 1 : Vocabulaire tunisien immobilier (synonymes + expressions)
#   Normalise les termes spécifiques au marché tunisien avant vectorisation
#   ex: "dar arabiya" → "maison arabe standing" (enrichi sémantiquement)
#
# Couche 2 : Scores thématiques immobiliers tunisiens
#   Détecte 8 thèmes clés qui influencent le prix en Tunisie :
#   standing, vue mer, neuf, meublé, proximité, superficie, équipements, localisation luxe
#   Chaque thème a un poids calibré sur le marché tunisien
#
# Couche 3 : Embeddings vectoriels (BERT ou LSA)
#   Représentation vectorielle de la description enrichie
#   text_embedding_score = combinaison des 3 couches
# ================================================================

EMBED_DIM = 32

# ── Couche 1 : Vocabulaire tunisien immobilier ────────────────────
# Expressions spécifiques au marché tunisien → tokens standardisés
# pour que TF-IDF/BERT les traite comme des concepts cohérents
TUN_VOCAB: dict[str, str] = {
    # Types de biens tunisiens
    's+0': 'studio',
    's+1': 'appartement une chambre',
    's+2': 'appartement deux chambres',
    's+3': 'appartement trois chambres',
    's+4': 'appartement quatre chambres',
    's+5': 'appartement cinq chambres',
    'dar arabiya': 'maison arabe standing heritage',
    'dar': 'maison',
    'villa prestige': 'villa luxe standing haut gamme',
    'standing': 'luxe haut gamme prestige',
    'rdc': 'rez de chaussee',
    'duplex': 'duplex deux niveaux',
    'triplex': 'triplex trois niveaux',
    'penthouse': 'penthouse dernier etage luxe',
    # Localisation premium Tunisie
    'pieds dans l eau': 'bord mer front mer vue mer luxe',
    'bord de mer': 'vue mer front mer littoral',
    'vue panoramique lac': 'berges du lac luxe vue lac',
    'berges du lac': 'berges lac tunis luxe standing',
    'les berges': 'berges lac luxe',
    'sidi bou said': 'sidi bou said luxe village artistique',
    'gammarth': 'gammarth zone luxe prestige',
    'la marsa': 'la marsa balneare prestige',
    'jardins de carthage': 'jardins carthage luxe residence',
    'kantaoui': 'port el kantaoui touristique mer',
    'yasmine hammamet': 'yasmine hammamet zone touristique luxe',
    'djerba': 'djerba ile touristique mer',
    # Qualificatifs standing
    'bien expose': 'orientation soleil lumineux',
    'tres lumineux': 'lumineux ensoleille',
    'entierement renove': 'renove neuf etat neuf',
    'cle en main': 'pret habiter etat neuf',
    'neuf jamais habite': 'neuf etat parfait',
    'residence fermee': 'residence securisee gardiennage standing',
    'residence securisee': 'securite gardiennage standing',
    'avec piscine': 'piscine luxe',
    'pied dans l eau': 'bord mer luxe',
    'vue sur mer': 'vue mer premium',
    # Equipements
    'double vitrage': 'isolation qualite construction',
    'climatisation centralisee': 'climatisation luxe',
    'domotique': 'domotique smart home luxe',
    'ascenseur': 'ascenseur immeuble standing',
    'parking couvert': 'parking garage securise',
    'cave': 'cave stockage',
    'terrasse panoramique': 'terrasse vue panoramique luxe',
    # Marché
    'prix negociable': 'prix flexible negociation',
    'prix ferme': 'prix fixe pas negociation',
    'urgent': 'vente urgente prix bas',
    'investissement': 'investissement rentabilite',
    'rendement locatif': 'rentabilite location investissement',
    # Arabe translittéré fréquent dans annonces tunisiennes
    'haouma': 'quartier voisinage',
    'houma': 'quartier',
    'dar el baraka': 'maison benediction standing',
    'cite el wafa': 'cite el wafa quartier',
    'cite ennasr': 'cite ennasr quartier',
}

# ── Couche 2 : Scores thématiques tunisiens ───────────────────────
# 8 thèmes calibrés sur le marché immobilier tunisien
# Chaque thème a des mots-clés et un poids sur le prix estimé

TUN_THEMES: dict[str, dict] = {
    'standing_luxe': {
        'keywords': ['luxe', 'standing', 'prestige', 'haut gamme', 'premium',
                     'somptueux', 'exceptionnel', 'rare', 'unique', 'exclusif',
                     'dar arabiya', 'villa prestige', 'penthouse', 'grand standing',
                     'haut standing', 'tres standing', 'standing eleve',
                     'propriete de luxe', 'villa luxueuse', 'residence prestige'],
        'weight': 0.25,
    },
    'vue_mer_eau': {
        # Vue directe uniquement — "lac" seul exclu (= quartier Lac 2 Tunis, pas vue)
        # "plage", "proche mer" exclus car trop génériques (30% false positive)
        'keywords': ['vue mer', 'vue sur mer', 'vue sur la mer',
                     'bord de mer', 'bord mer', 'front de mer',
                     'pieds dans l eau', 'pied dans l eau',
                     'face a la mer', 'face mer', 'acces direct mer',
                     'a 100m de la mer', 'a 50m de la mer',
                     'berges du lac', 'bord du lac', 'vue sur le lac',
                     'vue panoramique sur', 'vue imprenable sur'],
        'weight': 0.20,
    },
    'neuf_renove': {
        # Exclut "finition" seul — trop générique
        'keywords': ['neuf', 'nouvelle construction', 'jamais habite', 'jamais habité',
                     'livraison', 'cle en main', 'clé en main', 'pret a habiter',
                     'entierement renove', 'entièrement rénové', 'refait a neuf',
                     'refait neuf', 'remis a neuf', 'double vitrage',
                     'finition haut gamme', 'construction recente'],
        'weight': 0.15,
    },
    'equipements': {
        # Équipements premium uniquement — terrasse/balcon/jardin génériques exclus
        'keywords': ['piscine', 'piscine privee', 'piscine chauffee',
                     'ascenseur', 'domotique',
                     'garage ferme', 'parking couvert', 'parking prive',
                     'terrasse panoramique', 'grande terrasse', 'roof top',
                     'jardin privatif', 'suite parentale', 'dressing',
                     'cuisine equipee', 'cuisine équipée', 'cuisine amenagee'],
        'weight': 0.15,
    },
    'securite_residence': {
        # "securise" seul exclu — trop générique
        'keywords': ['residence fermee', 'résidence fermée',
                     'residence securisee', 'résidence sécurisée',
                     'gardiennage', 'gardien 24h', 'interphone video',
                     'badge acces', 'portail automatique', 'camera surveillance',
                     'vigile', 'surveillance 24h'],
        'weight': 0.10,
    },
    'localisation_premium': {
        # Zones premium confirmées — ennasr/carthage retirés (trop génériques)
        # la marsa conservé car vraiment premium côtier
        'keywords': ['sidi bou said', 'gammarth', 'la marsa',
                     'jardins de carthage', 'berges du lac', 'les berges du lac',
                     'port el kantaoui', 'yasmine hammamet',
                     'cite diplomatique', 'mutuelleville', 'monfleury'],
        'weight': 0.10,
    },
    'meuble_equipe': {
        # Meublé complet uniquement — "equipe" seul exclu (trop générique)
        'keywords': ['meuble', 'meublé', 'tout equipe', 'tout équipé',
                     'furnished', 'electromenager', 'électroménager',
                     'cuisine equipee', 'cuisine équipée',
                     'entierement meuble', 'entièrement meublé'],
        'weight': 0.03,
    },
    'investissement': {
        'keywords': ['investissement', 'rentabilité', 'rendement', 'locatif',
                     'revenu', 'touriste', 'airbnb', 'saisonnier'],
        'weight': 0.02,
    },
}


def _remove_accents(text: str) -> str:
    """Supprime les accents pour normalisation robuste des textes scrapés."""
    import unicodedata
    return ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')


def normalize_tunisian_text(text: str) -> str:
    """
    Couche 1 : Normalise le vocabulaire tunisien immobilier.
    - Supprime les accents (textes scrapés souvent sans accents)
    - Remplace les expressions spécifiques par leurs équivalents enrichis
    - Nettoie la ponctuation parasite
    """
    if not text: return text
    t = text.lower().strip()
    # Normalise les accents → version sans accent pour matching robuste
    t_no_accent = _remove_accents(t)
    # Build vocab without accents too for matching
    result = t_no_accent
    for expr, replacement in sorted(TUN_VOCAB.items(), key=lambda x: len(x[0]), reverse=True):
        expr_no_acc = _remove_accents(expr.lower())
        result = result.replace(expr_no_acc, _remove_accents(replacement))
    return result


def compute_tunisian_theme_score(text: str) -> float:
    """
    Couche 2 : Score thématique tunisien immobilier.
    Détecte la présence de chaque thème et retourne un score pondéré.
    Fonctionne avec ou sans accents (textes scrapés).
    """
    if not text: return 0.0
    t = _remove_accents(text.lower())  # normalise accents pour matching robuste
    total_score = 0.0
    themes_detected = []
    for theme, cfg in TUN_THEMES.items():
        hits = sum(1 for kw in cfg['keywords'] if kw in t)
        if hits > 0:
            # Score par thème : présent=1, renforcé si plusieurs mots-clés
            theme_score = min(hits / max(len(cfg['keywords']) * 0.3, 1), 1.0)
            total_score += theme_score * cfg['weight']
            themes_detected.append(theme)
    return round(min(total_score, 1.0), 4)


def build_rich_text(row) -> str:
    """
    Construit le texte enrichi pour embedding.
    Applique la normalisation tunisienne (Couche 1) + enrichit avec les métadonnées.
    """
    parts: list[str] = []

    # Description principale — normalisée avec le vocabulaire tunisien
    desc = str(row.get('description', '') or '')
    if desc:
        parts.append(normalize_tunisian_text(desc))

    # Type de bien
    tb_code  = row.get('type_bien')
    tb_label = TYPE_BIEN_DEC.get(int(tb_code), 'bien') if not pd.isna(tb_code) else 'bien'
    parts.append(f"type {tb_label}")

    # Localisation (enrichie aussi avec vocabulaire tunisien)
    if pd.notna(row.get('ville')):
        ville_norm = normalize_tunisian_text(str(row['ville']))
        parts.append(f"ville {ville_norm}")

    # Surface
    if pd.notna(row.get('surface_m2')):
        s = float(row['surface_m2'])
        # Catégoriser la surface pour aider le modèle
        if   s < 50:  cat = 'tres petite surface studio'
        elif s < 100: cat = 'petite surface appartement'
        elif s < 200: cat = 'surface moyenne maison'
        elif s < 400: cat = 'grande surface villa'
        else:          cat = 'tres grande surface propriete'
        parts.append(f"surface {s:.0f} m2 {cat}")

    # Features NLP binaires déjà calculées
    nlp_tokens = [feat.replace('nlp_', '') for feat in NLP_KW if row.get(feat, False)]
    if nlp_tokens:
        parts.append(' '.join(nlp_tokens))

    # Type de transaction
    tt = row.get('type_transaction')
    if not pd.isna(tt):
        parts.append('location' if int(tt) == 1 else 'vente')

    return ' '.join(parts) if parts else 'annonce immobilier tunisie'


def embedding_to_score(embed_vector, row) -> float:
    """
    Score final combinant les 3 couches NLP.

    Couche 1+2 = score thématique tunisien (depuis le texte brut)
    Couche 3   = score vectoriel (depuis l'embedding)
    """
    desc = str(row.get('description', '') or '')

    # Couche 3 : score vectoriel (norme embedding normalisée)
    norm_score = float(np.linalg.norm(embed_vector))
    vec_score  = min(norm_score, 1.0)

    # Couche 2 : score thématique tunisien
    theme_score = compute_tunisian_theme_score(desc)

    # Richesse de description
    len_score = min(len(desc) / 500.0, 1.0)

    # Features NLP binaires
    nlp_score = min(sum(1 for f in NLP_KW if row.get(f, False)) / len(NLP_KW), 1.0)

    # Diversité lexicale
    words = desc.lower().split()
    lex_div = (len(set(words)) / len(words)) if len(words) > 5 else 0.3

    # Combinaison pondérée
    # theme_score pèse plus car calibré sur le marché tunisien
    score = (
        theme_score * 0.35 +   # thèmes tunisiens (standing, vue mer, luxe...)
        vec_score   * 0.25 +   # embedding vectoriel
        len_score   * 0.20 +   # richesse description
        nlp_score   * 0.15 +   # features NLP binaires
        lex_div     * 0.05     # diversité vocabulaire
    )
    return round(float(np.clip(score, 0.0, 1.0)), 4)


def run_text_embedding(df: pd.DataFrame) -> tuple[pd.DataFrame, np.ndarray]:
    section("ETAPE 10 — NLP TUNISIEN IMMOBILIER")
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.decomposition import TruncatedSVD
    from sklearn.preprocessing import normalize as sk_normalize

    embed_dim = EMBED_DIM

    # Calcul du score thématique tunisien (Couche 2) sur toutes les annonces
    log("  Couche 1+2 : Vocabulaire tunisien + scores thématiques...")
    theme_scores = []
    for _, row in df.iterrows():
        desc = str(row.get('description', '') or '')
        theme_scores.append(compute_tunisian_theme_score(desc))
    df['_theme_score'] = theme_scores
    log(f"  Score thématique tunisien : mean={sum(theme_scores)/len(theme_scores):.3f}")

    # Thèmes les plus détectés
    theme_counts = {t: 0 for t in TUN_THEMES}
    for _, row in df.iterrows():
        desc = str(row.get('description', '') or '').lower()
        for theme, cfg in TUN_THEMES.items():
            if any(kw in desc for kw in cfg['keywords']):
                theme_counts[theme] += 1
    log("  Thèmes détectés :")
    for t, cnt in sorted(theme_counts.items(), key=lambda x: x[1], reverse=True):
        log(f"    {t:<25} : {cnt:>6} annonces ({cnt/len(df)*100:.1f}%)")

    # Construire les textes enrichis (Couche 1 appliquée)
    log("  Couche 3 : Embeddings vectoriels...")
    texts = df.apply(build_rich_text, axis=1).tolist()
    log(f"  {len(texts)} documents préparés")

    BERT_OK = False
    try:
        # Test PyTorch avant d'importer sentence_transformers
        import torch as _torch_test
        _ = _torch_test.tensor([1.0])  # test réel de chargement
        from sentence_transformers import SentenceTransformer
        # Modèle multilingue — comprend l'arabe, le français et le tunisien
        st_model   = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
        embeddings = st_model.encode(
            texts, batch_size=64, show_progress_bar=True, convert_to_numpy=True
        )
        embed_dim  = embeddings.shape[1]
        BERT_OK    = True
        log(f"  ✔ BERT multilingue : shape={embeddings.shape}")
    except Exception as e:
        log(f"  ⚠ BERT non disponible : {type(e).__name__}: {str(e)[:120]}")
        log(f"  → Solution : pip uninstall torch -y && pip install torch==2.1.2 --index-url https://download.pytorch.org/whl/cpu")
        log(f"  → Fallback LSA activé (qualité inférieure)")
        # TF-IDF avec vocabulaire immobilier tunisien spécifique
        vectorizer = TfidfVectorizer(
            max_features=10000,
            ngram_range=(1, 3),      # trigrams pour capturer "vue sur mer", "bord de mer"
            sublinear_tf=True,
            min_df=2,
            strip_accents='unicode',
            analyzer='word',
            token_pattern=r'\b[a-zA-ZÀ-ÿ]{2,}\b',
        )
        tfidf_matrix = vectorizer.fit_transform(texts)
        n_components = min(embed_dim, tfidf_matrix.shape[1] - 1, tfidf_matrix.shape[0] - 1)
        svd          = TruncatedSVD(n_components=n_components, random_state=42, n_iter=10)
        embeddings   = sk_normalize(svd.fit_transform(tfidf_matrix), norm='l2')
        embed_dim    = embeddings.shape[1]
        log(f"  ✔ LSA embeddings : shape={embeddings.shape}")

    # embed_cols défini ICI — après try/except, quelle que soit la branche
    embed_cols = [f'bert_dim_{i:02d}' for i in range(embed_dim)]
    df_embed   = pd.DataFrame(embeddings, index=df.index, columns=embed_cols)
    for col in embed_cols:
        df[col] = df_embed[col]

    scores = [embedding_to_score(embeddings[i], df.iloc[i]) for i in range(len(df))]
    df['text_embedding_score'] = scores
    log(f"  text_embedding_score : mean={df['text_embedding_score'].mean():.3f}")

    try:
        np.save('text_embeddings.npy', embeddings)
    except Exception as e:
        log(f"  [WARN] Sauvegarde .npy : {e}")

    return df, embeddings


def run_image_embedding_save(df: pd.DataFrame, img_embed_cols: list[str]) -> None:
    try:
        np.save('image_embeddings.npy', df[img_embed_cols].values)
    except Exception as e:
        log(f"  [WARN] Sauvegarde image_embeddings.npy : {e}")


# ================================================================
# MULTIMODAL FUSION
# ================================================================

MODAL_WEIGHTS = {'struct': 0.50, 'text': 0.30, 'vision': 0.20}
_STRUCT_FIELDS = ['prix', 'surface_m2', 'gouvernorat', 'type_bien', 'lat', 'lon',
                  'market_tension', 'cycle_marche', 'date_publication']


def compute_modal_struct_score(row) -> float:
    filled = sum(1 for f in _STRUCT_FIELDS if not pd.isna(row.get(f)))
    return round(filled / len(_STRUCT_FIELDS), 4)


def run_multimodal_fusion(df: pd.DataFrame) -> pd.DataFrame:
    section("ETAPE 10b — FUSION MULTIMODALE")
    df['modal_struct_score'] = df.apply(compute_modal_struct_score, axis=1)
    df['multimodal_score'] = (
        df['modal_struct_score']   * MODAL_WEIGHTS['struct'] +
        df['text_embedding_score'] * MODAL_WEIGHTS['text']   +
        df['image_quality_score']  * MODAL_WEIGHTS['vision']
    ).round(4)
    log(f"  multimodal_score : mean={df['multimodal_score'].mean():.3f}")
    return df


# ================================================================
# TARGET — prix_transaction_estimated
# Uses real nego_rates from INS immobilier.
# df is rebuilt from clean groupes_clean at this point → all rows have prix.
# ================================================================

CURRENT_YEAR = 2026
MIN_YEAR     = 2022


def compute_target(df: pd.DataFrame, nego_rates: dict) -> pd.DataFrame:
    """
    Estimates prix_transaction_estimated using real INS-derived nego rates.
    At this point df contains only clean rows (prix NaN already dropped in step 5).
    Every row should have a valid prix → prix_transaction_estimated computed for all.
    """
    section("ETAPE 11 — PRIX TRANSACTION ESTIME (TARGET)")

    # ── Segment-specific pressure curves ────────────────────────────
    # Each segment has its own time-pressure profile derived from
    # observed Tunisian market behavior:
    #
    #  Location residentiel : renters move fast → pressure builds quickly
    #    after 4 weeks, landlord typically drops price
    #  Vente appartement    : moderate pressure, buyers have alternatives
    #  Vente villa/maison   : sellers less pressured (emotional attachment)
    #  Terrain/Ferme        : very illiquid → high pressure after 3 months
    #  Commercial           : moderate, depends heavily on location
    #
    # Formula: base_rate + time_pressure(days, segment) + market_adjustments
    #
    # time_pressure uses a logarithmic curve (not linear):
    #   pressure = slope * log(1 + days/halflife)
    # This reflects real behavior: first weeks matter most,
    # then diminishing returns as stubborn sellers self-select.

    # Segment pressure profiles: (slope, halflife_days, max_pressure)
    PRESSURE_PROFILES = {
        # (tt, type_cat) : (slope, halflife, max_add)
        (1, 'Appartement'):  (0.045, 21,  0.09),  # location appart: fast pressure
        (1, 'Chambre'):      (0.050, 14,  0.10),  # location chambre: very fast
        (1, 'Maison'):       (0.040, 28,  0.08),  # location maison: moderate
        (1, 'Villa'):        (0.035, 35,  0.07),  # location villa: slow
        (2, 'Appartement'):  (0.030, 45,  0.08),  # vente appart: moderate
        (2, 'Maison'):       (0.025, 60,  0.06),  # vente maison: slower
        (2, 'Villa'):        (0.020, 75,  0.05),  # vente villa: slowest (emotional)
        (2, 'Terrain'):      (0.040, 30,  0.12),  # vente terrain: high illiquidity
        (2, 'Ferme'):        (0.045, 30,  0.13),  # vente ferme: very illiquid
        (2, 'Local Commercial'): (0.030, 45, 0.10),
        (2, 'Bureau'):       (0.030, 45,  0.10),
        (2, 'Chambre'):      (0.030, 45,  0.08),
    }
    _DEFAULT_PRESSURE = (0.030, 45, 0.08)

    def estimate_negotiation_rate(row) -> float:
        type_cat = str(row.get('type_categorise', '') or '').strip()
        tt       = int(row.get('type_transaction', 2) or 2)
        base     = nego_rates.get(type_cat, nego_rates.get('Autre', 0.05))
        days     = float(row.get('temp_days_on_market', 30) or 30)
        liq      = float(row.get('market_liquidity_score', 0.5) or 0.5)
        tension  = float(row.get('market_rental_tension',  0.3) or 0.3)
        cycle    = int(row.get('cycle_marche', 0) or 0)
        season   = int(row.get('temp_high_season', 0) or 0)
        attract  = float(row.get('score_attractivite', 0.5) or 0.5)
        cycle_label = CYCLE_MARCHE_DEC.get(cycle, 'Unknown')

        # Segment-specific time pressure (logarithmic curve)
        slope, halflife, max_add = PRESSURE_PROFILES.get(
            (tt, type_cat), _DEFAULT_PRESSURE
        )
        time_pressure = min(slope * np.log1p(days / halflife), max_add)

        # Market adjustments
        liquidity_adj  = (1.0 - liq) * 0.04          # illiquid market → more negotiation
        cycle_adj      = 0.02 if cycle_label in ('stabilization', 'recovery', 'decline') else 0.0
        tension_adj    = -tension * 0.025              # high rental demand → less negotiation
        season_adj     = -0.01 if season == 1 else 0.0
        attract_adj    = -(attract - 0.5) * 0.02       # attractive zone → less discount

        rate = base + time_pressure + liquidity_adj + cycle_adj + tension_adj + season_adj + attract_adj

        # Clip: minimum 3% (always some negotiation), maximum varies by segment
        max_rate = 0.20 if type_cat in ('Terrain', 'Ferme') else 0.15
        return round(float(np.clip(rate, 0.03, max_rate)), 4)

    df['negotiation_rate'] = df.apply(estimate_negotiation_rate, axis=1)

    # Prix transaction = prix * (1 - nego_rate)
    # df is rebuilt from clean data — prix NaN should be 0 here
    df['prix_transaction_estimated'] = (df['prix'] * (1 - df['negotiation_rate'])).round(2)

    n_computed = df['prix_transaction_estimated'].notna().sum()
    n_nan      = df['prix_transaction_estimated'].isna().sum()
    if n_nan > 0:
        log(f"  [WARN] {n_nan} prix_transaction_estimated NaN — prix source toujours absent")
    log(f"  prix_transaction_estimated : {n_computed:,} calculés | {n_nan} NaN")
    log(f"  Taux négociation moyen : {df['negotiation_rate'].mean()*100:.1f}%")

    df = df.drop(columns=['negotiation_rate'], errors='ignore')
    df['type_transaction'] = pd.to_numeric(df['type_transaction'], errors='coerce').fillna(1).astype(int)

    # date_publication → integer year
    def extract_year(val) -> int:
        if pd.isna(val): return CURRENT_YEAR
        s = str(val).strip().lower()
        if any(kw in s for kw in ['ago', 'hour', 'minute', 'day', 'week', 'month', 'year', 'just now', 'just']):
            m_n = re.search(r'(\d+)', s)
            n   = int(m_n.group(1)) if m_n else 1
            if any(kw in s for kw in ['minute', 'hour', 'day', 'week']): return CURRENT_YEAR
            if 'month' in s: return CURRENT_YEAR if n <= 6 else CURRENT_YEAR - 1
            if 'year'  in s: return max(CURRENT_YEAR - n, MIN_YEAR)
            return CURRENT_YEAR
        try:
            dt = pd.to_datetime(val, errors='coerce')
            if pd.isna(dt): return CURRENT_YEAR
            y = dt.year
            return y if MIN_YEAR <= y <= CURRENT_YEAR else MIN_YEAR
        except: return CURRENT_YEAR

    df['date_publication'] = df['date_publication'].apply(extract_year)
    log(f"  date_publication : {df['date_publication'].value_counts().sort_index().to_dict()}")
    return df




# ================================================================
# SAMPLE WEIGHTS — correction déséquilibres géographique + temporel
#
# Pourquoi sample_weight ?
#   - Tunis = 40% des annonces → le modèle biais vers Tunis
#   - Béja/Gafsa/Kasserine = <30 ann. → le modèle les ignore
#   - Peak 2025-2026 = 90% → le modèle ne voit quasiment pas stabilization/decline
#
# Solution : weight = geo_weight × temporal_weight
#   - Annonce sous-représentée → weight > 1 (le modèle y prête plus attention)
#   - Annonce sur-représentée  → weight < 1 (le modèle la normalise)
#
# Utilisation dans XGBoost/RF :
#   model.fit(X_train, y_train, sample_weight=df['sample_weight'])
# ================================================================

def compute_sample_weights(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcule sample_weight + prix_region_median — stratégie en 3 niveaux.

    NIVEAU A (≥100 ann.) : données fiables → poids normal, prix propre du gouvernorat
    NIVEAU B (30-99 ann.): données limitées → poids ×1.5, emprunt 30% région
    NIVEAU C (<30 ann.)  : données insuffisantes → poids plafonné à 2.0, emprunt 70% région

    RÉGIONS TUNISIENNES (groupées par marché similaire) :
      Grand Tunis  : Tunis, Ariana, Ben Arous, Manouba
      Nord-Est     : Nabeul, Bizerte, Béja, Jendouba
      Centre-Est   : Sousse, Monastir, Mahdia, Sfax, Kairouan
      Sud          : Médenine, Tataouine, Tozeur
      Centre-Ouest : Gafsa, Kasserine, Kébili, Sidi Bouzid, Siliana
      Nord-Ouest   : Le Kef, Zaghouan

    prix_region_median = feature ML qui donne au modèle le contexte de prix régional.
    Pour Gafsa (1 annonce à 1,148,850 TND aberrant) → remplacé par 70% médiane Centre-Ouest.
    """
    import mappings as _maps

    n = len(df)

    # ── Régions géographiques ─────────────────────────────────────
    REGIONS = {
        'Grand Tunis':   [1, 3, 13, 23],
        'Nord-Est':      [2, 4, 7, 16],
        'Centre-Est':    [8, 12, 15, 17, 20],
        'Sud':           [14, 21, 22],
        'Centre-Ouest':  [6, 9, 10, 18, 19],
        'Nord-Ouest':    [11, 24],
    }
    gov_to_region = {c: r for r, codes in REGIONS.items() for c in codes}

    # Médiane par région (robuste — basée sur toutes les annonces de la région)
    global_median = float(df['prix_transaction_estimated'].median())
    region_medians = {}
    for region, codes in REGIONS.items():
        data = df[df['gouvernorat'].isin(codes)]['prix_transaction_estimated']
        region_medians[region] = float(data.median()) if len(data) >= 5 else global_median

    # Médiane par gouvernorat
    gov_medians = df.groupby('gouvernorat')['prix_transaction_estimated'].median().to_dict()
    gov_counts  = df['gouvernorat'].value_counts()

    # ── Feature : prix_region_median (emprunt hiérarchique) ──────
    def compute_prix_region(gov_code):
        cnt     = gov_counts.get(gov_code, 0)
        gov_med = gov_medians.get(gov_code, np.nan)
        region  = gov_to_region.get(gov_code, 'Centre-Est')
        reg_med = region_medians.get(region, global_median)
        if pd.isna(gov_med): gov_med = reg_med
        if cnt >= 100: return gov_med                          # Niveau A
        elif cnt >= 30: return 0.70 * gov_med + 0.30 * reg_med  # Niveau B
        else:           return 0.30 * gov_med + 0.70 * reg_med  # Niveau C

    df['prix_region_median'] = df['gouvernorat'].apply(compute_prix_region).round(2)

    # ── 1. Poids géographique adaptatif ─────────────────────────
    n_govs    = df['gouvernorat'].nunique()
    ideal_pct = 1.0 / max(n_govs, 1)
    geo_weights = np.ones(n)
    for gov_code, cnt in gov_counts.items():
        actual_pct = cnt / n
        raw_w = ideal_pct / actual_pct
        if cnt >= 100:  w = np.clip(raw_w, 0.2, 3.0)       # Niveau A
        elif cnt >= 30: w = np.clip(raw_w * 1.5, 0.5, 4.0) # Niveau B
        else:           w = 2.0                              # Niveau C — cap strict
        mask = df['gouvernorat'] == gov_code
        geo_weights[mask.values] = w

    # ── 2. Poids temporel ─────────────────────────────────────────
    cycle_counts = df['cycle_marche'].value_counts()
    n_cycles     = df['cycle_marche'].nunique()
    ideal_cycle  = 1.0 / max(n_cycles, 1)
    temp_weights = np.ones(n)
    for cycle_code, cnt in cycle_counts.items():
        actual_pct = cnt / n
        w = np.clip(ideal_cycle / actual_pct, 0.3, 4.0)
        mask = df['cycle_marche'] == cycle_code
        temp_weights[mask.values] = w

    # ── Combinaison et normalisation ──────────────────────────────
    combined = geo_weights * temp_weights
    combined = combined / combined.mean()
    combined = np.clip(combined, 0.1, 6.0)
    combined = combined / combined.mean()

    df['sample_weight'] = np.round(combined, 4)

    # ── Log résumé ───────────────────────────────────────────────
    log(f"  sample_weight    : mean={df['sample_weight'].mean():.3f} | "
        f"min={df['sample_weight'].min():.3f} | max={df['sample_weight'].max():.3f}")
    log(f"  prix_region_median : mean={df['prix_region_median'].mean():,.0f} TND")
    n_a = sum(1 for c in gov_counts.index if gov_counts[c] >= 100)
    n_b = sum(1 for c in gov_counts.index if 30 <= gov_counts[c] < 100)
    n_c = sum(1 for c in gov_counts.index if gov_counts[c] < 30)
    log(f"  Niveau A (>=100): {n_a} gouvernorats → données propres")
    log(f"  Niveau B (30-99): {n_b} gouvernorats → emprunt 30% région")
    log(f"  Niveau C (<30)  : {n_c} gouvernorats → emprunt 70% région")
    log("  Poids par gouvernorat (sous-représentés) :")
    gov_w = df.groupby('gouvernorat')['sample_weight'].mean().sort_values(ascending=False)
    for gov_code, w in gov_w.head(6).items():
        gov_name = _maps.GOUVERNORAT_DEC.get(int(gov_code), '?')
        cnt = gov_counts.get(gov_code, 0)
        level = 'A' if cnt >= 100 else ('B' if cnt >= 30 else 'C')
        log(f"    [{level}] {gov_name:<15} : weight={w:.2f} ({cnt} ann.)")
    log("  Poids par cycle_marche :")
    cycle_dec = {1:'stabilization', 2:'growth', 3:'peak', 4:'recovery', 5:'decline'}
    for code, w in df.groupby('cycle_marche')['sample_weight'].mean().items():
        name = cycle_dec.get(int(code), '?')
        cnt  = cycle_counts.get(code, 0)
        log(f"    {name:<15} : weight={w:.2f} ({cnt} ann., {cnt/n*100:.1f}%)")
    return df

# ================================================================
# SAVE ENCODING MAPPINGS
# ================================================================

def save_encoding_mappings(ville_rank_map: dict = None, n_top: int = 0) -> None:
    import mappings as _maps_mod
    # ville_rank_map : {ville_key: code}  ex: {'tunis': 23001, 'la marsa': 23002, ...}
    # code format    : gouvernorat * 1000 + rang_interne
    #   23001 = Tunis gouvernorat (23), rang 1 (plus fréquente)
    #   23000 = ville rare dans Tunis
    #   0     = gouvernorat inconnu

    # Build ville decoder: {code: description lisible}
    ville_dec = {}
    if ville_rank_map:
        # Villes connues : inverser le dict
        for ville, code in ville_rank_map.items():
            gov_code = code // 1000
            rang     = code % 1000
            gov_name = _maps_mod.GOUVERNORAT_DEC.get(gov_code, str(gov_code))
            if rang > 0:
                ville_dec[str(code)] = f"{ville} [{gov_name}]"
            else:
                ville_dec[str(code)] = f"ville_rare_{gov_name}"
        # Villes rares par gouvernorat (GOV*1000)
        for gov_code, gov_name in _maps_mod.GOUVERNORAT_DEC.items():
            if gov_code > 0:
                rare_code = gov_code * 1000
                if str(rare_code) not in ville_dec:
                    ville_dec[str(rare_code)] = f"ville_rare_{gov_name}"
        # Inconnu
        ville_dec['0'] = 'ville_inconnue'

    enc_data = {
        'gouvernorat':      {str(v): k for k, v in _maps_mod.GOUVERNORAT_ENC.items()},
        'type_bien':        {str(v): k for k, v in _maps_mod.TYPE_BIEN_ENC.items()},
        'type_transaction': {'1': 'Location', '2': 'Vente'},
        'cycle_marche':     {str(v): k for k, v in _maps_mod.CYCLE_MARCHE_ENC.items()},
        'ville_encoded':    ville_dec,
    }
    try:
        with open('encoding_mappings.json', 'w', encoding='utf-8') as f:
            json.dump(enc_data, f, ensure_ascii=False, indent=2)
        log(f"  Mappings sauvegardés : encoding_mappings.json")
        log(f"  ville_encoded : {len(ville_dec)} codes documentés")
    except Exception as e:
        log(f"  [WARN] encoding_mappings.json : {e}")


# ================================================================
# FINAL EXPORT
# ================================================================

# ── 14 colonnes pour l'objectif de valuation multi-modale ──
#
# Features géographiques (4)  : gouvernorat, ville_encoded, lat, lon
#   → XGBoost/RF : encodage direct
#   → GNN : lat/lon pour construction du graphe spatial (voisinage géographique)
#
# Features bien (4)           : type_bien, surface_m2, nb_pieces, type_transaction
#   → Features structurelles fondamentales pour tout modèle de pricing
#
# Features marché (3)         : score_attractivite, market_tension, cycle_marche
#   → Contexte économique réel (BCT + INS + Google Maps)
#
# Features multimodales (2)   : text_embedding_score, image_url
#   → text_embedding_score : score NLP depuis description (XGBoost/RF/Bayésien)
#   → image_url            : URL image réelle pour CNN/CLIP (analyse visuelle directe)
#     Note : CNN/CLIP télécharge et analyse l'image depuis cette URL.
#     image_quality_score est un proxy fallback — pas un substitut à la vraie image.
#
# Target unique (1)           : prix_transaction_estimated
#   → Prix réel estimé = prix_affiché × (1 - taux_négociation_INS)
#   → C'est ce que le modèle doit apprendre à prédire
#   → prix (affiché) n'est PAS le target — c'est le prix vendeur, pas le prix marché

TARGET_COLS = [
    # ── Géographie — XGBoost/RF + GNN ───────────────────────────
    'gouvernorat',             # code 1-24
    'ville_encoded',           # code gov*1000+rang (hiérarchique)
    'lat',                     # latitude GPS
    'lon',                     # longitude GPS
    # ── Bien — XGBoost/RF + GNN ─────────────────────────────────
    'type_bien',               # Appartement=1 … Villa=9
    'surface_m2',              # m²
    'nb_pieces',               # résidentiel seulement
    'type_transaction',        # Location=1 / Vente=2
    # ── Marché — XGBoost/RF + GNN ───────────────────────────────
    'score_attractivite',      # Google Maps signaux 0-1
    'market_tension',          # tension locative 0-1
    'cycle_marche',            # cycle BCT+INS
    'prix_region_median',      # contexte régional (zones sous-représentées)
    # ── NLP — XGBoost/RF ────────────────────────────────────────
    'text_embedding_score',    # score BERT/LSA 0-1
    # ── Vision — CNN/CLIP uniquement (exclure de X XGBoost) ─────
    'image_url',               # URL image
    # ── Target ──────────────────────────────────────────────────
    'prix_transaction_estimated',
    # ── Poids entraînement (exclure de X, passer en sample_weight=) ─
    'sample_weight',
]
TARGET_COLS_NO_PIECES = [c for c in TARGET_COLS if c != 'nb_pieces']
GROUPES_SANS_PIECES   = ('Foncier', 'Commercial', 'Divers')

FICHIERS_ML = {
    'Residentiel': ('residentiel_BO2.xlsx', '1565C0'),
    'Foncier':     ('foncier_BO2.xlsx',     '2E7D32'),
    'Commercial':  ('commercial_BO2.xlsx',  'E65100'),
    'Divers':      ('divers_BO2.xlsx',      '880E4F'),
}

PROPAGATE_COLS = [
    'lat', 'lon', 'market_tension', 'market_liquidity_score', 'market_rental_tension',
    'score_attractivite', 'population_gouvernorat', 'nb_immo_direct',
    'cycle_marche', 'taux_directeur', 'temp_days_on_market', 'temp_high_season',
    'prix', 'type_transaction', 'date_publication',
    'text_embedding_score', 'image_quality_score',
    'multimodal_score', 'prix_transaction_estimated', 'nb_pieces',
    'gouvernorat', 'ville_encoded', 'type_bien', 'prix_region_median',
]


def write_excel(df_out: pd.DataFrame, filename: str, color_hex: str, sheet_name: str) -> None:
    wb = Workbook(); ws = wb.active; ws.title = sheet_name
    hfill = PatternFill("solid", start_color=color_hex)
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    for r in dataframe_to_rows(df_out, index=False, header=True):
        clean = []
        for cell in r:
            if isinstance(cell, str):
                cell = cell.encode('ascii', 'ignore').decode('ascii')
                cell = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cell)
                cell = re.sub(r'\s+', ' ', cell).strip()
            elif isinstance(cell, bool):        cell = str(cell)
            elif isinstance(cell, np.integer):  cell = int(cell)
            elif isinstance(cell, np.floating): cell = round(float(cell), 4) if not np.isnan(cell) else None
            elif hasattr(cell, 'item'):          cell = cell.item()  # numpy scalar -> Python natif
            clean.append(cell)
        ws.append(clean)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        w = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
    ws.freeze_panes = "A2"
    wb.save(filename)
    log(f"  {filename:<38}: {len(df_out):>6} lignes | {len(df_out.columns)} colonnes")


def resegment_and_export(
    df: pd.DataFrame,
    groupes_clean: dict[str, pd.DataFrame],
    img_embed_cols: list[str],
) -> dict[str, pd.DataFrame]:
    section("ETAPE 12 — EXPORT ML — COLONNES FINALES")

    df_all_groups = pd.concat(groupes_clean.values(), ignore_index=False)
    moves = []
    for idx, row in df_all_groups.iterrows():
        tb_code = row.get('type_bien', 0)
        tb      = TYPE_BIEN_DEC.get(int(tb_code) if not pd.isna(tb_code) else 0, 'Unknown')
        current = next((g for g, dg in groupes_clean.items() if idx in dg.index), None)
        target  = RESEGMENT_MAP.get(tb, 'Divers' if tb in ('Unknown', 'Autre', '') else None)
        if target and current and target != current:
            moves.append((idx, current, target))

    for idx, frm, to in moves:
        groupes_clean[frm] = groupes_clean[frm].drop(index=idx, errors='ignore')
        row_data = df_all_groups.loc[[idx]]
        if to not in groupes_clean: groupes_clean[to] = row_data
        else: groupes_clean[to] = pd.concat([groupes_clean[to], row_data])
    log(f"  Re-segmentation : {len(moves)} lignes déplacées")

    df = df.drop(columns=['_gouvernorat_str'], errors='ignore')
    for groupe in groupes_clean:
        groupes_clean[groupe] = groupes_clean[groupe].drop(columns=['_gouvernorat_str'], errors='ignore')

    # Ensure date_publication is int in groupes_clean to match df
    CURRENT_YEAR = datetime.now().year
    MIN_YEAR = 2020
    def extract_year(val) -> int:
        if pd.isna(val): return CURRENT_YEAR
        s = str(val).strip().lower()
        if any(kw in s for kw in ['ago', 'hour', 'minute', 'day', 'week', 'month', 'year', 'just now', 'just']):
            m_n = re.search(r'(\d+)', s)
            n   = int(m_n.group(1)) if m_n else 1
            if any(kw in s for kw in ['minute', 'hour', 'day', 'week']): return CURRENT_YEAR
            if 'month' in s: return CURRENT_YEAR if n <= 6 else CURRENT_YEAR - 1
            if 'year'  in s: return max(CURRENT_YEAR - n, MIN_YEAR)
            return CURRENT_YEAR
        try:
            dt = pd.to_datetime(val, errors='coerce')
            if pd.isna(dt): return CURRENT_YEAR
            y = dt.year
            return y if MIN_YEAR <= y <= CURRENT_YEAR else MIN_YEAR
        except: return CURRENT_YEAR

    for groupe in groupes_clean:
        if 'date_publication' in groupes_clean[groupe].columns:
            groupes_clean[groupe]['date_publication'] = groupes_clean[groupe]['date_publication'].apply(extract_year)

    all_propagate = PROPAGATE_COLS + img_embed_cols

    for groupe, dg in groupes_clean.items():
        fname, color = FICHIERS_ML[groupe]
        valid_idx = [i for i in dg.index if i in df.index]

        # Propagate ALL columns from df — overwrite existing ones too
        # (previous logic skipped columns already in dg, leaving stale NaN values)
        for col in all_propagate:
            if col in df.columns and valid_idx:
                dg.loc[valid_idx, col] = df.loc[valid_idx, col].values

        if valid_idx and 'gouvernorat' in df.columns:
            dg.loc[valid_idx, 'gouvernorat'] = df.loc[valid_idx, 'gouvernorat'].values

        # Fix residual Unknown governorates
        mask_unk = dg['gouvernorat'] == 0
        if mask_unk.sum() > 0 and 'lat' in dg.columns:
            lats_s = dg.loc[mask_unk, 'lat'].fillna(36.8).values
            lons_s = dg.loc[mask_unk, 'lon'].fillna(10.1).values
            dlat_s = (lats_s[:, None] - _gov_lats[None, :]) * _KM_PER_LAT
            dlon_s = (lons_s[:, None] - _gov_lons[None, :]) * _KM_PER_LON
            best_s = np.argmin(np.sqrt(dlat_s ** 2 + dlon_s ** 2), axis=1)
            gov_s  = [_gov_names[i] for i in best_s]
            dg.loc[mask_unk, 'gouvernorat'] = (
                pd.Series(gov_s, index=dg.loc[mask_unk].index)
                .map(GOUVERNORAT_ENC).fillna(0).astype(int))
            log(f"  [{groupe}] {mask_unk.sum()} Unknown résiduels corrigés par géo")

        # Safety re-encode if any column slipped to string
        if 'gouvernorat' in dg.columns and dg['gouvernorat'].dtype == object:
            dg['gouvernorat'] = dg['gouvernorat'].apply(lambda x: GOUVERNORAT_ENC.get(str(x).strip(), 0))
        if 'type_bien' in dg.columns and dg['type_bien'].dtype == object:
            dg['type_bien'] = dg['type_bien'].apply(encode_type_bien)
        if 'cycle_marche' in dg.columns and dg['cycle_marche'].dtype == object:
            dg['cycle_marche'] = dg['cycle_marche'].apply(encode_cycle_marche)

        if 'lat' in dg.columns:
            dg['lat'] = dg['lat'].fillna(36.8)
            dg['lon'] = dg['lon'].fillna(10.1)

        # Final safety — fix NaN AND invalid zeros per group
        # score_attractivite=0 / market_tension=0 / cycle_marche=0 are all errors
        _med_mt  = df['market_tension'].replace(0, np.nan).median()   if 'market_tension'   in df.columns else 0.5
        _med_sa  = df['score_attractivite'].replace(0, np.nan).median() if 'score_attractivite' in df.columns else 0.5
        _med_te  = df['text_embedding_score'].median()                 if 'text_embedding_score' in df.columns else 0.5
        _med_mt  = float(_med_mt)  if not pd.isna(_med_mt)  else 0.5
        _med_sa  = float(_med_sa)  if not pd.isna(_med_sa)  else 0.5
        _med_te  = float(_med_te)  if not pd.isna(_med_te)  else 0.5

        # (fill_value, also_fix_zeros)
        _checks = {
            'score_attractivite':   (_med_sa, True),
            'market_tension':       (_med_mt, True),
            'cycle_marche':         (2,       True),
            'text_embedding_score': (_med_te, False),
        }
        for _c, (_v, _fix_zero) in _checks.items():
            if _c not in dg.columns: continue
            _n_nan  = dg[_c].isna().sum()
            _n_zero = (dg[_c] == 0).sum() if _fix_zero else 0
            if _n_nan > 0:
                dg[_c] = dg[_c].fillna(_v)
            if _fix_zero and _n_zero > 0:
                dg.loc[dg[_c] == 0, _c] = _v
            if _n_nan > 0 or _n_zero > 0:
                log(f"  [{groupe}] {_c} : {_n_nan} NaN + {_n_zero} zéros → {round(_v,3)}")

        groupes_clean[groupe] = dg
        cols_sel = TARGET_COLS_NO_PIECES if groupe in GROUPES_SANS_PIECES else TARGET_COLS
        df_export = dg[[c for c in cols_sel if c in dg.columns]].copy()
        # Forcer date_publication en int Python pur (evite erreur datetime64 openpyxl)
        if 'date_publication' in df_export.columns:
            df_export['date_publication'] = (
                pd.to_numeric(df_export['date_publication'], errors='coerce')
                .fillna(2025).astype(int)
            )
        write_excel(df_export, fname, color, f"{groupe}_ML")

    return groupes_clean


def print_final_report(
    groupes_clean: dict[str, pd.DataFrame],
    n_sources: int, n0: int, n_apres_dedup: int,
    bct: dict, nego_rates: dict,
) -> None:
    section("RAPPORT FINAL")
    total = sum(len(dg) for dg in groupes_clean.values())
    log(f"Sources chargees              : {n_sources}")
    log(f"Annonces brutes               : {n0:>7,}")
    log(f"Apres deduplication           : {n_apres_dedup:>7,}  (-{n0 - n_apres_dedup})")
    log(f"Apres nettoyage complet       : {total:>7,}")
    log("")
    log(f"BCT taux directeur            : {bct.get('taux_directeur', '?')}% ({bct.get('date', '?')})")
    log(f"Taux négociation INS réels    :")
    for k, v in nego_rates.items():
        log(f"  {k:<20}: {v*100:.1f}%")
    log("")
    log("  Colonnes imputées             : surface_m2 (médiane) | nb_pieces (médiane) | type_categorise (mode)")
    log("  Nouvelles features            : ville_encoded (fréquentiel) | taux négociation courbe log par segment")
    log("  prix NaN                  : lignes supprimées (pas d'estimation possible sans prix)")
    log("Valeurs illogiques            : lignes supprimées (pas imputées)")
    for groupe, (fname, _) in FICHIERS_ML.items():
        log(f"  {fname:<38}: {len(groupes_clean.get(groupe, [])):>6} annonces")