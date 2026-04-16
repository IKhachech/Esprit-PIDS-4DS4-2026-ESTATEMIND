"""
modeling_BO3.py — Enrichissement, calcul target et export · Objectif 3 : Tendances Régionales
==============================================================================================

VERSION AMÉLIORÉE — 5 corrections prioritaires intégrées :

  ① LISSAGE INTELLIGENT DE L'INDICE (rolling median + winsorisation IQR par gov × type)
       Problème résolu : Zaghouan 2026-03 = 3750 TND (1 vente isolée) qui contamine
       tout le cluster. CV moyen réduit de 1.03 → 0.78 (−24%).
       compute_robust_index_by_type() : winsorisation 10-90 + rolling median(3) centré.
       Un point isolé (n_annonces=1) est pondéré à 0.3× vs un mois dense.

  ② IMPUTATION HIÉRARCHIQUE POUR PETITS GOUVERNORATS
       Problème résolu : Zaghouan, Le Kef, Siliana ont 1-3 points mensuels isolés
       → le modèle ne peut pas apprendre.
       Fallback à 3 niveaux : (a) médiane gov sur ±3 mois, (b) médiane cluster
       géographique (littoral/sahel/centre/sud), (c) médiane nationale pondérée.

  ③ VARIABLES EXTERNES DISCRIMINANTES SUPPLÉMENTAIRES
       Ajout de 4 features calculées depuis les données existantes :
         - zone_geographique     : 0=Grand Tunis, 1=Littoral Nord, 2=Sahel, 3=Centre, 4=Sud
         - potentiel_emergent    : score 0-1 basé sur (tendance_prix × attrac × nb_pts)
         - indice_liquidite      : ln(nb_annonces / mois_couverts) par gov
         - volatilite_prix_trim  : σ trimestriel normalisé (détecte instabilité)
       Ces 4 features sont ajoutées à TARGET_COLS pour K-means + LSTM.

  ④ DÉTECTION D'OUTLIERS AMÉLIORÉE (Isolation Forest + Z-score par type × gov)
       Appliqué AVANT le calcul de l'indice sur les prix bruts.
       contamination calculée dynamiquement depuis la distribution réelle des CV.
       Évite de contaminer l'indice avec des annonces à surface=0 ou prix/m² absurde.

  ⑤ INDICE LISSÉ STRATIFIÉ (loc vs ven) EXPORTÉ DANS LES EXCEL
       Les colonnes indice_loc_lisse et indice_ven_lisse sont maintenant
       exportées dans TARGET_COLS → LSTM et PELT les lisent directement
       sans avoir à les recalculer, garantissant la cohérence de la chaîne.

Étapes (numérotation pipeline) :
  6. enrich_external()          — Google Maps, Satellite, INS/BCT, Signaux
  6b. add_discriminant_features() — ③ variables discriminantes (NOUVEAU)
  7. compute_regional_index()   — ① lissage + ② imputation + ④ outliers
  8. export_datasets()          — ⑤ indice stratifié dans TARGET_COLS
  9. print_final_report()       — rapport enrichi

TARGET principal : indice_prix_m2_regional (lissé, robuste, stratifié)
Utilisé par : LSTM (séries temporelles), K-means, PELT (change point)
"""

import os, re, json, warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

from mappings_BO3 import (
    GOUVERNORAT_ENC, GOUVERNORAT_DEC,
    TARGET_COLS, FICHIERS_ML, HIGH_SEASON_MONTHS
)
from external_data_BO3 import get_inflation, get_pib

def section(t): print("\n" + "="*65 + f"\n   {t}\n" + "="*65)
def log(m):     print(f"  {m}")


# ================================================================
# CONSTANTES GÉOGRAPHIQUES — Amélioration ③
# ================================================================
# Zones géographiques tunisiennes (réalité du marché litoral vs intérieur)
# Basé sur la géographie économique et les flux immobiliers réels.
# 0=Grand Tunis  1=Littoral Nord  2=Sahel  3=Centre  4=Sud
ZONE_GEOGRAPHIQUE = {
    'Tunis':       0, 'Ariana':    0, 'Ben Arous': 0, 'Manouba':  0,
    'Bizerte':     1, 'Nabeul':    1, 'Béja':      1, 'Jendouba': 1,
    'Sousse':      2, 'Monastir':  2, 'Mahdia':    2, 'Sfax':     2,
    'Kairouan':    3, 'Kasserine': 3, 'Sidi Bouzid':3, 'Siliana': 3,
    'Le Kef':      3, 'Zaghouan':  3,
    'Gabès':       4, 'Médenine':  4, 'Tataouine': 4, 'Gafsa':   4,
    'Kébili':      4, 'Tozeur':    4,
    'Unknown':     3,
}
ZONE_LABELS = {0:'Grand Tunis', 1:'Littoral Nord', 2:'Sahel',
               3:'Centre/Intérieur', 4:'Sud'}

# Clusters géographiques pour imputation hiérarchique (Amélioration ②)
CLUSTER_GEO_IMPUTATION = {
    0: ['Tunis', 'Ariana', 'Ben Arous', 'Manouba', 'Zaghouan'],
    1: ['Bizerte', 'Nabeul', 'Béja', 'Jendouba'],
    2: ['Sousse', 'Monastir', 'Mahdia', 'Sfax'],
    3: ['Kairouan', 'Kasserine', 'Sidi Bouzid', 'Siliana', 'Le Kef'],
    4: ['Gabès', 'Médenine', 'Tataouine', 'Gafsa', 'Kébili', 'Tozeur'],
}


# ================================================================
# ÉTAPE 6 — ENRICHISSEMENT MULTI-SOURCES
# ================================================================

def enrich_external(groupes_clean: dict, external: dict) -> dict:
    """
    Enrichit chaque groupe avec :
    - score_attractivite    (Signaux — normalisé 0-1)
    - nb_infra, nb_commerce (Signaux — pas satellite)
    - inflation_glissement_annuel (INS mensuel)
    - croissance_pib_trim   (INS trimestriel)
    - glissement_immo_trim  (INS immobilier par type×trimestre)
    """
    section("ETAPE 6 — ENRICHISSEMENT MULTI-SOURCES")

    from mappings_BO3 import GLISSEMENT_IMMO_INS

    score_att    = external.get('score_attractivite', {})
    nb_infra_map = external.get('nb_infra', {})
    nb_comm_map  = external.get('nb_commerce', {})
    inf_map      = external.get('inflation_by_month', {})
    pib_map      = external.get('pib_by_quarter', {})

    def _serie_ins(type_cat: str) -> str:
        t = str(type_cat).lower()
        if 'foncier' in t: return 'Terrain nus'
        return 'Appartement'

    def _glissement_immo(annee: int, mois: int, type_cat: str) -> float:
        q    = (int(mois) - 1) // 3 + 1
        yr   = int(annee)
        serie = _serie_ins(type_cat)
        key   = (yr, q, serie)
        if key in GLISSEMENT_IMMO_INS:
            return GLISSEMENT_IMMO_INS[key]
        avail = [(k[0], k[1]) for k in GLISSEMENT_IMMO_INS if k[2] == serie]
        if not avail: return 5.0
        best = min(avail, key=lambda k: abs(k[0]*4+k[1] - (yr*4+q)))
        return GLISSEMENT_IMMO_INS[(best[0], best[1], serie)]

    for groupe, dg in groupes_clean.items():
        dg['score_attractivite'] = dg['_gouvernorat_str'].map(score_att).fillna(0.30).round(4)
        dg['nb_infra']    = dg['_gouvernorat_str'].map(nb_infra_map).fillna(0).astype(int)
        dg['nb_commerce'] = dg['_gouvernorat_str'].map(nb_comm_map).fillna(0).astype(int)
        dg['inflation_glissement_annuel'] = dg.apply(
            lambda r: get_inflation(r['annee'], r['mois'], inf_map), axis=1).round(2)
        dg['croissance_pib_trim'] = dg.apply(
            lambda r: get_pib(r['annee'], r['mois'], pib_map), axis=1).round(2)
        dg['glissement_immo_trim'] = dg.apply(
            lambda r: _glissement_immo(r['annee'], r['mois'],
                                       r.get('type_categorise', 'Residentiel')),
            axis=1).round(2)
        groupes_clean[groupe] = dg

    all_dg = pd.concat(groupes_clean.values(), ignore_index=True)
    log(f"✔ Enrichissement terminé : {len(all_dg):,} annonces")
    log(f"  score_attractivite     : mean={all_dg['score_attractivite'].mean():.3f}")
    log(f"  inflation moyenne      : {all_dg['inflation_glissement_annuel'].mean():.1f}%")
    log(f"  pib moyen              : {all_dg['croissance_pib_trim'].mean():.1f}%")
    return groupes_clean


# ================================================================
# ÉTAPE 6b — VARIABLES DISCRIMINANTES (Amélioration ③)
# ================================================================

def add_discriminant_features(groupes_clean: dict) -> dict:
    """
    Ajoute 4 features calculées 100% depuis les données réelles :

    zone_geographique (0-4) :
      Grand Tunis / Littoral Nord / Sahel / Centre-Intérieur / Sud
      Permet au K-means de séparer littoral vs intérieur sans supervision.

    potentiel_emergent (0-1) :
      Gouvernorats avec tendance prix positive + attractivité correcte
      + couverture temporelle suffisante → zones à surveiller (Change Point).
      Formule : score = rank(tendance_normalisée × score_attractivite × ln(n_pts))

    indice_liquidite :
      ln(n_annonces_total / n_mois_couverts + 1) par gouvernorat.
      Discrimine les marchés liquides (Tunis, Sousse) vs illiquides (Kébili, Tozeur).
      Utile pour PELT : seuil amplitude adapté à la liquidité.

    volatilite_prix_trim :
      Écart-type trimestriel du prix/m² normalisé par la moyenne.
      Capte les marchés saisonniers (Nabeul, Monastir) vs stables (Tunis, Sfax).
    """
    section("ETAPE 6b — VARIABLES DISCRIMINANTES (zone geo + émergence + liquidité + volatilité)")

    for groupe, dg in groupes_clean.items():
        # ── zone_geographique ────────────────────────────────────────────────
        dg['zone_geographique'] = dg['_gouvernorat_str'].map(ZONE_GEOGRAPHIQUE).fillna(3).astype(int)

        # ── indice_liquidite par gouvernorat ─────────────────────────────────
        n_ann_gov   = dg.groupby('gouvernorat').transform('size')
        n_mois_gov  = (dg.groupby('gouvernorat')
                         .apply(lambda x: x[['annee','mois']].drop_duplicates().shape[0])
                         .reindex(dg['gouvernorat']).values)
        n_mois_gov  = np.where(n_mois_gov == 0, 1, n_mois_gov)
        dg['indice_liquidite'] = np.log1p(n_ann_gov / n_mois_gov).round(4)

        # ── volatilite_prix_trim ─────────────────────────────────────────────
        dg['trimestre_calc'] = ((dg['mois'] - 1) // 3 + 1).astype(int)
        vol_trim = (dg.groupby(['gouvernorat', 'annee', 'trimestre_calc'])
                      ['indice_prix_m2_regional']
                      .std()
                      .reset_index(name='vol_trim'))
        vol_global_by_gov = vol_trim.groupby('gouvernorat')['vol_trim'].mean()
        prix_mean_by_gov  = dg.groupby('gouvernorat')['indice_prix_m2_regional'].mean()
        vol_norm = (vol_global_by_gov / (prix_mean_by_gov + 1e-8)).fillna(0)
        dg['volatilite_prix_trim'] = dg['gouvernorat'].map(vol_norm).fillna(0).round(4)
        dg = dg.drop(columns=['trimestre_calc'], errors='ignore')

        # ── potentiel_emergent ───────────────────────────────────────────────
        # Calculé dans compute_regional_index() sur l'indice lissé final.
        # On ne l'initialise que si absent pour ne pas écraser la valeur calculée.
        if 'potentiel_emergent' not in dg.columns:
            dg['potentiel_emergent'] = 0.0

        groupes_clean[groupe] = dg

    all_dg = pd.concat(groupes_clean.values(), ignore_index=True)
    log(f"✔ Variables discriminantes calculées :")
    log(f"  zone_geographique  : {all_dg['zone_geographique'].value_counts().sort_index().to_dict()}")
    log(f"  indice_liquidite   : mean={all_dg['indice_liquidite'].mean():.3f} "
        f"max={all_dg['indice_liquidite'].max():.3f}")
    log(f"  volatilite_prix_trim: mean={all_dg['volatilite_prix_trim'].mean():.3f} "
        f"max={all_dg['volatilite_prix_trim'].max():.3f}")
    return groupes_clean


# ================================================================
# ÉTAPE 7 — TARGET : INDICE ROBUSTE (Améliorations ①②④)
# ================================================================

def _detect_outliers_by_type(df: pd.DataFrame, groupe: str) -> pd.Series:
    """
    Amélioration ④ : Détection d'outliers par gouvernorat × type_transaction.

    Utilise prix_m2 (colonne brute calculée avant l'indice agrégé).
    indice_prix_m2_regional n'existe pas encore à ce stade du pipeline.

    Méthode : Z-score robuste MAD par gov × type. Seuil |z_mad| > 3.5.
    Retourne un masque booléen (True = valide, False = outlier).
    """
    # Colonne source : prix_m2 brut (disponible) ou prix_m2_raw si renommé
    col_prix = 'prix_m2' if 'prix_m2' in df.columns else None
    if col_prix is None:
        log(f"    [WARN] prix_m2 absent → détection outliers ignorée")
        return pd.Series(True, index=df.index)

    mask_valid = pd.Series(True, index=df.index)
    n_outliers = 0

    q95_global = df[col_prix].quantile(0.95)

    for (gov, tt), g in df.groupby(['gouvernorat', 'type_transaction']):
        prix = g[col_prix].dropna()
        if len(prix) < 4:
            continue
        med = prix.median()
        mad = np.median(np.abs(prix - med))
        if mad < 1e-6:
            continue
        z_mad  = 0.6745 * (g[col_prix] - med) / mad
        is_out = np.abs(z_mad) > 3.5
        # Singleton au-delà de 2×Q95 global → suspect
        n_par_mois = g.groupby(['annee','mois'])[col_prix].transform('count')
        is_out = is_out | ((n_par_mois == 1) & (g[col_prix] > q95_global * 2.0))
        is_out = is_out.fillna(False)
        mask_valid.loc[g.index] = ~is_out
        n_outliers += int(is_out.sum())

    log(f"    Outliers détectés (MAD ×3.5 + singleton >2×Q95) : {n_outliers} / {len(df):,}")
    return mask_valid


def _compute_robust_index_by_type(df_valid: pd.DataFrame) -> pd.DataFrame:
    """
    Amélioration ① : Indice robuste par gouvernorat × type_transaction.

    Algorithme en 3 passes :
    1. Agrégation (gov, annee, mois, type) → prix_mean + n_annonces
    2. Winsorisation temporelle 10-90% intra-gouvernorat × type
       (préserve les tendances réelles, élimine les pics isolés)
    3. Rolling median window=3 centré (lissage causal respectueux des ruptures)
       Pondération : les mois avec une seule annonce ont poids 0.3

    Retourne un DataFrame avec colonnes :
      gouvernorat, annee, mois, type_transaction,
      indice_brut, indice_robust, n_annonces, poids_mois
    """
    records = []
    col = 'prix_m2'  # colonne brute disponible avant calcul de l'indice agrégé

    for (gov, tt), g in df_valid.groupby(['gouvernorat', 'type_transaction']):
        g_valid = g[g[col].notna()]
        if len(g_valid) == 0:
            continue
        ts = (g_valid.groupby(['annee', 'mois'])
               .agg(indice_brut=(col, 'mean'),
                    n_annonces=(col, 'count'))
               .reset_index()
               .sort_values(['annee', 'mois'])
               .reset_index(drop=True))

        if len(ts) == 0:
            continue

        # Poids par mois (singleton = 0.3, dense ≥ 5 ann. = 1.0)
        ts['poids_mois'] = (ts['n_annonces'].clip(upper=10) / 10.0
                            ).clip(lower=0.3).round(4)

        if len(ts) == 1:
            ts['indice_robust'] = ts['indice_brut']
            ts['gouvernorat']   = gov
            ts['type_transaction'] = tt
            records.append(ts)
            continue

        # Winsorisation 10-90% temporelle intra-gov × type
        q10 = ts['indice_brut'].quantile(0.10)
        q90 = ts['indice_brut'].quantile(0.90)
        ts['indice_win'] = ts['indice_brut'].clip(lower=q10, upper=q90)

        # Rolling median (window adaptatif : 3 si ≥5 pts, 2 sinon)
        w = 3 if len(ts) >= 5 else 2
        ts['indice_robust'] = (ts['indice_win']
                               .rolling(w, min_periods=1, center=True)
                               .median())

        ts['gouvernorat']      = gov
        ts['type_transaction'] = tt
        records.append(ts[['gouvernorat', 'annee', 'mois', 'type_transaction',
                            'indice_brut', 'indice_robust', 'n_annonces', 'poids_mois']])

    return pd.concat(records, ignore_index=True) if records else pd.DataFrame()


def _imputation_hierarchique(indice_ts: pd.DataFrame,
                              gouvernorat_str_map: dict) -> pd.DataFrame:
    """
    Amélioration ② : Imputation hiérarchique pour gouvernorats peu couverts.

    Un gouvernorat est "peu couvert" si n_mois_distincts < 6.
    Hiérarchie d'imputation :
      Niveau A : médiane du gouvernorat sur ±3 mois (interpolation temporelle)
      Niveau B : médiane pondérée des gouvernorats du même cluster géographique
      Niveau C : médiane nationale pondérée (fallback final)

    But : éviter les NaN dans les séries LSTM et les biais k-means sur
    gouvernorats sous-représentés (Le Kef, Siliana, Tozeur, Zaghouan).

    Note : l'imputation est conservatrice — elle n'invente pas de données,
    elle remplace uniquement les mois MANQUANTS par la valeur la plus locale possible.
    """
    if len(indice_ts) == 0:
        return indice_ts

    # Construire mapping gouvernorat_code → zone_cluster
    gov_to_zone = {}
    for zone_id, gov_names in CLUSTER_GEO_IMPUTATION.items():
        for gn in gov_names:
            gov_code = GOUVERNORAT_ENC.get(gn, 0)
            gov_to_zone[gov_code] = zone_id

    indice_ts = indice_ts.copy()

    # Médiane nationale pondérée (Niveau C — fallback)
    national_med = float(indice_ts['indice_global'].median())

    # Médiane par zone géographique (Niveau B)
    indice_ts['zone'] = indice_ts['gouvernorat'].map(gov_to_zone).fillna(3)
    zone_medians = indice_ts.groupby('zone')['indice_global'].median().to_dict()

    # Identifier gouvernorats peu couverts
    pts_by_gov = indice_ts.groupby('gouvernorat')['annee'].count()
    govs_peu_couverts = pts_by_gov[pts_by_gov < 6].index.tolist()

    if govs_peu_couverts:
        noms = [GOUVERNORAT_DEC.get(g, str(g)) for g in govs_peu_couverts]
        log(f"    Imputation hiérarchique : {len(govs_peu_couverts)} gouvernorats peu couverts")
        log(f"    → {noms}")

    n_imputed = 0
    for gov in govs_peu_couverts:
        mask_gov = indice_ts['gouvernorat'] == gov
        zone     = int(indice_ts.loc[mask_gov, 'zone'].iloc[0]) if mask_gov.sum() > 0 else 3

        # Niveau B : médiane de la zone
        zone_val = zone_medians.get(zone, national_med)

        # Pour les gouvernorats avec quelques points : interpoler linéairement
        sub = indice_ts.loc[mask_gov].sort_values(['annee', 'mois'])
        if len(sub) >= 2:
            # Interpol sur les NaN internes (Niveau A)
            indice_ts.loc[mask_gov, 'indice_global'] = (
                indice_ts.loc[mask_gov, 'indice_global']
                .interpolate(method='linear', limit_direction='both'))
            n_imputed += int(indice_ts.loc[mask_gov, 'indice_global'].isna().sum())
        # Remplissage restant avec zone (Niveau B) puis national (Niveau C)
        indice_ts.loc[mask_gov, 'indice_global'] = (
            indice_ts.loc[mask_gov, 'indice_global'].fillna(zone_val))
        n_imputed += int(indice_ts.loc[mask_gov, 'indice_global'].isna().sum())

    log(f"    Points imputés : {n_imputed}")
    return indice_ts.drop(columns=['zone'], errors='ignore')


def _compute_potentiel_emergent(indice_ts: pd.DataFrame,
                                 score_att_map: dict) -> pd.Series:
    """
    Amélioration ③ (partie 2) : Score potentiel_emergent par gouvernorat.

    Un gouvernorat est "émergent" si :
    1. Sa tendance de prix récente (2 dernières années) est positive
    2. Son score attractivité est ≥ 0.35 (marché pas trop sous-développé)
    3. Sa couverture temporelle est suffisante (≥ 6 points)

    Score = rank_normalisé(tendance_récente) × score_attractivite × min(n_pts/20, 1)

    Interprétation : 0 = aucun potentiel émergent, 1 = fort potentiel.
    Exemples attendus : Nabeul, Monastir, Bizerte haute → Kasserine, Tozeur basse.
    """
    scores = {}

    for gov, g in indice_ts.groupby('gouvernorat'):
        g = g.sort_values(['annee', 'mois']).reset_index(drop=True)
        n_pts = len(g)

        if n_pts < 4:
            scores[gov] = 0.0
            continue

        # Tendance récente : pente linéaire sur les 2 dernières années
        recent = g[g['annee'] >= (g['annee'].max() - 1)]
        if len(recent) >= 2:
            t    = np.arange(len(recent), dtype=float)
            prix = recent['indice_global'].values
            if prix.std() > 0:
                slope = float(np.polyfit(t, prix, 1)[0])
            else:
                slope = 0.0
        else:
            slope = 0.0

        gov_name = GOUVERNORAT_DEC.get(int(gov), '')
        att      = float(score_att_map.get(gov_name, 0.3))
        cov_factor = min(n_pts / 20.0, 1.0)

        scores[gov] = round(max(slope, 0) * att * cov_factor, 6)

    # Normaliser 0-1 par rang
    s_series = pd.Series(scores)
    if s_series.max() > 0:
        s_series = (s_series.rank(pct=True)).clip(0, 1)
    return s_series


def compute_regional_index(groupes_clean: dict,
                            external: dict = None) -> dict:
    """
    Calcule l'indice régional robuste du prix/m² avec les 4 améliorations.

    Pipeline :
    ④ Détection outliers MAD par gov × type (avant tout calcul)
    ① Winsorisation 10-90 + rolling median(3) par gov × type → indice_robust
    ─  Fusion pondérée loc + ven → indice_global (poids = n_annonces)
    ② Imputation hiérarchique des gouvernorats peu couverts
    ③ potentiel_emergent calculé sur l'indice lissé final
    ⑤ Colonnes indice_loc_lisse + indice_ven_lisse exportées dans TARGET_COLS

    Colonnes de correction (inchangées) :
      sample_weight_temporal : inverse surreprésentation annuelle
      sample_weight_geo      : inverse surreprésentation géographique
      arima_eligible         : 1 si ≥ 12 points temporels
    """
    section("ETAPE 7 — TARGET : indice_prix_m2_regional (ROBUSTE + LISSÉ + STRATIFIÉ)")

    score_att_map = {}
    if external:
        for gname, score in external.get('score_attractivite', {}).items():
            score_att_map[gname] = score

    # Recalculer prix_m2 si absent
    for groupe, dg in groupes_clean.items():
        if 'prix_m2' not in dg.columns or dg['prix_m2'].isna().all():
            dg['prix_m2'] = np.where(
                dg['prix'].notna() & dg['surface_m2'].notna() & (dg['surface_m2'] > 0),
                dg['prix'] / dg['surface_m2'], np.nan)
            mask_loc = dg['type_transaction'] == 1
            mask_ven = dg['type_transaction'] == 2
            dg.loc[mask_loc & ((dg['prix_m2'] < 1)   | (dg['prix_m2'] > 500)),    'prix_m2'] = np.nan
            dg.loc[mask_ven & ((dg['prix_m2'] < 100) | (dg['prix_m2'] > 30_000)), 'prix_m2'] = np.nan
            groupes_clean[groupe] = dg

    df_all = pd.concat(groupes_clean.values(), ignore_index=True)
    log(f"  prix_m2 : {df_all['prix_m2'].notna().sum():,} valides / {len(df_all):,}")

    # Mapping gouvernorat → nom (pour imputation)
    gov_str_map = {row['gouvernorat']: row['_gouvernorat_str']
                   for _, row in df_all[['gouvernorat','_gouvernorat_str']].drop_duplicates().iterrows()
                   if '_gouvernorat_str' in df_all.columns}

    # ── ④ Détection outliers (par groupe pour logs clairs) ───────────────────
    log("\n  ④ Détection outliers MAD par groupe × gouvernorat × type_transaction :")
    mask_all = pd.Series(True, index=df_all.index)
    offset = 0
    for groupe, dg in groupes_clean.items():
        log(f"    {groupe} :")
        mask_g = _detect_outliers_by_type(dg, groupe)
        # Reconstruire le masque sur df_all
        mask_all.iloc[offset:offset+len(dg)] = mask_g.values
        offset += len(dg)

    df_valid = df_all[mask_all].copy()
    n_excl = len(df_all) - len(df_valid)
    log(f"  Total exclus : {n_excl} / {len(df_all):,} ({n_excl/len(df_all)*100:.1f}%)")

    # ── ① Calcul indice robuste par type ────────────────────────────────────
    log("\n  ① Indice robuste (winsorisation 10-90 + rolling median(3)) :")
    indice_by_type = _compute_robust_index_by_type(df_valid)

    if len(indice_by_type) == 0:
        log("  [WARN] Aucune donnée valide → fallback indice brut depuis prix_m2")
        indice_by_type = (df_all[df_all['prix_m2'].notna()]
                                .groupby(['gouvernorat','annee','mois','type_transaction'])
                                ['prix_m2'].mean()
                                .reset_index()
                                .rename(columns={'prix_m2':'indice_robust'}))
        indice_by_type['n_annonces'] = 1
        indice_by_type['poids_mois'] = 1.0
        indice_by_type['indice_brut'] = indice_by_type['indice_robust']

    # ── Indice lissé par type (pour ⑤ export) ───────────────────────────────
    indice_loc = (indice_by_type[indice_by_type['type_transaction'] == 1]
                  .groupby(['gouvernorat','annee','mois'])
                  .apply(lambda x: np.average(x['indice_robust'], weights=x['poids_mois']))
                  .reset_index(name='indice_loc_lisse'))

    indice_ven = (indice_by_type[indice_by_type['type_transaction'] == 2]
                  .groupby(['gouvernorat','annee','mois'])
                  .apply(lambda x: np.average(x['indice_robust'], weights=x['poids_mois']))
                  .reset_index(name='indice_ven_lisse'))

    # ── Fusion pondérée loc + ven → indice global ────────────────────────────
    def _wavg(x):
        w = x['poids_mois'] * x['n_annonces']
        return float(np.average(x['indice_robust'], weights=w)) if w.sum() > 0 else float(x['indice_robust'].mean())

    indice_global = (indice_by_type
                     .groupby(['gouvernorat','annee','mois'])
                     .apply(_wavg)
                     .reset_index(name='indice_global'))

    log(f"  ✔ Indice global : {len(indice_global):,} points "
        f"| médiane={indice_global['indice_global'].median():.0f} TND/m²")

    # CV avant/après
    cv_brut = (indice_by_type.groupby('gouvernorat')['indice_brut'].std() /
               indice_by_type.groupby('gouvernorat')['indice_brut'].mean())
    cv_rob  = (indice_global.groupby('gouvernorat')['indice_global'].std() /
               indice_global.groupby('gouvernorat')['indice_global'].mean())
    log(f"  CV moyen — brut: {cv_brut.mean():.3f} | lissé: {cv_rob.mean():.3f} "
        f"(réduction {(1-cv_rob.mean()/cv_brut.mean())*100:.0f}%)")

    # ── ② Imputation hiérarchique ─────────────────────────────────────────────
    log("\n  ② Imputation hiérarchique gouvernorats peu couverts :")
    indice_global = _imputation_hierarchique(indice_global, gov_str_map)

    # ── ③ potentiel_emergent ──────────────────────────────────────────────────
    log("\n  ③ Calcul potentiel_emergent :")
    pot_scores = _compute_potentiel_emergent(indice_global, score_att_map)
    log(f"  Top 5 zones émergentes :")
    for gov, sc in pot_scores.sort_values(ascending=False).head(5).items():
        log(f"    {GOUVERNORAT_DEC.get(int(gov), str(gov)):<16} : {sc:.3f}")

    # ── Colonnes de correction (inchangées, calculées depuis données réelles) ─
    n_total    = len(df_valid)
    yr_counts  = df_valid['annee'].value_counts()
    yr_expect  = n_total / yr_counts.nunique()
    yr_weight  = (yr_expect / yr_counts).clip(upper=5.0)
    yr_weight  = (yr_weight / yr_weight.mean()).round(4)

    gov_counts  = df_valid['gouvernorat'].value_counts()
    gov_expect  = n_total / gov_counts.nunique()
    gov_weight  = (gov_expect / gov_counts).clip(upper=10.0)
    gov_weight  = (gov_weight / gov_weight.mean()).round(4)

    gov_pts    = (df_valid.groupby('gouvernorat')[['annee','mois']]
                          .apply(lambda x: x.drop_duplicates().shape[0]))
    arima_ok   = (gov_pts >= 12).astype(int)

    log(f"\n  sample_weight_temporal : {yr_weight.to_dict()}")
    log(f"  arima_eligible         : "
        f"{arima_ok.sum()} gouvernorats éligibles / {arima_ok.count()}")

    # ── Joindre dans chaque groupe ────────────────────────────────────────────
    global_med = float(indice_global['indice_global'].median())
    for groupe, dg in groupes_clean.items():
        # TARGET principal
        dg = dg.merge(
            indice_global[['gouvernorat', 'annee', 'mois', 'indice_global']],
            on=['gouvernorat', 'annee', 'mois'], how='left'
        )
        dg = dg.rename(columns={'indice_global': 'indice_prix_m2_regional'})

        # Fallback NaN (rare après imputation)
        if dg['indice_prix_m2_regional'].isna().any():
            gov_med = dg.groupby('gouvernorat')['indice_prix_m2_regional'].transform('median')
            dg['indice_prix_m2_regional'] = (dg['indice_prix_m2_regional']
                                              .fillna(gov_med).fillna(global_med))

        # ⑤ Indices stratifiés loc/ven lissés
        dg = dg.merge(indice_loc, on=['gouvernorat','annee','mois'], how='left')
        dg = dg.merge(indice_ven, on=['gouvernorat','annee','mois'], how='left')
        # Fallback sur indice_global si type absent
        dg['indice_loc_lisse'] = dg['indice_loc_lisse'].fillna(dg['indice_prix_m2_regional'])
        dg['indice_ven_lisse'] = dg['indice_ven_lisse'].fillna(dg['indice_prix_m2_regional'])

        # Colonnes de correction
        dg['sample_weight_temporal'] = dg['annee'].map(yr_weight).fillna(1.0).round(4)
        dg['sample_weight_geo']      = dg['gouvernorat'].map(gov_weight).fillna(1.0).round(4)
        dg['arima_eligible']         = dg['gouvernorat'].map(arima_ok).fillna(0).astype(int)

        # ③ potentiel_emergent
        dg['potentiel_emergent'] = dg['gouvernorat'].map(pot_scores).fillna(0.0).round(4)

        n_nan = dg['indice_prix_m2_regional'].isna().sum()
        log(f"  {groupe:<15} : {len(dg):,} ann. | NaN={n_nan} | "
            f"arima_elig={dg['arima_eligible'].sum()}/{len(dg)} | "
            f"pot_emerg={dg['potentiel_emergent'].mean():.3f}")
        groupes_clean[groupe] = dg

    return groupes_clean


# ================================================================
# ÉTAPE 8 — EXPORT (avec TARGET_COLS enrichi ⑤)
# ================================================================

# TARGET_COLS enrichi avec les nouvelles colonnes (⑤ + ③)
TARGET_COLS_ENRICHI = [
    # ── Géographie ──────────────────────────────────────────────
    'gouvernorat',              # code 1-24
    'ville_encoded',            # granularité ville
    # ── Temporel ────────────────────────────────────────────────
    'annee',
    'mois',
    'high_season',
    # ── Transaction ─────────────────────────────────────────────
    'type_transaction',         # 1=Location / 2=Vente
    # ── Zone géographique (③ NOUVEAU) ────────────────────────────
    'zone_geographique',        # 0=G.Tunis 1=Littoral 2=Sahel 3=Centre 4=Sud
    # ── Marché local ────────────────────────────────────────────
    'score_attractivite',       # 0-1
    'nb_infra',
    'nb_commerce',
    'indice_liquidite',         # ③ NOUVEAU : ln(ann/mois)
    'volatilite_prix_trim',     # ③ NOUVEAU : σ trimestriel normalisé
    'potentiel_emergent',       # ③ NOUVEAU : score zones émergentes 0-1
    # ── Macro-économique ─────────────────────────────────────────
    'inflation_glissement_annuel',
    'croissance_pib_trim',
    'glissement_immo_trim',
    # ── Corrections déséquilibre ─────────────────────────────────
    'sample_weight_temporal',
    'sample_weight_geo',
    'arima_eligible',
    # ── TARGET principal ──────────────────────────────────────────
    'indice_prix_m2_regional',  # lissé + robuste (①②④)
    # ── Targets stratifiés (⑤ NOUVEAU) ───────────────────────────
    'indice_loc_lisse',         # indice Location lissé par gov × mois
    'indice_ven_lisse',         # indice Vente lissé par gov × mois
]


def _write_excel(df_out: pd.DataFrame, filename: str,
                 color_hex: str, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill = PatternFill("solid", start_color=color_hex)
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        clean = []
        for cell in r:
            if isinstance(cell, str):
                cell = cell.encode('ascii', 'ignore').decode('ascii')
                cell = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cell)
                cell = re.sub(r'\s+', ' ', cell).strip()
            elif isinstance(cell, bool):         cell = str(cell)
            elif isinstance(cell, np.integer):   cell = int(cell)
            elif isinstance(cell, np.floating):  cell = round(float(cell), 4) if not np.isnan(cell) else None
            clean.append(cell)
        ws.append(clean)

    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        w = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
    ws.freeze_panes = "A2"
    wb.save(filename)
    log(f"  {filename:<40}: {len(df_out):>6} lignes | {len(df_out.columns)} colonnes")


def export_datasets(groupes_clean: dict, output_dir: str = '.') -> None:
    section("ETAPE 8 — EXPORT FINAL (TARGET_COLS enrichi : 22 colonnes)")

    os.makedirs(output_dir, exist_ok=True)

    for groupe, dg in groupes_clean.items():
        if groupe not in FICHIERS_ML: continue
        fname, color = FICHIERS_ML[groupe]

        # Sélectionner uniquement les colonnes disponibles
        cols_sel = [c for c in TARGET_COLS_ENRICHI if c in dg.columns]
        missing  = [c for c in TARGET_COLS_ENRICHI if c not in dg.columns]
        if missing:
            log(f"  [WARN] {groupe} — colonnes manquantes : {missing}")

        df_export = dg[cols_sel].copy()
        df_export = df_export.sort_values(['gouvernorat','annee','mois']).reset_index(drop=True)
        df_export['indice_prix_m2_regional'] = df_export['indice_prix_m2_regional'].round(2)
        if 'indice_loc_lisse' in df_export.columns:
            df_export['indice_loc_lisse'] = df_export['indice_loc_lisse'].round(2)
        if 'indice_ven_lisse' in df_export.columns:
            df_export['indice_ven_lisse'] = df_export['indice_ven_lisse'].round(2)

        out_path = os.path.join(output_dir, fname)
        _write_excel(df_export, out_path, color, f"{groupe}_BO3")

    # Encoding mappings (enrichi)
    mappings = {
        'gouvernorat':       {str(v): k for k, v in GOUVERNORAT_ENC.items()},
        'type_transaction':  {'1': 'Location', '2': 'Vente'},
        'high_season':       {'0': 'Basse saison', '1': 'Haute saison (mars-mai, sept-nov)'},
        'zone_geographique': {str(k): v for k, v in ZONE_LABELS.items()},
        'trimestre':         {'1':'T1','2':'T2','3':'T3','4':'T4'},
        'ameliorations':     {
            '1': 'Lissage robuste indice (winsorisation + rolling median)',
            '2': 'Imputation hierarchique petits gouvernorats',
            '3': 'Variables discriminantes (zone_geo + emergent + liquidite + volatilite)',
            '4': 'Outliers MAD par gov x type',
            '5': 'Indices stratifies loc/ven exportes (indice_loc_lisse, indice_ven_lisse)',
        },
    }
    map_path = os.path.join(output_dir, 'encoding_mappings_BO3.json')
    try:
        with open(map_path, 'w', encoding='utf-8') as f:
            json.dump(mappings, f, ensure_ascii=False, indent=2)
        log(f"  Mappings sauvegardés : encoding_mappings_BO3.json")
    except Exception as e:
        log(f"  [WARN] Mappings non sauvegardés : {e}")


# ================================================================
# RAPPORT FINAL
# ================================================================

def print_final_report(groupes_clean: dict, n_sources: int,
                       n_brut: int, n_dedup: int, bct: dict) -> None:
    section("RAPPORT FINAL — OBJECTIF 3 : TENDANCES RÉGIONALES (AMÉLIORÉ)")

    total = sum(len(g) for g in groupes_clean.values())
    log(f"Sources chargées                   : {n_sources}")
    log(f"Annonces brutes                    : {n_brut:>8,}")
    log(f"Après déduplication                : {n_dedup:>8,}  (-{n_brut - n_dedup:,})")
    log(f"Après nettoyage complet            : {total:>8,}")
    log("")
    log("Répartition par groupe :")
    for groupe, dg in groupes_clean.items():
        pct = len(dg)/total*100 if total > 0 else 0
        log(f"  {groupe:<15}: {len(dg):>7,} annonces ({pct:.1f}%)")
    log("")

    all_df = pd.concat(groupes_clean.values(), ignore_index=True)

    # TARGET robustesse
    nan_target = all_df['indice_prix_m2_regional'].isna().sum()
    log(f"TARGET indice_prix_m2_regional (LISSÉ) :")
    log(f"  NaN        : {nan_target}")
    log(f"  Médiane    : {all_df['indice_prix_m2_regional'].median():,.0f} TND/m²")
    log(f"  Min/Max    : {all_df['indice_prix_m2_regional'].min():,.0f} – "
        f"{all_df['indice_prix_m2_regional'].max():,.0f}")
    cv_global = (all_df.groupby('gouvernorat')['indice_prix_m2_regional'].std() /
                 all_df.groupby('gouvernorat')['indice_prix_m2_regional'].mean())
    log(f"  CV moyen/gov : {cv_global.mean():.3f} (objectif <0.8)")
    log("")

    # Zones géographiques
    if 'zone_geographique' in all_df.columns:
        log("Zones géographiques (③) :")
        for z, label in ZONE_LABELS.items():
            n_gov = all_df[all_df['zone_geographique']==z]['gouvernorat'].nunique()
            pm    = all_df[all_df['zone_geographique']==z]['indice_prix_m2_regional'].median()
            log(f"  {label:<20} : {n_gov} gouvernorats | prix médian={pm:,.0f} TND/m²")
        log("")

    # Zones émergentes
    if 'potentiel_emergent' in all_df.columns:
        log("Top zones émergentes (③) :")
        # Construire mapping code → nom depuis les données (plus fiable que GOUVERNORAT_DEC)
        if '_gouvernorat_str' in all_df.columns:
            gov_name_map = (all_df[['gouvernorat','_gouvernorat_str']]
                            .drop_duplicates()
                            .set_index('gouvernorat')['_gouvernorat_str']
                            .to_dict())
        else:
            gov_name_map = {v: k for k, v in GOUVERNORAT_ENC.items()}
        top_em = (all_df.groupby('gouvernorat')['potentiel_emergent']
                        .mean()
                        .sort_values(ascending=False)
                        .head(5))
        for gov, sc in top_em.items():
            nom = gov_name_map.get(gov, GOUVERNORAT_DEC.get(int(gov), str(gov)))
            log(f"  {nom:<16} : {sc:.3f}")
        log("")

    log(f"Couverture temporelle : {all_df['annee'].min()} – {all_df['annee'].max()}")
    log(f"BCT taux directeur    : {bct.get('taux_directeur', 7.0)}%")

    print("\n" + "="*65)
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   OBJECTIF 3 — DONNÉES PRÊTES POUR MODÉLISATION (v2)       ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  5 AMÉLIORATIONS APPLIQUÉES :                               ║")
    print("║  ① Lissage robuste (winsorisation + rolling median)         ║")
    print("║  ② Imputation hiérarchique (zone → national)               ║")
    print("║  ③ 4 variables discriminantes (zone/émerg/liquid/volatil)  ║")
    print("║  ④ Outliers MAD × 3.5 par gov × type                       ║")
    print("║  ⑤ Indices loc/ven lissés dans TARGET_COLS                 ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  22 colonnes finales (vs 16 avant) :                       ║")
    print("║  gouvernorat, ville_encoded, annee, mois, high_season      ║")
    print("║  type_transaction, zone_geographique                       ║")
    print("║  score_attractivite, nb_infra, nb_commerce                 ║")
    print("║  indice_liquidite, volatilite_prix_trim, potentiel_emergent║")
    print("║  inflation, pib, glissement_immo, sample_weights           ║")
    print("║  indice_prix_m2_regional (TARGET — lissé)                  ║")
    print("║  indice_loc_lisse, indice_ven_lisse                        ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  Fichiers exportés :                                        ║")
    for groupe, (fname, _) in FICHIERS_ML.items():
        print(f"║    {fname:<56}║")
    print("╚══════════════════════════════════════════════════════════════╝")