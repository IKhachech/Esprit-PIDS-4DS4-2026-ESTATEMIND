"""
external_data_BO3.py — Données externes pour l'Objectif 3 : Tendances Régionales.

Sources :
  - BCT          : taux directeur + historique
  - INS          : inflation mensuelle + croissance PIB trimestrielle
  - Signaux      : score attractivité par gouvernorat
  - Satellite    : nb_infra + nb_commerce par gouvernorat
  - Google Maps  : note_google_moyenne par gouvernorat

Spécificités BO3 vs BO2 :
  - Inflation chargée au niveau MENSUEL (pas trimestriel)
  - PIB chargé au niveau TRIMESTRIEL
  - Satellite : agrégation par gouvernorat (pas par ville)
  - Google Maps : chargé depuis raw_data.json
"""

import os, re, io, json, zipfile, warnings
import numpy as np
import pandas as pd
import openpyxl

warnings.filterwarnings('ignore')

from mappings_BO3 import (
    GOUVERNORAT_ENC, VILLE_TO_GOUVERNORAT,
    INFLATION_FALLBACK, PIB_FALLBACK, DEFAULT_ATTRACTIVITE
)

DEFAULT_PATHS = {
    'bct':        '../bct/bct_dataset_20260219_1332.xlsx',
    'ins':        '../Ins/ins_dataset_20260219_1307.xlsx',
    'signaux':    '../goolgeMaps/signaux_immobilier_tunisie.xlsx',
    'satellite':  '../Tunisia_Satellite_scraper/tunisia_satellite_data',
    'google_json':'../goolgeMaps/raw_data.json',
}


# ================================================================
# HELPERS
# ================================================================

def _open_xlsx(path: str):
    buf = io.BytesIO()
    with zipfile.ZipFile(path, 'r', metadata_encoding='utf-8') as zin:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
    buf.seek(0)
    return openpyxl.load_workbook(buf, read_only=True, data_only=True)


def _sheet_to_df(path: str, sheet_name: str, header_row: int) -> pd.DataFrame:
    wb   = _open_xlsx(path)
    ws   = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows or header_row >= len(rows):
        return pd.DataFrame()
    headers = [str(c) if c is not None else f'col_{i}' for i, c in enumerate(rows[header_row])]
    data    = [r for r in rows[header_row + 1:] if any(v is not None for v in r)]
    return pd.DataFrame(data, columns=headers)


def normalize_gouvernorat(val: str) -> str:
    """Normalise un nom de gouvernorat ou ville → nom officiel."""
    if pd.isna(val) or str(val).strip() in ('', 'nan', 'None', 'NaN'):
        return 'Unknown'
    v = str(val).strip().lower()
    if v in VILLE_TO_GOUVERNORAT:
        return VILLE_TO_GOUVERNORAT[v]
    for key, gov in VILLE_TO_GOUVERNORAT.items():
        if key in v or v in key:
            return gov
    for gov_name in GOUVERNORAT_ENC:
        if gov_name.lower() == v:
            return gov_name
    return 'Unknown'


# ================================================================
# BCT
# ================================================================

def load_bct(path: str) -> dict:
    defaults = {'taux_directeur': 7.0, 'taux_moyen': 7.08, 'tre': 6.0, 'date': 'inconnu'}
    if not os.path.exists(path):
        print(f"  [WARN] BCT absent → fallback")
        return defaults
    try:
        df = _sheet_to_df(path, '🗂️ Dataset Complet', header_row=3)
        df['Valeur'] = pd.to_numeric(df['Valeur'], errors='coerce')
        result = dict(defaults)
        for _, row in df.iterrows():
            ind = str(row.get('Indicateur', '')).lower()
            val = row.get('Valeur')
            if pd.isna(val): continue
            if 'directeur' in ind:
                result['taux_directeur'] = float(val)
                d = row.get('Date')
                if d and not pd.isna(d): result['date'] = str(d)
            elif ind.strip() == 'taux':
                result['taux_moyen'] = float(val)
            elif any(x in ind for x in ['épargne', 'epargne', 'tre']):
                result['tre'] = float(val)
        print(f"  ✔ BCT chargé : taux_directeur={result['taux_directeur']}% | date={result['date']}")
        return result
    except Exception as e:
        print(f"  [WARN] BCT erreur : {e} → fallback")
        return defaults


# ================================================================
# INS — INFLATION MENSUELLE + PIB TRIMESTRIEL
# ================================================================

def load_ins_bo3(path: str) -> dict:
    """
    Charge INS pour BO3 :
    - inflation_by_month  : {(annee, mois): glissement_annuel_%}
    - pib_by_quarter      : {(annee, trimestre): croissance_%}

    Spécificité BO3 : inflation au niveau mensuel (pas trimestriel comme BO2).
    """
    result = {
        'inflation_by_month': dict(INFLATION_FALLBACK),
        'pib_by_quarter':     dict(PIB_FALLBACK),
    }
    if not os.path.exists(path):
        print(f"  [WARN] INS absent → fallback hardcodé")
        return result
    try:
        wb = _open_xlsx(path)
        sheet_names = wb.sheetnames
        wb.close()

        MOIS = {'janvier':1,'février':2,'mars':3,'avril':4,'mai':5,'juin':6,
                'juillet':7,'août':8,'septembre':9,'octobre':10,'novembre':11,'décembre':12}

        # Inflation mensuelle
        inf_sheet = next((s for s in sheet_names if 'inflation' in s.lower() or 'prix' in s.lower()), None)
        if inf_sheet:
            df_inf = _sheet_to_df(path, inf_sheet, header_row=3)
            df_inf['Valeur'] = pd.to_numeric(df_inf['Valeur'], errors='coerce')
            for _, row in df_inf.iterrows():
                p   = str(row.get('Période', '')).lower()
                m   = re.search(r'(\d{4})', p)
                val = row.get('Valeur')
                if not m or pd.isna(val): continue
                yr   = int(m.group(1))
                mois = next((v for k, v in MOIS.items() if k in p), None)
                if mois:
                    result['inflation_by_month'][(yr, mois)] = float(val)
            print(f"  ✔ INS inflation : {len(result['inflation_by_month'])} mois")

        # PIB trimestriel
        pib_sheet = next((s for s in sheet_names if 'pib' in s.lower() or 'croissance' in s.lower()), None)
        if pib_sheet:
            df_pib = _sheet_to_df(path, pib_sheet, header_row=3)
            df_pib['Valeur'] = pd.to_numeric(df_pib['Valeur'], errors='coerce')
            for _, row in df_pib.iterrows():
                p   = str(row.get('Période', '')).lower()
                m   = re.search(r'(\d{4})', p)
                val = row.get('Valeur')
                if not m or pd.isna(val): continue
                yr = int(m.group(1))
                if   'premi'  in p: q = 1
                elif 'deuxi'  in p: q = 2
                elif 'troisi' in p: q = 3
                elif 'quatri' in p: q = 4
                else: continue
                result['pib_by_quarter'][(yr, q)] = float(val)
            print(f"  ✔ INS PIB : {len(result['pib_by_quarter'])} trimestres")

        return result
    except Exception as e:
        print(f"  [WARN] INS erreur : {e} → fallback")
        return result


def get_inflation(annee: int, mois: int, inflation_map: dict) -> float:
    """Retourne le glissement annuel de l'inflation pour un mois donné."""
    key = (int(annee), int(mois))
    if key in inflation_map:
        return inflation_map[key]
    keys = sorted(inflation_map.keys())
    if not keys: return 5.0
    closest = min(keys, key=lambda k: abs(k[0]*12+k[1] - (annee*12+mois)))
    return inflation_map[closest]


def get_pib(annee: int, mois: int, pib_map: dict) -> float:
    """Retourne la croissance PIB trimestrielle pour un mois donné."""
    q   = (int(mois) - 1) // 3 + 1
    key = (int(annee), q)
    if key in pib_map:
        return pib_map[key]
    keys = sorted(pib_map.keys())
    if not keys: return 2.0
    closest = min(keys, key=lambda k: abs(k[0]*4+k[1] - (annee*4+q)))
    return pib_map[closest]


# ================================================================
# SIGNAUX — SCORE ATTRACTIVITÉ
# ================================================================

def load_signaux_bo3(path: str) -> dict:
    """
    Charge depuis signaux_immobilier_tunisie.xlsx (feuille Score Attractivité) :
      - score_attractivite  : normalisé 0-1 (÷100) — même échelle que BO2
      - nb_infra            : nb lieux infrastructure par gouvernorat (Google Maps)
      - nb_commerce         : nb lieux commerciaux par gouvernorat (Google Maps)

    Corrélations confirmées (faibles → colonnes indépendantes) :
      score ↔ nb_infra    : 0.40 → info distincte ✔
      score ↔ nb_commerce : 0.16 → info distincte ✔

    Avantage : même fichier que score_attractivite → 0 dépendance satellite.
    """
    defaults_score = {k: round(v / 100.0, 4) for k, v in DEFAULT_ATTRACTIVITE.items()}
    defaults_infra    = {k: 0 for k in DEFAULT_ATTRACTIVITE}
    defaults_commerce = {k: 0 for k in DEFAULT_ATTRACTIVITE}

    if not os.path.exists(path):
        print(f"  [WARN] Signaux absent → fallback score/nb_infra/nb_commerce")
        return {
            'score':    defaults_score,
            'nb_infra': defaults_infra,
            'nb_commerce': defaults_commerce,
        }
    try:
        df = _sheet_to_df(path, '🏆 Score Attractivité', header_row=0)
        df = df.dropna(subset=['gouvernorat'])

        score_map    = dict(defaults_score)
        infra_map    = dict(defaults_infra)
        commerce_map = dict(defaults_commerce)

        for col in ['score_attractivite', 'nb_infra', 'nb_commerce']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        for _, row in df.iterrows():
            gov = normalize_gouvernorat(str(row.get('gouvernorat', '')))
            if gov == 'Unknown': continue

            score = row.get('score_attractivite')
            if pd.notna(score):
                score_map[gov] = round(float(score) / 100.0, 4)

            nb_i = row.get('nb_infra')
            if pd.notna(nb_i):
                infra_map[gov] = int(nb_i)

            nb_c = row.get('nb_commerce')
            if pd.notna(nb_c):
                commerce_map[gov] = int(nb_c)

        print(f"  ✔ Signaux chargés : {len(score_map)} gouvernorats")
        print(f"    score_attractivite : 0-1 (normalisé /100)")
        print(f"    nb_infra           : mean={sum(infra_map.values())/len(infra_map):.1f}")
        print(f"    nb_commerce        : mean={sum(commerce_map.values())/len(commerce_map):.1f}")
        print(f"    Source unique : signaux_immobilier_tunisie.xlsx (0 dépendance satellite)")
        return {
            'score':       score_map,
            'nb_infra':    infra_map,
            'nb_commerce': commerce_map,
        }
    except Exception as e:
        print(f"  [WARN] Signaux erreur : {e} → fallback")
        return {
            'score':       defaults_score,
            'nb_infra':    defaults_infra,
            'nb_commerce': defaults_commerce,
        }


# ================================================================
# GOOGLE MAPS — NOTE MOYENNE
# ================================================================

def load_google_maps_bo3(json_path: str) -> dict:
    """
    Charge raw_data.json → {gouvernorat: note_google_moyenne}.
    Structure attendue : {gov_name: [{note_google: X, total_avis: Y, ...}]}
    """
    result = {}
    if not os.path.exists(json_path):
        print(f"  [WARN] Google Maps JSON absent → note_google_moyenne=0")
        return result
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            raw = json.load(f)
        for gov_name, lieux_list in raw.items():
            if not lieux_list: continue
            gov_norm = normalize_gouvernorat(gov_name)
            if gov_norm == 'Unknown': continue
            notes  = [l.get('note_google', 0) for l in lieux_list if l.get('note_google', 0) > 0]
            result[gov_norm] = round(float(np.mean(notes)), 2) if notes else 0.0
        print(f"  ✔ Google Maps chargé : {len(result)} gouvernorats")
        return result
    except Exception as e:
        print(f"  [WARN] Google Maps erreur : {e}")
        return result


# ================================================================
# SATELLITE — NB_INFRA + NB_COMMERCE
# ================================================================

def load_satellite_bo3(sat_dir: str) -> dict:
    """
    Charge les fichiers satellite CSV → {gouvernorat: {nb_infra, nb_commerce}}.
    Agrège par gouvernorat (pas par ville).
    """
    result = {}
    if not os.path.exists(sat_dir):
        print(f"  [WARN] Satellite dir absent → nb_infra/nb_commerce=0")
        return result
    try:
        files = os.listdir(sat_dir)
        for key, prefix in [('nb_infra', 'tunisia_amenities'),
                             ('nb_commerce', 'tunisia_buildings')]:
            matching = [f for f in files if prefix in f.lower() and f.endswith('.csv')]
            if not matching: continue
            fpath  = os.path.join(sat_dir, max(matching,
                        key=lambda x: os.path.getmtime(os.path.join(sat_dir, x))))
            df_sat = pd.read_csv(fpath)
            if 'city' not in df_sat.columns: continue
            for city in df_sat['city'].unique():
                if pd.isna(city): continue
                gov = normalize_gouvernorat(str(city))
                if gov == 'Unknown': continue
                if gov not in result:
                    result[gov] = {'nb_infra': 0, 'nb_commerce': 0}
                result[gov][key] += len(df_sat[df_sat['city'] == city])
        print(f"  ✔ Satellite chargé : {len(result)} gouvernorats")
        return result
    except Exception as e:
        print(f"  [WARN] Satellite erreur : {e}")
        return result


# ================================================================
# ENTRY POINT
# ================================================================

def load_all(
    bct_path:    str = DEFAULT_PATHS['bct'],
    ins_path:    str = DEFAULT_PATHS['ins'],
    sig_path:    str = DEFAULT_PATHS['signaux'],
    sat_dir:     str = DEFAULT_PATHS['satellite'],
    google_path: str = DEFAULT_PATHS['google_json'],
) -> dict:
    print("\n" + "="*65)
    print("   CHARGEMENT DONNÉES EXTERNES — OBJECTIF 3 : TENDANCES")
    print("="*65)
    bct      = load_bct(bct_path)
    ins      = load_ins_bo3(ins_path)
    signaux  = load_signaux_bo3(sig_path)       # {score, nb_infra, nb_commerce}
    note_g   = load_google_maps_bo3(google_path) # optionnel — pas dans TARGET_COLS
    # Satellite optionnel — nb_infra/nb_commerce déjà dans signaux
    # load_satellite_bo3(sat_dir) non appelé car redondant
    return {
        'bct':                bct,
        'inflation_by_month': ins['inflation_by_month'],
        'pib_by_quarter':     ins['pib_by_quarter'],
        'score_attractivite': signaux['score'],       # {gov: 0.0-1.0}
        'nb_infra':           signaux['nb_infra'],    # {gov: int}  — depuis signaux
        'nb_commerce':        signaux['nb_commerce'], # {gov: int}  — depuis signaux
        'note_google':        note_g,                 # optionnel
    }
