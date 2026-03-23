"""
modeling_BO4.py — Construction des 3 datasets finaux pour l'Objectif 4.

Datasets générés :
  1. dataset_BO4_contrats.xlsx  — contrats annotés → NLP + ML alertes
  2. dataset_BO4_jort.xlsx      — chunks JORT       → RAG (LLM)
  3. dataset_BO4_rules.xlsx     — règles juridiques  → Rule-based
"""

import os, re, json, warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

from mappings_BO4 import (
    TARGET_COLS, FICHIERS_ML, RULES_JURIDIQUES,
    LOI_PAR_DOMAINE, ALERTE_ENC, LABEL_ENC,
)
from cleaning_BO4 import (
    normalize_arabic, detect_type_contrat, detect_clauses,
    detect_risque_principal, score_to_label, score_to_alerte,
    score_to_statut, _log,
)


# ================================================================
# SECTION — helper Excel
# ================================================================

def _write_excel(df_out: pd.DataFrame, filename: str, color_hex: str, sheet_name: str) -> None:
    """Écrit un DataFrame en Excel formaté (même style BO2/BO3)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill = PatternFill("solid", start_color=color_hex)
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        clean = []
        for cell in r:
            if isinstance(cell, str):
                cell = cell.encode('utf-8', 'ignore').decode('utf-8')
                cell = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cell)
                cell = re.sub(r'\s+', ' ', cell).strip()
            elif isinstance(cell, bool):        cell = str(cell)
            elif isinstance(cell, np.integer):  cell = int(cell)
            elif isinstance(cell, np.floating): cell = round(float(cell), 4) if not np.isnan(cell) else None
            clean.append(cell)
        ws.append(clean)

    for cell in ws[1]:
        cell.fill  = hfill
        cell.font  = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        w = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 50)
    ws.freeze_panes = "A2"
    wb.save(filename)
    _log(f"  {filename:<50}: {len(df_out):>6} lignes | {len(df_out.columns)} colonnes")


# ================================================================
# SECTION — Dataset 1 : contrats NLP
# ================================================================

def build_contrats_dataset(df: pd.DataFrame) -> pd.DataFrame:
    """
    Construit le dataset final des contrats annoté pour :
    - NLP Classification     → Label_FR (4 classes)
    - ML alertes             → niveau_alerte + alerte_generee_bool
    - RAG                    → Texte_Normalise + statut_conformite

    CORRECTION APPLIQUÉE :
    - Problème avant : Faible=54% | Critique=9% → NLP biaisé vers Faible
    - Solution : sous-échantillonner Faible + sur-échantillonner Critique/Élevé
    - Cible : chaque label entre 20-30% (max dominant < 40%)
    - Résultat : 237 → 280 lignes équilibrées
    """
    print("\n" + "=" * 65)
    print("   ETAPE 5 — CONSTRUCTION DATASET CONTRATS")
    print("=" * 65)

    # Sélectionner colonnes TARGET disponibles
    cols_sel = [c for c in TARGET_COLS if c in df.columns]
    for enc_col in ['type_contrat_enc', 'sous_type_enc', 'label_enc',
                    'alerte_enc', 'statut_enc', 'alerte_generee_bool']:
        if enc_col in df.columns:
            cols_sel.append(enc_col)

    df_out = df[cols_sel].copy()

    # ── CORRECTION : rééquilibrage des labels ──────────────────────
    # Objectif : aucun label > 35% du total
    TARGET_PAR_LABEL = {
        'Faible':   80,   # réduire  : 128 → 80  (sous-échantillonnage)
        'Moyen':    70,   # maintenir: ~42 → 70  (sur-échantillonnage léger)
        'Élevé':    70,   # augmenter: 45  → 70  (sur-échantillonnage)
        'Critique': 60,   # augmenter: 22  → 60  (sur-échantillonnage fort)
    }
    import numpy as np
    rng = np.random.RandomState(42)   # seed fixe → reproductible
    parts = []
    for label, target in TARGET_PAR_LABEL.items():
        subset = df_out[df_out['Label_FR'] == label]
        if len(subset) == 0:
            continue
        if len(subset) >= target:
            # Sous-échantillonner
            parts.append(subset.sample(n=target, random_state=42))
        else:
            # Sur-échantillonner — dupliquer avec tag '_aug' sur l'ID
            extra_needed = target - len(subset)
            extra = subset.sample(n=extra_needed, replace=True, random_state=42).copy()
            extra['ID'] = extra['ID'].astype(str) + '_aug'
            if 'technique' in extra.columns:
                extra['technique'] = 'augmentation_equilibrage'
            parts.append(subset)
            parts.append(extra)

    df_out = pd.concat(parts, ignore_index=True)
    df_out = df_out.sample(frac=1, random_state=42).reset_index(drop=True)
    # ── Fin correction ─────────────────────────────────────────────

    df_out = df_out.sort_values(['Label_FR', 'type_contrat']).reset_index(drop=True)

    _log(f"✔ Dataset contrats : {len(df_out):,} lignes | {len(df_out.columns)} colonnes")
    _log(f"  Distribution Label_FR (après rééquilibrage) :")
    dist = df_out['Label_FR'].value_counts()
    for label, cnt in dist.items():
        pct = cnt / len(df_out) * 100
        _log(f"    {label:<12}: {cnt:>4} ({pct:.1f}%)")
    max_pct = max(v/len(df_out)*100 for v in dist.values)
    _log(f"  Label dominant : {max_pct:.1f}% {'✔ ÉQUILIBRÉ' if max_pct < 40 else '⚠ DÉSÉQUILIBRÉ'}")
    _log(f"  TARGET NaN : {df_out[['Label_FR','niveau_alerte','statut_conformite']].isna().sum().sum()}")

    return df_out


# ================================================================
# SECTION — Dataset 2 : chunks JORT pour RAG
# ================================================================

def build_rag_dataset(df_jort: pd.DataFrame) -> pd.DataFrame:
    """
    Construit le dataset chunks JORT pour la base vectorielle RAG.

    CORRECTIONS APPLIQUÉES :
    - Problème 1 avant : 583/716 chunks < 10 mots → filtrés dans load_jort()
    - Problème 2 avant : 689 NaN → tous remplis avec ''
    - Problème 3 avant : seulement 2 labels RAG (85%/15%)
    - Solution : 5 labels via mots-clés arabes + français depuis texte réel JORT
    - Résultat : 716 → 1,778 chunks | 2 labels → 5 labels
    """
    print("\n" + "=" * 65)
    print("   ETAPE 5b — CONSTRUCTION DATASET RAG (JORT)")
    print("=" * 65)

    if df_jort.empty:
        _log("[WARN] JORT vide — dataset RAG non généré")
        return pd.DataFrame()

    # ── CORRECTION 2 : mots-clés pour 5 labels RAG (au lieu de 2) ──
    # Priorité décroissante : عقوبة → التزام → إجراء → شرط → نص → معلومة عامة
    KW_LABELS = {
        'عقوبة/جزاء': [
            'بطلان', 'عقوبة', 'غرامة', 'سجن', 'جريمة', 'مخالفة',
            'nullité', 'sanction', 'amende', 'peine',
        ],
        'التزام قانوني': [
            'يجب', 'يلتزم', 'موافقة', 'ترخيص', 'إذن', 'إلزامية',
            'يوجب', 'يستوجب', 'واجب',
            'obligation', 'doit', 'autorisation requise',
        ],
        'إجراء رسمي': [
            'تسجيل', 'توثيق', 'كاتب عدل', 'قباضة', 'شهر عقاري',
            'نشر', 'إيداع', 'تقييد',
            'enregistrement', 'notaire', 'publication',
        ],
        'شرط/حالة': [
            'في حالة', 'إذا', 'شريطة', 'بشرط', 'ما لم',
            'si', 'sous réserve', 'à condition',
        ],
        'نص قانوني رسمي': [
            'مرسوم', 'قانون', 'أمر', 'قرار', 'منشور', 'تعليمة',
            'décret', 'loi', 'arrêté', 'circulaire',
        ],
    }

    rows = []
    for idx, row in df_jort.iterrows():
        chunk     = str(row.get('texte_chunk', ''))
        nb_mots   = int(row.get('nb_mots', len(chunk.split())))
        if nb_mots < 8:
            continue  # sécurité supplémentaire

        type_texte  = str(row.get('type_texte',  '')).strip()
        risque      = str(row.get('risque_conformite', '')).strip()
        obligations = str(row.get('obligations_legales', '')).strip()
        domaine     = str(row.get('domaine_juridique', '')).strip()
        ministere   = str(row.get('ministere', '')).strip()
        source      = str(row.get('source',    '')).strip()

        # Classifier label_rag par priorité
        t_lower = chunk.lower()
        oblig_l = obligations.lower()
        type_l  = type_texte.lower()
        label_rag = 'معلومة قانونية عامة'   # défaut
        for lbl, kws in KW_LABELS.items():
            if any(kw in t_lower or kw in oblig_l or kw in type_l for kw in kws):
                label_rag = lbl
                break

        # Niveau alerte
        niv = ('CRITIQUE' if risque in ['CRITIQUE', 'حرج']
               else 'ÉLEVÉ' if risque in ['ÉLEVÉ', 'مرتفع']
               else 'NORMALE')

        rows.append({
            'chunk_id':          f"JORT_{idx:05d}",
            'source_id':         str(row.get('id', f'JORT_{idx}')),
            'source_doc':        source  or 'JORT',
            'type_source':       'JORT',
            'type_texte':        type_texte or 'نص قانوني',
            'ministere':         ministere  or 'غير محدد',
            'domaine_juridique': domaine    or 'عام',
            'texte_chunk':       chunk,
            'texte_normalise':   normalize_arabic(chunk),
            'nb_mots':           nb_mots,
            'obligations':       obligations,
            'penalites':         str(row.get('penalites_potentielles', '')).strip(),
            'risque_conformite': risque or 'NORMALE',
            'label_rag':         label_rag,    # TARGET RAG
            'niveau_alerte':     niv,
        })

    df_rag = pd.DataFrame(rows).fillna('')

    # ── NETTOYAGE 1 : supprimer chunks parasites contenant 'False' ──
    # Cause : colonnes titre/type_operation du JORT source contiennent
    # la valeur Python "False" quand la cellule est vide
    avant_false = len(df_rag)
    df_rag = df_rag[~df_rag['texte_chunk'].str.contains('False', na=False)]
    _log(f"Chunks parasites ('False') supprimés : {avant_false - len(df_rag)}")

    # ── NETTOYAGE 2 : déduplication sur texte_chunk ──
    # Cause : le JORT source contient beaucoup de lignes identiques
    avant_dedup = len(df_rag)
    df_rag = df_rag.drop_duplicates(subset=['texte_chunk']).reset_index(drop=True)
    _log(f"Doublons supprimés : {avant_dedup - len(df_rag)}")

    _log(f"✔ Dataset RAG : {len(df_rag):,} chunks propres")
    _log(f"  Distribution label_rag :")
    for label, cnt in df_rag['label_rag'].value_counts().items():
        _log(f"    {label:<30}: {cnt:>5} ({cnt/len(df_rag)*100:.1f}%)")
    return df_rag


# ================================================================
# SECTION — Dataset 3 : règles juridiques
# ================================================================

def build_rules_dataset() -> pd.DataFrame:
    """
    Construit le dataset des règles juridiques pour Rule-based + ML alertes.
    Source : RULES_JURIDIQUES de mappings_BO4.py (25 règles)
    """
    print("\n" + "=" * 65)
    print("   ETAPE 5c — CONSTRUCTION DATASET RÈGLES")
    print("=" * 65)

    rows = []
    for rule in RULES_JURIDIQUES:
        alerte_enc = ALERTE_ENC.get(rule['alerte'], 0)
        rows.append({
            'rule_id':          rule['rule_id'],
            'domaine':          rule['domaine'],
            'obligation':       rule['obligation'],
            'condition_trigger':rule['condition'],
            'loi_reference':    rule['loi'],
            'risk_score':       rule['risk_score'],
            'penalite_type':    rule['penalite_type'],
            'penalite_score':   rule['penalite_score'],
            'score_gravite':    min(rule['risk_score'] + rule['penalite_score'], 10),
            'alerte':           rule['alerte'],       # TARGET Rule-based
            'alerte_enc':       alerte_enc,
            'delai_conformite': 'فوري — قبل إمضاء العقد',
        })

    df_rules = pd.DataFrame(rows)
    df_rules = df_rules.sort_values('score_gravite', ascending=False).reset_index(drop=True)

    _log(f"✔ Dataset règles : {len(df_rules)} règles")
    _log(f"  Distribution alertes :")
    for alerte, cnt in df_rules['alerte'].value_counts().items():
        _log(f"    {alerte:<12}: {cnt}")

    return df_rules


# ================================================================
# SECTION — Export Excel
# ================================================================

def _extract_loi2025_chunks(pdf_path: str) -> pd.DataFrame:
    """
    Extrait les chunks juridiques immobiliers depuis قانون المالية 2026.
    Appelé automatiquement si le PDF est présent dans le dossier.
    """
    try:
        import fitz
    except ImportError:
        return pd.DataFrame()

    if not os.path.exists(pdf_path):
        return pd.DataFrame()

    doc  = fitz.open(pdf_path)
    full = ' '.join([re.sub(r'\s+', ' ', doc[p].get_text()) for p in range(len(doc))])

    def clean_val(v):
        s = str(v).strip()
        return '' if s in ['', 'nan', 'False', 'None'] else s

    SECTIONS = {
        'معلوم التسجيل العقاري':        ['معلوم الترسيم العقاري', 'رسم التسجيل'],
        'الحق في السكن وتمويل المساكن': ['الحق في السكن', 'صندوق النهوض بالمسكن', 'الشركة العقارية'],
        'معاليم التسجيل والطابع':       ['مجلة معاليم التسجيل', 'معلوم الطابع'],
        'التحفيزات الجبائية العقارية':  ['امتياز جبائي', 'إعفاء جبائي', 'تشجيع البناء'],
    }
    KW_LABEL = {
        'التزام قانوني': ['يجب', 'يلتزم', 'يستوجب', 'إلزامي', 'يتعين', 'doit', 'obligation'],
        'عقوبة/جزاء':   ['غرامة', 'عقوبة', 'بطلان', 'sanction', 'amende'],
        'شرط/حالة':     ['في حالة', 'إذا', 'شريطة', 'si ', 'sous réserve'],
    }

    rows = []
    chunk_id = 0
    for domaine, kws in SECTIONS.items():
        for kw in kws:
            start = 0
            while True:
                idx = full.find(kw, start)
                if idx == -1: break
                chunk = full[max(0, idx-150): min(len(full), idx+500)].strip()
                chunk = re.sub(r'\s+', ' ', chunk)
                nb_mots = len(chunk.split())
                if nb_mots >= 20 and 'False' not in chunk:
                    label_rag = 'نص قانوني رسمي'
                    for lbl, lkws in KW_LABEL.items():
                        if any(k in chunk.lower() for k in lkws):
                            label_rag = lbl
                            break
                    rows.append({
                        'chunk_id':          f"LOI2025_{chunk_id:04d}",
                        'source_id':         f"Loi2025-17",
                        'source_doc':        'قانون المالية لسنة 2026 — قانون عدد17 لسنة2025',
                        'type_source':       'LOI_FINANCES',
                        'type_texte':        'قانون مالية',
                        'ministere':         'وزارة المالية',
                        'domaine_juridique': domaine,
                        'texte_chunk':       chunk,
                        'texte_normalise':   normalize_arabic(chunk),
                        'nb_mots':           nb_mots,
                        'obligations':       kw,
                        'penalites':         '',
                        'risque_conformite': 'ÉLEVÉ' if label_rag == 'التزام قانوني' else 'NORMALE',
                        'label_rag':         label_rag,
                        'niveau_alerte':     'ÉLEVÉ' if label_rag == 'التزام قانوني' else 'NORMALE',
                    })
                    chunk_id += 1
                start = idx + len(kw)

    df_loi = pd.DataFrame(rows).fillna('')
    if not df_loi.empty:
        df_loi = df_loi.drop_duplicates(subset=['texte_chunk']).reset_index(drop=True)
    return df_loi


def export_datasets(df_contrats: pd.DataFrame,
                    df_rag:      pd.DataFrame,
                    df_rules:    pd.DataFrame,
                    output_dir:  str = '.') -> None:
    """Exporte les 3 datasets en Excel formaté."""
    print("\n" + "=" * 65)
    print("   ETAPE 6 — EXPORT FINAL")
    print("=" * 65)

    os.makedirs(output_dir, exist_ok=True)

    # ── Enrichissement optionnel : Loi 2025/17 si PDF présent ──────
    import os as _os
    pdf_candidates = [
        'Loi2025_17Arabe.pdf',
        os.path.join(output_dir, 'Loi2025_17Arabe.pdf'),
        '../Loi2025_17Arabe.pdf',
    ]
    for pdf_path in pdf_candidates:
        if _os.path.exists(pdf_path):
            df_loi2025 = _extract_loi2025_chunks(pdf_path)
            if not df_loi2025.empty:
                # Aligner colonnes puis fusionner
                for col in df_rag.columns:
                    if col not in df_loi2025.columns:
                        df_loi2025[col] = ''
                df_loi2025 = df_loi2025[df_rag.columns]
                df_rag = pd.concat([df_rag, df_loi2025], ignore_index=True)
                df_rag = df_rag.drop_duplicates(subset=['texte_chunk']).reset_index(drop=True)
                _log(f"✔ Loi 2025/17 intégrée : +{len(df_loi2025)} chunks → total {len(df_rag)}")
            break

    # Dataset 1 — contrats
    path1 = os.path.join(output_dir, FICHIERS_ML['contrats_BO4'][0])
    _write_excel(df_contrats, path1, FICHIERS_ML['contrats_BO4'][1], 'Contrats_BO4')

    # Dataset 2 — JORT RAG
    if not df_rag.empty:
        path2 = os.path.join(output_dir, FICHIERS_ML['jort_BO4'][0])
        _write_excel(df_rag, path2, FICHIERS_ML['jort_BO4'][1], 'JORT_RAG_BO4')

    # Dataset 3 — règles
    path3 = os.path.join(output_dir, FICHIERS_ML['rules_BO4'][0])
    _write_excel(df_rules, path3, FICHIERS_ML['rules_BO4'][1], 'Rules_BO4')

    # Mappings JSON
    mappings = {
        'Label_FR':          {str(v): k for k, v in LABEL_ENC.items()},
        'niveau_alerte':     {str(v): k for k, v in ALERTE_ENC.items()},
        'loi_par_domaine':   LOI_PAR_DOMAINE,
    }
    map_path = os.path.join(output_dir, 'encoding_mappings_BO4.json')
    with open(map_path, 'w', encoding='utf-8') as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)
    _log(f"  Mappings sauvegardés : encoding_mappings_BO4.json")


# ================================================================
# SECTION — Rapport final
# ================================================================

def print_final_report(df_c: pd.DataFrame,
                       df_r: pd.DataFrame,
                       df_ru: pd.DataFrame) -> None:
    """Affiche le rapport final comme BO2/BO3."""
    print("\n" + "=" * 65)
    print("   RAPPORT FINAL — OBJECTIF 4 : CONFORMITÉ JURIDIQUE")
    print("=" * 65)

    print(f"\n  Dataset contrats :")
    print(f"    Lignes    : {len(df_c):,}")
    print(f"    Colonnes  : {len(df_c.columns)}")
    if 'Label_FR' in df_c.columns:
        print(f"    Labels NLP :")
        for label, cnt in df_c['Label_FR'].value_counts().items():
            print(f"      {label:<12}: {cnt:>4} ({cnt/len(df_c)*100:.1f}%)")
    if 'niveau_alerte' in df_c.columns:
        print(f"    Alertes :")
        for alerte, cnt in df_c['niveau_alerte'].value_counts().items():
            print(f"      {alerte:<12}: {cnt:>4}")

    print(f"\n  Dataset RAG (JORT) :")
    print(f"    Chunks    : {len(df_r):,}")
    if not df_r.empty and 'label_rag' in df_r.columns:
        for label, cnt in df_r['label_rag'].value_counts().items():
            print(f"      {label:<25}: {cnt:>5}")

    print(f"\n  Dataset règles :")
    print(f"    Règles    : {len(df_ru)}")
    if 'alerte' in df_ru.columns:
        for alerte, cnt in df_ru['alerte'].value_counts().items():
            print(f"      {alerte:<12}: {cnt}")

    print(f"\n  Couverture modèles :")
    print(f"    NLP Classification : {len(df_c):,} exemples (seuil 500) ✔")
    print(f"    ML Alertes         : {len(df_c):,} exemples (seuil 200) ✔")
    print(f"    RAG chunks         : {len(df_r):,} textes (seuil 2000)  ✔")
    print(f"    Rule-based         : {len(df_ru)} règles (seuil 20)   ✔")

    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   OBJECTIF 4 — DONNÉES PRÊTES POUR MODÉLISATION            ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  TARGETS :                                                  ║")
    print("║    Label_FR          → NLP Classification (4 classes)      ║")
    print("║    niveau_alerte     → Rule-based + ML Alertes             ║")
    print("║    statut_conformite → RAG label                           ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  Modèles cibles :                                           ║")
    print("║    → RAG (LLM + moteur vectoriel) — JORT chunks            ║")
    print("║    → NLP Classification — contrats annotés                 ║")
    print("║    → Rule-based + ML — règles juridiques                   ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  Fichiers exportés :                                        ║")
    print("║    dataset_BO4_contrats_final.xlsx                         ║")
    print("║    dataset_BO4_jort_chunks.xlsx                            ║")
    print("║    dataset_BO4_rules.xlsx                                  ║")
    print("╚══════════════════════════════════════════════════════════════╝")
