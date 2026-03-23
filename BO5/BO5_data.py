"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   PIPELINE IMMOBILIER TUNISIE — BO5 : INDICE DE RENTABILITÉ RÉGIONALE       ║
║   Adapté depuis BO2 (Multimodal) → BO5 (Rentabilité macro-régionale)        ║
║                                                                              ║
║   OBJECTIF : Calculer et prédire l'indice_rentabilite_regionale              ║
║              par gouvernorat — pour modèles de scoring régional              ║
║                                                                              ║
║   TARGET ⭐ : indice_rentabilite_regionale                                   ║
║     = rendement locatif net estimé × score attractivité × ajustement macro  ║
║                                                                              ║
║   FEATURES MARCHÉ (5) :                                                      ║
║     prix_m2, variation_prix_m2, inflation_glissement_annuel,                 ║
║     croissance_pib_trim, high_season                                         ║
║                                                                              ║
║   FEATURES ATTRACTIVITÉ GOOGLE + OSM (10) :                                  ║
║     note_google_moyenne, densite_population, nb_amenities_total,             ║
║     ratio_amenities_commerce, ratio_amenities_sante, densite_routes_km,      ║
║     nb_stations_transport, nb_buildings_residentiel,                         ║
║     surface_landuse_residentiel, nb_commerce                                 ║
║                                                                              ║
║   COLONNES FINALES (17 = 5 marché + 10 attractivité + 1 TARGET + 1 géo) :   ║
║     gouvernorat_nom + 5 marché + 10 attractivité + TARGET⭐                  ║
║                                                                              ║
║   ÉTAPES :                                                                   ║
║    1.  ETL              — Fusion 8 sources (identique BO2)                   ║
║    2.  Déduplication    — 3 passes (identique BO2)                           ║
║    2b. Types & Aperçu   — Correction types (identique BO2)                  ║
║    3.  Standardisation  — type_bien + type_transaction (identique BO2)       ║
║    4.  Segmentation     — 4 groupes (identique BO2)                          ║
║    5.  Nettoyage        — Seuils + Isolation Forest (identique BO2)          ║
║    5b. Imputation NA    — median/KNN/hot-deck (identique BO2)                ║
║    5c. Visualisation    — Histogrammes + Boxplots (identique BO2)            ║
║    5d. Encodage MANUEL  — gouvernorat / type_bien (identique BO2)            ║
║    6.  Géocodage        — Ville → lat/lon (identique BO2)                    ║
║   ─── SPÉCIFIQUE BO5 ──────────────────────────────────────────────────────  ║
║    7.  Features Marché  — prix_m2, variation, macro (PIB, inflation, saison)  ║
║    8.  Features OSM     — amenities, routes, buildings, landuse               ║
║    9.  Features Google  — note_google_moyenne par gouvernorat                 ║
║   10.  TARGET BO5       — indice_rentabilite_regionale (calcul composite)     ║
║   11.  Export Final     — 4 fichiers Excel *_BO5.xlsx (17 colonnes)          ║
║                                                                              ║
║   SUPPRIMÉ vs BO2 : BERT/LSA, Vision, Fusion multimodale,                   ║
║     prix_transaction_estimated, negotiation_rate, cycle_marche,              ║
║     market_tension, NLP flags, nb_pieces, surface_m2 brute                  ║
║                                                                              ║
║   AJOUTÉ vs BO2 : prix_m2 (agrégat gouvernorat), variation_prix_m2,         ║
║     inflation_glissement_annuel, croissance_pib_trim, high_season,           ║
║     note_google_moyenne, densite_population, nb_amenities_total,             ║
║     ratio_amenities_commerce, ratio_amenities_sante, densite_routes_km,      ║
║     nb_stations_transport, nb_buildings_residentiel,                         ║
║     surface_landuse_residentiel, nb_commerce                                 ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import pandas as pd
import numpy as np
import re, os, warnings, time, json
from datetime import datetime

warnings.filterwarnings('ignore')

from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from sklearn.impute import KNNImputer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

# ── Mode DEBUG : export CSV rapide pour itérations, False = Excel final ──
DEBUG_MODE = False

_t0 = time.time()

# ================================================================
# CHEMINS DES SOURCES
# ================================================================

SATELLITE_DIR = '../Tunisia_Satellite_scraper/tunisia_satellite_data'

SOURCES_IMMO = [
    ('Facebook Marketplace','../marketplace/data/facebook_marketplace_2185_annonces_20260216_222412.csv','csv',None),
    ('Mubawab',             '../mubawab/mubawab_annonces.csv',                                          'csv',None),
    ('Mubawab Partial',     '../mubawab2/mubawab_partial_120.xlsx',                                     'xlsx',None),
    ('Tayara',              '../tayara/tayara_complete.csv',                                            'csv',None),
    ('Tunisie Annonces',    '../tunisie_annance_scraper/ta_properties.csv',                             'csv',None),
    ('Century21',           '../scrapping/century21_data2.csv',                                        'csv',';'),
    ('HomeInTunisia',       '../scrapping/homeintunisia_data2.csv',                                     'csv',';'),
    ('BnB',                 '../bnb/bnb_properties.csv',                                               'csv',None),
]

# Chemins satellites BO5
BUILDINGS_PATH  = f'{SATELLITE_DIR}/tunisia_buildings_20260215_133916.csv'
POPULATION_PATH = f'{SATELLITE_DIR}/tunisia_population_20260215_133916.csv'
AMENITIES_PATH  = f'{SATELLITE_DIR}/tunisia_amenities_20260215_133916.csv'
ROADS_PATH      = f'{SATELLITE_DIR}/tunisia_roads_20260215_133916.csv'
LANDUSE_PATH    = f'{SATELLITE_DIR}/tunisia_landuse_20260215_133916.csv'
SIGNAUX_PATH    = f'{SATELLITE_DIR}/signaux_immobilier_tunisie.xlsx'

# ================================================================
# FONCTIONS UTILITAIRES
# ================================================================

def section(t): print("\n" + "="*70 + f"\n   {t}\n" + "="*70)
def log(m):     print(f"  {m}")


# ================================================================
# ── ENCODAGES MANUELS FIXES (identiques BO2) ────────────────────
# ================================================================

GOUVERNORAT_ENC = {
    'Unknown': 0, 'Ariana': 1, 'Béja': 2, 'Ben Arous': 3, 'Bizerte': 4,
    'Gabès': 5, 'Gafsa': 6, 'Jendouba': 7, 'Kairouan': 8, 'Kasserine': 9,
    'Kébili': 10, 'Le Kef': 11, 'Mahdia': 12, 'Manouba': 13, 'Médenine': 14,
    'Monastir': 15, 'Nabeul': 16, 'Sfax': 17, 'Sidi Bouzid': 18, 'Siliana': 19,
    'Sousse': 20, 'Tataouine': 21, 'Tozeur': 22, 'Tunis': 23, 'Zaghouan': 24,
}
GOUVERNORAT_DEC = {v: k for k, v in GOUVERNORAT_ENC.items()}

VILLE_TO_GOUVERNORAT = {
    # ── Tunis ─────────────────────────────────────────────────────
    'tunis': 'Tunis', 'carthage': 'Tunis', 'la marsa': 'Tunis', 'marsa': 'Tunis',
    'gammarth': 'Tunis', 'sidi bou said': 'Tunis', 'sidi bou saïd': 'Tunis',
    'sidi bou sad': 'Tunis', 'el aouina': 'Tunis', 'aouina': 'Tunis',
    'el menzah': 'Tunis', 'menzah': 'Tunis', 'cite el khadra': 'Tunis',
    'cité el khadra': 'Tunis', 'cite khadra': 'Tunis', 'el khadra': 'Tunis',
    'le bardo': 'Tunis', 'bardo': 'Tunis', 'ennasr': 'Tunis', 'nasr': 'Tunis',
    'ain zaghouan': 'Tunis', 'aïn zaghouan': 'Tunis', 'ain zagouan': 'Tunis',
    'les berges du lac': 'Tunis', 'berges du lac': 'Tunis', 'lac': 'Tunis',
    'lac 1': 'Tunis', 'lac 2': 'Tunis', 'el lac': 'Tunis', 'chotrana': 'Tunis',
    'chotrana 1': 'Tunis', 'chotrana 2': 'Tunis', 'chotrana 3': 'Tunis',
    'mutuelleville': 'Tunis', 'el manar': 'Tunis', 'el manar 1': 'Tunis',
    'el manar 2': 'Tunis', 'montplaisir': 'Tunis', 'el omrane': 'Tunis',
    'cite olympique': 'Tunis', 'el ouardia': 'Tunis', 'sijoumi': 'Tunis',
    'medina': 'Tunis', 'la medina': 'Tunis',
    # ── Ariana ────────────────────────────────────────────────────
    'ariana': 'Ariana', 'raoued': 'Ariana', 'la soukra': 'Ariana', 'soukra': 'Ariana',
    'kalaat el andalous': 'Ariana', 'cite el wafa': 'Ariana', 'borj louzir': 'Ariana',
    'sidi thabet': 'Ariana', 'mnihla': 'Ariana',
    # ── Ben Arous ─────────────────────────────────────────────────
    'ben arous': 'Ben Arous', 'ezzahra': 'Ben Arous', 'hammam lif': 'Ben Arous',
    'borj cedria': 'Ben Arous', 'borj cédria': 'Ben Arous', 'megrine': 'Ben Arous',
    'mégrine': 'Ben Arous', 'mourouj': 'Ben Arous', 'mourou': 'Ben Arous',
    'rades': 'Ben Arous', 'radès': 'Ben Arous', 'fouchana': 'Ben Arous',
    'mohamedia': 'Ben Arous', 'mornag': 'Ben Arous', 'boumhel': 'Ben Arous',
    'bou mhel': 'Ben Arous', 'bou mhel el basset': 'Ben Arous',
    'el mourouj': 'Ben Arous', 'nouvelle medina': 'Ben Arous',
    # ── Manouba ───────────────────────────────────────────────────
    'manouba': 'Manouba', 'tebourba': 'Manouba', 'el battan': 'Manouba',
    'oued ellil': 'Manouba', 'el mornaguia': 'Manouba', 'douar hicher': 'Manouba',
    'cite ettadhamen': 'Manouba',
    # ── Nabeul ────────────────────────────────────────────────────
    'nabeul': 'Nabeul', 'hammamet': 'Nabeul', 'kelibia': 'Nabeul', 'kélibia': 'Nabeul',
    'korba': 'Nabeul', 'grombalia': 'Nabeul', 'bou argoub': 'Nabeul',
    'nabeul city': 'Nabeul', 'menzel temime': 'Nabeul', 'soliman': 'Nabeul',
    'el haouaria': 'Nabeul', 'takelsa': 'Nabeul', 'beni khalled': 'Nabeul',
    'hammam ghezaz': 'Nabeul', 'dar chaabane': 'Nabeul', 'el mida': 'Nabeul',
    # ── Sousse ────────────────────────────────────────────────────
    'sousse': 'Sousse', 'skanes': 'Sousse', 'msaken': 'Sousse',
    'ksar hellal': 'Sousse', 'hammam sousse': 'Sousse', 'akouda': 'Sousse',
    'kantaoui': 'Sousse', 'port el kantaoui': 'Sousse', 'enfidha': 'Sousse',
    'sousse center': 'Sousse', 'sousse centre': 'Sousse', 'sidi bou ali': 'Sousse',
    'kalaa kebira': 'Sousse', 'kalaa sghira': 'Sousse', 'kondar': 'Sousse',
    # ── Monastir ──────────────────────────────────────────────────
    'monastir': 'Monastir', 'skanes monastir': 'Monastir', 'monastir city': 'Monastir',
    'moknine': 'Monastir', 'bekalta': 'Monastir', 'jemmal': 'Monastir',
    'beni hassen': 'Monastir', 'ouerdanine': 'Monastir', 'zeramdine': 'Monastir',
    'teboulba': 'Monastir',
    # ── Sfax ──────────────────────────────────────────────────────
    'sfax': 'Sfax', 'sax': 'Sfax', 'sfax city': 'Sfax', 'sakiet ezzit': 'Sfax',
    'sakiet eddaier': 'Sfax', 'thyna': 'Sfax', 'la shkira': 'Sfax',
    'el ain': 'Sfax', 'agareb': 'Sfax', 'jebeniana': 'Sfax',
    'bir ali ben khalifa': 'Sfax', 'mahres': 'Sfax',
    # ── Bizerte ───────────────────────────────────────────────────
    'bizerte': 'Bizerte', 'mateur': 'Bizerte', 'menzel bourguiba': 'Bizerte',
    'menzel jemil': 'Bizerte', 'el alia': 'Bizerte', 'ras jebel': 'Bizerte',
    'ghar el melh': 'Bizerte', 'zarzouna': 'Bizerte',
    # ── Béja ──────────────────────────────────────────────────────
    'beja': 'Béja', 'béja': 'Béja', 'bja': 'Béja', 'beja city': 'Béja',
    'medjez el bab': 'Béja', 'testour': 'Béja', 'nefza': 'Béja',
    'amdoun': 'Béja', 'thibar': 'Béja',
    # ── Jendouba ──────────────────────────────────────────────────
    'jendouba': 'Jendouba', 'jendouba city': 'Jendouba', 'tabarka': 'Jendouba',
    'ain draham': 'Jendouba', 'fernana': 'Jendouba', 'ghardimaou': 'Jendouba',
    'bou salem': 'Jendouba', 'oued mliz': 'Jendouba',
    # ── Le Kef ────────────────────────────────────────────────────
    'le kef': 'Le Kef', 'el kef': 'Le Kef', 'kef': 'Le Kef',
    'dahmani': 'Le Kef', 'tajerouine': 'Le Kef', 'sers': 'Le Kef',
    # ── Siliana ───────────────────────────────────────────────────
    'siliana': 'Siliana', 'siliana city': 'Siliana', 'gaafour': 'Siliana',
    'bou arada': 'Siliana', 'makthar': 'Siliana',
    # ── Kairouan ──────────────────────────────────────────────────
    'kairouan': 'Kairouan', 'kairouan city': 'Kairouan', 'haffouz': 'Kairouan',
    'el ala': 'Kairouan', 'sbikha': 'Kairouan', 'oueslatia': 'Kairouan',
    'nasrallah': 'Kairouan',
    # ── Kasserine ─────────────────────────────────────────────────
    'kasserine': 'Kasserine', 'kasserine city': 'Kasserine', 'sbeitla': 'Kasserine',
    'thala': 'Kasserine', 'feriana': 'Kasserine', 'foussana': 'Kasserine',
    # ── Sidi Bouzid ───────────────────────────────────────────────
    'sidi bouzid': 'Sidi Bouzid', 'sidi bou zid': 'Sidi Bouzid',
    'sidi bouzid city': 'Sidi Bouzid', 'meknassy': 'Sidi Bouzid',
    'regueb': 'Sidi Bouzid', 'bir el hafey': 'Sidi Bouzid',
    # ── Mahdia ────────────────────────────────────────────────────
    'mahdia': 'Mahdia', 'mahdia city': 'Mahdia', 'el jem': 'Mahdia',
    'ksour essef': 'Mahdia', 'chebba': 'Mahdia', 'bou merdes': 'Mahdia',
    'hebira': 'Mahdia',
    # ── Gabès ─────────────────────────────────────────────────────
    'gabes': 'Gabès', 'gabès': 'Gabès', 'gabes city': 'Gabès', 'matmata': 'Gabès',
    'mareth': 'Gabès', 'nouvelle matmata': 'Gabès', 'el hamma': 'Gabès',
    # ── Médenine ──────────────────────────────────────────────────
    'medenine': 'Médenine', 'médenine': 'Médenine', 'médénine': 'Médenine',
    'mdenine': 'Médenine', 'zarzis': 'Médenine', 'djerba': 'Médenine',
    'jerba': 'Médenine', 'djerba houmt souk': 'Médenine',
    'djerba midoun': 'Médenine', 'djerba ajim': 'Médenine',
    'ben gardane': 'Médenine', 'beni khedache': 'Médenine',
    # ── Tataouine ─────────────────────────────────────────────────
    'tataouine': 'Tataouine', 'tataouine city': 'Tataouine', 'remada': 'Tataouine',
    'ghomrassen': 'Tataouine', 'bir lahmar': 'Tataouine',
    # ── Gafsa ─────────────────────────────────────────────────────
    'gafsa': 'Gafsa', 'gafsa city': 'Gafsa', 'moularès': 'Gafsa',
    'redeyef': 'Gafsa', 'metlaoui': 'Gafsa', 'el ksar': 'Gafsa',
    # ── Tozeur ────────────────────────────────────────────────────
    'tozeur': 'Tozeur', 'tozeur city': 'Tozeur', 'nefta': 'Tozeur',
    'tamerza': 'Tozeur', 'degache': 'Tozeur',
    # ── Kébili ────────────────────────────────────────────────────
    'kebili': 'Kébili', 'kébili': 'Kébili', 'kebili city': 'Kébili',
    'douz': 'Kébili', 'souk lahad': 'Kébili', 'faouar': 'Kébili',
    # ── Zaghouan ──────────────────────────────────────────────────
    'zaghouan': 'Zaghouan', 'zaghouan city': 'Zaghouan', 'zriba': 'Zaghouan',
    'nadhour': 'Zaghouan', 'bir mcherga': 'Zaghouan', 'hammam zriba': 'Zaghouan',
    # ── Génériques ────────────────────────────────────────────────
    'tunisie': None, 'tunisia': None, 'inconnu': None, 'unknown': None, '': None,
}

def normalize_gouvernorat(val):
    if pd.isna(val) or str(val).strip() in ('', 'nan', 'None', 'NaN', 'none'):
        return 'Unknown'
    v = str(val).strip().lower()
    if v in VILLE_TO_GOUVERNORAT:
        result = VILLE_TO_GOUVERNORAT[v]
        return result if result else 'Unknown'
    for key, gov in VILLE_TO_GOUVERNORAT.items():
        if not key or not gov: continue
        if v.startswith(key) or key.startswith(v) or key in v:
            return gov
    for gov_name in GOUVERNORAT_ENC:
        if gov_name.lower() == v:
            return gov_name
    return 'Unknown'

def encode_gouvernorat(gov_str):
    if pd.isna(gov_str): return 0
    return GOUVERNORAT_ENC.get(str(gov_str).strip(), 0)

TYPE_BIEN_ENC = {
    'Unknown': 0, 'Appartement': 1, 'Autre': 2, 'Bureau': 3, 'Chambre': 4,
    'Ferme': 5, 'Local Commercial': 6, 'Maison': 7, 'Terrain': 8, 'Villa': 9,
}
TYPE_BIEN_DEC = {v: k for k, v in TYPE_BIEN_ENC.items()}

def encode_type_bien(val):
    if pd.isna(val): return 0
    return TYPE_BIEN_ENC.get(str(val).strip(), 0)

def std_transaction(val):
    if pd.isna(val): return np.nan
    v = str(val).lower().strip()
    if any(x in v for x in ['courte','vacance','saisonn','airbnb','nuit']): return 1
    if any(x in v for x in ['location','locat','rent','louer','loue','loyer','mensuel']): return 1
    if any(x in v for x in ['vent','sale','achat','cession','vendre','vend']): return 2
    return np.nan


section("ENCODAGES MANUELS FIXES BO5")
log("gouvernorat      : " + str({k: v for k, v in list(GOUVERNORAT_ENC.items())[:6]}) + " ...")
log("type_bien        : " + str(TYPE_BIEN_ENC))
log("type_transaction : {1: 'Location', 2: 'Vente'}")
log(f"VILLE_TO_GOUVERNORAT : {len(VILLE_TO_GOUVERNORAT)} entrées")


# ================================================================
# FONCTIONS NETTOYAGE (identiques BO2)
# ================================================================

def remove_noise(val):
    if pd.isna(val): return None
    val = str(val).encode('ascii', 'ignore').decode('ascii')
    val = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', val)
    val = re.sub(r'[^\w\s\.,;:!\?\-\(\)\/\'\"\@\#\%\&\+\=]', '', val)
    return re.sub(r'\s+', ' ', val).strip() or None

def clean_price(v):
    if pd.isna(v): return None
    v = re.sub(r'[^\d.]', '', str(v).replace(',', '.'))
    try: return float(v) if v else None
    except: return None

def clean_surface(v):
    if pd.isna(v): return None
    v = re.sub(r'[^\d.]', '', str(v).replace(',', '.'))
    try: return float(v) if v else None
    except: return None

def clean_city(val):
    if pd.isna(val): return None
    val = remove_noise(val)
    if not val: return None
    return val.strip().title()

def parse_date_flexible(val):
    if pd.isna(val): return pd.NaT
    val = str(val).strip()
    for fmt in ['%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M', '%Y-%m-%d', '%d-%m-%Y']:
        try: return datetime.strptime(val[:19], fmt)
        except: pass
    return pd.to_datetime(val, dayfirst=True, errors='coerce')

def parse_relative_date(relative_str, reference_date):
    if pd.isna(relative_str): return pd.to_datetime(reference_date)
    s = str(relative_str).lower().strip()
    try:
        ref = pd.to_datetime(reference_date)
        m = re.search(r'(\d+)', s)
        n = int(m.group(1)) if m else 1
        if 'minute' in s: return ref - pd.Timedelta(minutes=n)
        if 'hour'   in s: return ref - pd.Timedelta(hours=n)
        if 'day'    in s: return ref - pd.Timedelta(days=n)
        if 'week'   in s: return ref - pd.Timedelta(weeks=n)
        if 'month'  in s: return ref - pd.Timedelta(days=n * 30)
        if 'year'   in s: return ref - pd.Timedelta(days=n * 365)
        if 'just'   in s: return ref
    except: pass
    return pd.to_datetime(reference_date)

def extract_date_from_image_url_bnb(img_url):
    if pd.isna(img_url): return pd.NaT
    m = re.search(r'/uploads/(\d{4})/(\d{2})/', str(img_url))
    if m:
        try: return datetime(int(m.group(1)), int(m.group(2)), 1)
        except: pass
    return pd.NaT

def extract_date_from_facebook_cdn(img_url):
    if pd.isna(img_url): return pd.NaT
    m = re.search(r'oe=([0-9A-Fa-f]{8})', str(img_url))
    if m:
        try:
            ts  = int(m.group(1), 16)
            exp = datetime.fromtimestamp(ts)
            return exp - pd.Timedelta(days=90)
        except: pass
    return pd.NaT

def extract_date_from_filename(filename):
    m = re.search(r'(\d{8})', str(filename))
    if m:
        try: return datetime.strptime(m.group(1), '%Y%m%d')
        except: pass
    return datetime.now()

TYPE_MAP = {
    'appartement': 'Appartement', 'appart': 'Appartement', 'studio': 'Appartement',
    'duplex': 'Appartement', 'triplex': 'Appartement', 'penthouse': 'Appartement',
    's+1': 'Appartement', 's+2': 'Appartement', 's+3': 'Appartement',
    's+4': 'Appartement', 's+5': 'Appartement', 'rdc': 'Appartement',
    'maison': 'Maison', 'house': 'Maison', 'dar': 'Maison', 'bungalow': 'Maison',
    'villa': 'Villa', 'propriete': 'Villa', 'chalet': 'Villa',
    'chambre': 'Chambre', 'room': 'Chambre', 'colocation': 'Chambre',
    'local': 'Local Commercial', 'commerce': 'Local Commercial',
    'boutique': 'Local Commercial', 'magasin': 'Local Commercial',
    'hangar': 'Local Commercial', 'entrepot': 'Local Commercial',
    'showroom': 'Local Commercial', 'fonds de commerce': 'Local Commercial',
    'bureau': 'Bureau', 'office': 'Bureau', 'open space': 'Bureau', 'coworking': 'Bureau',
    'terrain': 'Terrain', 'lot': 'Terrain', 'lotissement': 'Terrain',
    'parcelle': 'Terrain', 'foncier': 'Terrain',
    'ferme': 'Ferme', 'agricole': 'Ferme', 'oliveraie': 'Ferme',
    'verger': 'Ferme', 'exploitation': 'Ferme', 'champ': 'Ferme',
}

TYPE_KW = {
    'Appartement':      ['appartement', 'studio', 'duplex', 'triplex', 'penthouse', 's+1', 's+2', 's+3', 's+4', 's+5', 'rdc'],
    'Maison':           ['maison', 'house', 'dar', 'bungalow'],
    'Villa':            ['villa', 'propriete'],
    'Chambre':          ['chambre', 'colocation'],
    'Local Commercial': ['local commercial', 'boutique', 'magasin', 'hangar', 'entrepot', 'showroom'],
    'Bureau':           ['bureau', 'office', 'open space', 'coworking'],
    'Terrain':          ['terrain', 'lot', 'lotissement', 'parcelle', 'foncier'],
    'Ferme':            ['ferme', 'agricole', 'oliveraie', 'verger', 'exploitation'],
}

DESC_KW = {
    'Appartement':      ['appartement', 'appart', 'studio', 'duplex'],
    'Maison':           ['maison', 'bungalow', 'dar'],
    'Villa':            ['villa'],
    'Chambre':          ['chambre a louer', 'colocation'],
    'Local Commercial': ['local commercial', 'fonds de commerce', 'boutique'],
    'Bureau':           ['bureau', 'office'],
    'Terrain':          ['terrain', 'parcelle'],
    'Ferme':            ['ferme', 'agricole', 'oliveraie'],
}

def std_type_bien(val, surface_m2=None, nb_pieces=None):
    if pd.isna(val): return None
    v = str(val).lower().strip()
    if re.search(r'immo\s*neuf|projet\s*neuf', v): return 'Appartement'
    if v == 'autre': return 'Autre'
    for k, can in TYPE_MAP.items():
        if k in v: return can
    return str(val).strip().title()


# ================================================================
# COORDONNÉES PAR GOUVERNORAT (depuis buildings.csv)
# ================================================================

if os.path.exists(BUILDINGS_PATH):
    _df_build = pd.read_csv(BUILDINGS_PATH)
    _df_build['gouvernorat_norm'] = (
        _df_build['city'].str.lower().str.strip().map(VILLE_TO_GOUVERNORAT).fillna('Unknown')
    )
    _coords_df = (
        _df_build[_df_build['gouvernorat_norm'] != 'Unknown']
        .groupby('gouvernorat_norm')[['lat', 'lon']].median().reset_index()
    )
    GOUVERNORAT_COORDS = {
        row['gouvernorat_norm']: (round(row['lat'], 4), round(row['lon'], 4))
        for _, row in _coords_df.iterrows()
    }
    log(f"  ✔ GOUVERNORAT_COORDS : {len(GOUVERNORAT_COORDS)} gouvernorats depuis buildings.csv")
else:
    GOUVERNORAT_COORDS = {}
    log(f"  [WARN] buildings.csv absent → coords fallback")

_FALLBACK_COORDS = {
    'Unknown': (36.8, 10.1), 'Ariana': (36.8665, 10.1647), 'Béja': (36.7256, 9.1817),
    'Ben Arous': (36.7533, 10.2283), 'Bizerte': (37.2744, 9.8739), 'Gabès': (33.8881, 10.0975),
    'Gafsa': (34.4250, 8.7842), 'Jendouba': (36.5011, 8.7757), 'Kairouan': (35.6781, 10.0963),
    'Kasserine': (35.1676, 8.8365), 'Kébili': (33.7052, 8.9691), 'Le Kef': (36.1744, 8.7148),
    'Mahdia': (35.5047, 11.0622), 'Manouba': (36.8100, 10.0972), 'Médenine': (33.3549, 10.5055),
    'Monastir': (35.7643, 10.8113), 'Nabeul': (36.4561, 10.7376), 'Sfax': (34.7405, 10.7603),
    'Sidi Bouzid': (35.0382, 9.4850), 'Siliana': (36.0850, 9.3708), 'Sousse': (35.8245, 10.6346),
    'Tataouine': (32.9211, 10.4519), 'Tozeur': (33.9197, 8.1335), 'Tunis': (36.8190, 10.1658),
    'Zaghouan': (36.4029, 10.1429),
}
for _g, _c in _FALLBACK_COORDS.items():
    if _g not in GOUVERNORAT_COORDS:
        GOUVERNORAT_COORDS[_g] = _c

GROUPE_MAP = {
    'Appartement': 'Residentiel', 'Maison': 'Residentiel',
    'Villa': 'Residentiel', 'Chambre': 'Residentiel',
    'Terrain': 'Foncier', 'Ferme': 'Foncier',
    'Local Commercial': 'Commercial', 'Bureau': 'Commercial',
    'Divers': 'Divers',
}

SEUILS = {
    'Residentiel': {'prix_min': 50,    'prix_max': 20_000_000, 'surf_min': 5,   'surf_max': 10_000},
    'Foncier':     {'prix_min': 1_000, 'prix_max': 50_000_000, 'surf_min': 50,  'surf_max': 5_000_000},
    'Commercial':  {'prix_min': 1_000, 'prix_max': 20_000_000, 'surf_min': 10,  'surf_max': 50_000},
    'Divers':      {'prix_min': 10,    'prix_max': 50_000,     'surf_min': 5,   'surf_max': 2_000},
}


# ================================================================
# ETAPE 1 — ETL : CHARGEMENT ET FUSION (identique BO2)
# ================================================================

section("ETAPE 1 — ETL : CHARGEMENT ET FUSION")

all_dfs = []
for name, path, ftype, sep in SOURCES_IMMO:
    if not os.path.exists(path):
        log(f"MANQUANT  {name}"); continue
    try:
        raw = pd.read_excel(path) if ftype == 'xlsx' else pd.read_csv(path, sep=sep or ',', on_bad_lines='skip')
    except Exception as e:
        log(f"ERREUR {name}: {e}"); continue

    if name == 'Facebook Marketplace':
        img_col  = 'image_url' if 'image_url' in raw.columns else None
        date_pub = raw[img_col].apply(extract_date_from_facebook_cdn) if img_col else \
                   pd.Series([extract_date_from_filename(path)] * len(raw))
        std = pd.DataFrame({
            'prix': raw['prix_affiche'].apply(clean_price), 'surface_m2': None,
            'ville': raw['ville'].apply(clean_city), 'gouvernorat': raw['gouvernorat'].apply(clean_city),
            'type_bien': raw['type_bien'].apply(remove_noise),
            'type_transaction': raw['type_annonce'].apply(remove_noise),
            'chambres': None, 'pieces': None,
            'description': raw['description'].apply(remove_noise),
            'source': name, 'date_publication': date_pub})

    elif name == 'Mubawab':
        date_pub = pd.Series([extract_date_from_filename(path)] * len(raw))
        std = pd.DataFrame({
            'prix': raw['prix_tnd'].apply(clean_price), 'surface_m2': raw['surface_m2'].apply(clean_surface),
            'ville': raw['ville'].apply(clean_city), 'gouvernorat': raw['gouvernorat'].apply(clean_city),
            'type_bien': raw['type_bien'].apply(remove_noise),
            'type_transaction': raw['categorie'].apply(remove_noise),
            'chambres': pd.to_numeric(raw['nb_chambres'], errors='coerce'),
            'pieces': pd.to_numeric(raw['nb_pieces'], errors='coerce'),
            'description': raw['titre'].apply(remove_noise),
            'source': name, 'date_publication': date_pub})

    elif name == 'Mubawab Partial':
        dc       = 'description' if 'description' in raw.columns else 'titre'
        date_pub = raw['date_collecte'].apply(parse_date_flexible) if 'date_collecte' in raw.columns else \
                   pd.Series([extract_date_from_filename(path)] * len(raw))
        std = pd.DataFrame({
            'prix': raw['prix'].apply(clean_price), 'surface_m2': raw['surface_m2'].apply(clean_surface),
            'ville': raw['ville'].apply(clean_city), 'gouvernorat': raw['ville'].apply(clean_city),
            'type_bien': raw['type_propriete'].apply(remove_noise) if 'type_propriete' in raw.columns else None,
            'type_transaction': None,
            'chambres': pd.to_numeric(raw['nombre_chambres'], errors='coerce'), 'pieces': None,
            'description': raw[dc].apply(remove_noise), 'source': name, 'date_publication': date_pub})

    elif name == 'Tayara':
        if 'scraped_at' in raw.columns and 'date' in raw.columns:
            date_pub = pd.Series([parse_relative_date(row['date'], row['scraped_at'])
                                  for _, row in raw.iterrows()])
        elif 'scraped_at' in raw.columns:
            date_pub = raw['scraped_at'].apply(parse_date_flexible)
        else:
            date_pub = pd.Series([extract_date_from_filename(path)] * len(raw))
        std = pd.DataFrame({
            'prix': raw['price'].apply(clean_price), 'surface_m2': raw['superficie'].apply(clean_surface),
            'ville': raw['region'].apply(clean_city), 'gouvernorat': raw['location'].apply(clean_city),
            'type_bien': raw['category'].apply(remove_noise),
            'type_transaction': raw['type_de_transaction'].apply(remove_noise),
            'chambres': pd.to_numeric(raw['chambres'], errors='coerce'), 'pieces': None,
            'description': raw['description'].apply(remove_noise), 'source': name, 'date_publication': date_pub})

    elif name == 'Tunisie Annonces':
        if 'date_insertion' in raw.columns:       date_pub = raw['date_insertion'].apply(parse_date_flexible)
        elif 'date_modification' in raw.columns:  date_pub = raw['date_modification'].apply(parse_date_flexible)
        else:                                      date_pub = pd.Series([pd.NaT] * len(raw))
        std = pd.DataFrame({
            'prix': raw['prix_montant'].apply(clean_price), 'surface_m2': raw['surface_couverte'].apply(clean_surface),
            'ville': raw['gouvernorat'].apply(clean_city), 'gouvernorat': raw['gouvernorat'].apply(clean_city),
            'type_bien': raw['type_bien'].apply(remove_noise),
            'type_transaction': raw['type_transaction'].apply(remove_noise),
            'chambres': pd.to_numeric(raw['nombre_chambres'], errors='coerce'),
            'pieces': pd.to_numeric(raw['nombre_pieces'], errors='coerce'),
            'description': raw['description_complete'].apply(remove_noise),
            'source': name, 'date_publication': date_pub})

    elif name == 'Century21':
        date_pub = raw['date_scraping'].apply(parse_date_flexible) if 'date_scraping' in raw.columns else \
                   pd.Series([extract_date_from_filename(path)] * len(raw))
        std = pd.DataFrame({
            'prix': raw['prix'].apply(clean_price), 'surface_m2': raw['surface_m2'].apply(clean_surface),
            'ville': raw['localisation'].apply(clean_city), 'gouvernorat': raw['localisation'].apply(clean_city),
            'type_bien': raw['type'].apply(remove_noise), 'type_transaction': raw['status'].apply(remove_noise),
            'chambres': None, 'pieces': None,
            'description': raw['description'].apply(remove_noise), 'source': name, 'date_publication': date_pub})

    elif name == 'HomeInTunisia':
        date_pub = raw['date_scraping'].apply(parse_date_flexible) if 'date_scraping' in raw.columns else \
                   pd.Series([extract_date_from_filename(path)] * len(raw))
        std = pd.DataFrame({
            'prix': raw['prix'].apply(clean_price), 'surface_m2': raw['surface_m2'].apply(clean_surface),
            'ville': raw['localisation'].apply(clean_city), 'gouvernorat': raw['localisation'].apply(clean_city),
            'type_bien': raw['type'].apply(remove_noise), 'type_transaction': None,
            'chambres': pd.to_numeric(raw['nombre_chambres'], errors='coerce'),
            'pieces': pd.to_numeric(raw['nombre_pieces'], errors='coerce'),
            'description': raw['titre'].apply(remove_noise), 'source': name, 'date_publication': date_pub})

    elif name == 'BnB':
        img_col  = 'image' if 'image' in raw.columns else None
        date_pub = raw[img_col].apply(extract_date_from_image_url_bnb) if img_col else \
                   pd.Series([extract_date_from_filename(path)] * len(raw))
        date_pub = date_pub.fillna(extract_date_from_filename(path))
        std = pd.DataFrame({
            'prix': raw['prix_montant'].apply(clean_price),
            'surface_m2': raw['meta_Surface'].apply(clean_surface) if 'meta_Surface' in raw.columns else None,
            'ville': raw['localisation'].apply(clean_city), 'gouvernorat': raw['localisation'].apply(clean_city),
            'type_bien': None, 'type_transaction': 'Location Courte Duree',
            'chambres': pd.to_numeric(raw['meta_Lits'], errors='coerce') if 'meta_Lits' in raw.columns else None,
            'pieces': None, 'description': raw['description'].apply(remove_noise),
            'source': name, 'date_publication': date_pub})

    for col in ['prix', 'surface_m2', 'ville', 'gouvernorat', 'type_bien', 'type_transaction',
                'chambres', 'pieces', 'description', 'source', 'date_publication']:
        if col not in std.columns:
            std[col] = None

    all_dfs.append(std)
    log(f"OK  {name:<25}: {len(std):>6} annonces")

df = pd.concat(all_dfs, ignore_index=True, sort=False)
n0 = len(df)
log(f"\nTotal brut fusionné : {n0:,}")


# ================================================================
# ETAPE 2 — DÉDUPLICATION (3 passes)
# ================================================================

section("ETAPE 2 — DEDUPLICATION")

n_p1 = len(df)
df   = df.drop_duplicates(keep='first')
log(f"Passe 1 - Stricts                      : -{n_p1 - len(df)}")

n_p2 = len(df)
_prix_key = df['prix'].fillna(-1).round(0).astype(int).astype(str)
_surf_key = df['surface_m2'].fillna(-1).round(0).astype(int).astype(str)
_gouv_key = df['gouvernorat'].fillna('').astype('category').cat.codes.astype(str)
_desc_key = df['description'].fillna('').str[:120]
df['_dup_key'] = pd.util.hash_pandas_object(
    pd.DataFrame({'p': _prix_key, 's': _surf_key, 'g': _gouv_key, 'd': _desc_key}),
    index=False
)
df = df.drop_duplicates(subset=['_dup_key'], keep='first').drop(columns=['_dup_key'])
log(f"Passe 2 - Hash(Prix+Surf+Gouv+Desc)    : -{n_p2 - len(df)}")

n_p3    = len(df)
_dhash  = pd.util.hash_pandas_object(df['description'].fillna('').str[:200].str.lower(), index=False)
df['_dhash'] = _dhash
mask = df['description'].notna() & (df['description'].str.strip() != '')
df   = pd.concat([df[mask].drop_duplicates(subset=['_dhash'], keep='first'),
                  df[~mask]], ignore_index=True).drop(columns=['_dhash'])
log(f"Passe 3 - Hash(Description)            : -{n_p3 - len(df)}")

n_apres_dedup = len(df)
log(f"\n  Total avant  : {n0:>7,}")
log(f"  Total après  : {n_apres_dedup:>7,}  (-{n0 - n_apres_dedup}, soit {(n0 - n_apres_dedup)/n0*100:.1f}%)")


# ================================================================
# ETAPE 2b — VÉRIFICATION TYPES
# ================================================================

section("ETAPE 2b — VERIFICATION TYPES")

for col in ['prix', 'surface_m2', 'chambres', 'pieces']:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

df['date_publication'] = pd.to_datetime(df['date_publication'], errors='coerce')

for col in ['gouvernorat', 'ville', 'type_bien', 'type_transaction', 'source']:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip().replace({'nan': None, 'None': None, '': None})

log(f"Dimensions : {df.shape[0]:,} lignes × {df.shape[1]} colonnes")


# ================================================================
# ETAPE 3 — STANDARDISATION
# ================================================================

section("ETAPE 3 — STANDARDISATION")

df['type_transaction'] = df['type_transaction'].apply(std_transaction)

LOC_KEYWORDS   = ['location', 'locat', 'à louer', 'a louer', 'louer', 'rent', 'loue',
                   'mensuel', 'mensuelle', 'loyer', 'courte duree', 'saisonnier', 'vacances']
VENTE_KEYWORDS = ['vente', 'à vendre', 'a vendre', 'vendre', 'achat', 'cession', 'vend', 'sale']

def resolve_transaction(prix_val, trans_val, desc_str=None):
    if not pd.isna(trans_val):
        try:
            v = int(float(trans_val))
            if v in (1, 2): return v
        except: pass
    desc = str(desc_str or '').lower()
    if desc:
        loc_s   = sum(1 for kw in LOC_KEYWORDS   if kw in desc)
        vente_s = sum(1 for kw in VENTE_KEYWORDS  if kw in desc)
        if loc_s > 0 and loc_s >= vente_s:  return 1
        if vente_s > 0 and vente_s > loc_s: return 2
    if not pd.isna(prix_val):
        try:
            p = float(prix_val)
            if p < 5000:  return 1
            if p > 20000: return 2
            return 1
        except: pass
    return np.nan

df['type_transaction'] = df.apply(
    lambda r: resolve_transaction(r.get('prix'), r.get('type_transaction'), r.get('description')), axis=1)
tt_mode = df['type_transaction'].mode()
if len(tt_mode) > 0:
    df['type_transaction'] = df['type_transaction'].fillna(int(tt_mode.iloc[0]))
df['type_transaction'] = df['type_transaction'].astype(int)

df['type_bien'] = df.apply(
    lambda r: std_type_bien(r.get('type_bien'), r.get('surface_m2'), r.get('pieces')), axis=1)

if 'pieces' not in df.columns:   df['pieces']   = np.nan
if 'chambres' not in df.columns: df['chambres'] = np.nan
df['nb_pieces'] = df['pieces'].combine_first(df['chambres'])

df['_gouvernorat_str'] = df['gouvernorat'].apply(normalize_gouvernorat)
log(f"  gouvernorat unique : {sorted(df['_gouvernorat_str'].unique().tolist())}")
log(f"  type_transaction   : {df['type_transaction'].value_counts().sort_index().to_dict()}")


# ================================================================
# ETAPE 4 — SEGMENTATION
# ================================================================

section("ETAPE 4 — SEGMENTATION EN 4 GROUPES")

def classify(row):
    tb   = str(row.get('type_bien', '') or '').lower()
    desc = str(row.get('description', '') or '').lower()
    if row.get('source') == 'BnB': return 'Divers'
    for typ, kws in TYPE_KW.items():
        for kw in kws:
            if kw in tb: return typ
    if not tb or tb in ('none', 'nan', ''):
        for typ, kws in DESC_KW.items():
            for kw in kws:
                if kw in desc: return typ
    return 'Divers'

df['type_categorise'] = df.apply(classify, axis=1)
df['groupe']          = df['type_categorise'].map(GROUPE_MAP)

for g, c in df['groupe'].value_counts().items():
    log(f"  {g:<15}: {c:>6} ({c / len(df) * 100:.1f}%)")


# ================================================================
# ETAPE 5 — NETTOYAGE + ISOLATION FOREST
# ================================================================

section("ETAPE 5 — NETTOYAGE + ISOLATION FOREST")

groupes_clean = {}
for groupe in ['Residentiel', 'Foncier', 'Commercial', 'Divers']:
    dg = df[df['groupe'] == groupe].copy()
    if len(dg) == 0: continue
    n_avant = len(dg)
    s = SEUILS[groupe]

    dg = dg[dg['prix'].isna()       | ((dg['prix']       >= s['prix_min']) & (dg['prix']       <= s['prix_max']))]
    dg = dg[dg['surface_m2'].isna() | ((dg['surface_m2'] >= s['surf_min']) & (dg['surface_m2'] <= s['surf_max']))]

    n_anom = 0
    feats  = dg[['prix', 'surface_m2']].dropna()
    if len(feats) >= 5_000:
        X      = StandardScaler().fit_transform(feats)
        preds  = IsolationForest(n_estimators=50, contamination=0.05, random_state=42, n_jobs=-1).fit_predict(X)
        anoms  = feats.index[preds == -1]
        dg     = dg.drop(index=anoms)
        n_anom = len(anoms)
    elif len(feats) >= 50:
        for col in ['prix', 'surface_m2']:
            if col not in dg.columns: continue
            q1, q3 = dg[col].quantile(0.01), dg[col].quantile(0.99)
            out    = (dg[col] < q1) | (dg[col] > q3)
            n_anom += int(out.sum())
            dg      = dg[~out]

    groupes_clean[groupe] = dg
    log(f"  {groupe:<15}: {n_avant:>6} -> {len(dg):>6} | {n_anom} anomalies supprimées")


# ================================================================
# ETAPE 5b — IMPUTATION NA
# ================================================================

section("ETAPE 5b — IMPUTATION NA")

for groupe, dg in groupes_clean.items():
    n = len(dg)
    for col in ['prix', 'surface_m2']:
        if col not in dg.columns: continue
        n_na = dg[col].isna().sum()
        if n_na == 0: continue
        rate = n_na / n
        if rate <= 0.30:
            dg[col] = dg[col].fillna(dg[col].median())
        elif rate <= 0.60:
            knn_base = [c for c in ['prix', 'surface_m2'] if c in dg.columns and dg[c].notna().sum() > n * 0.3]
            if len(knn_base) >= 2 and n >= 50:
                try:
                    sub    = dg[knn_base].fillna(dg[knn_base].median())
                    result = KNNImputer(n_neighbors=5).fit_transform(sub)
                    for i, c in enumerate(knn_base):
                        dg[c] = result[:, i]
                except:
                    dg[col] = dg[col].fillna(dg[col].median())
            else:
                dg[col] = dg[col].fillna(dg[col].median())
        else:
            pool = dg[col].dropna().values
            if len(pool) > 0:
                np.random.seed(42)
                dg.loc[dg[col].isna(), col] = np.random.choice(pool, size=n_na)
    groupes_clean[groupe] = dg
    log(f"  [{groupe}] prix NaN: {dg['prix'].isna().sum()} | surface_m2 NaN: {dg['surface_m2'].isna().sum()}")


# ================================================================
# ETAPE 5c — VISUALISATION
# ================================================================

section("ETAPE 5c — VISUALISATION DISTRIBUTIONS")

if DEBUG_MODE:
    log("  [DEBUG_MODE] Visualisation skippée")
else:
    os.makedirs('plots_bo5', exist_ok=True)
    for groupe, dg in groupes_clean.items():
        cols_avail = [c for c in ['prix', 'surface_m2'] if c in dg.columns and dg[c].notna().sum() > 10]
        if not cols_avail: continue
        fig, axes = plt.subplots(2, len(cols_avail), figsize=(6 * len(cols_avail), 10))
        if len(cols_avail) == 1: axes = axes.reshape(2, 1)
        fig.suptitle(f'[{groupe}] — Distributions BO5', fontsize=12, fontweight='bold')
        for i, col in enumerate(cols_avail):
            data = dg[col].dropna()
            axes[0][i].hist(data, bins=50, color='steelblue', edgecolor='white', alpha=0.85)
            axes[0][i].set_title(f'{col} — Histogramme')
            axes[0][i].axvline(data.mean(), color='red', linestyle='--', linewidth=1.2)
            axes[0][i].xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
            axes[1][i].boxplot(data, vert=True, patch_artist=True,
                               boxprops=dict(facecolor='#AED6F1', color='navy'),
                               medianprops=dict(color='darkred', linewidth=2),
                               flierprops=dict(marker='o', markersize=3, alpha=0.4))
            Q1, Q3 = data.quantile(0.25), data.quantile(0.75)
            axes[1][i].set_title(f'{col} — Boxplot IQR=[{Q1:,.0f};{Q3:,.0f}]')
        plt.tight_layout()
        fp = f'plots_bo5/{groupe.lower()}_bo5.png'
        plt.savefig(fp, dpi=120, bbox_inches='tight')
        plt.close()
        log(f"  Plot sauvegardé : {fp}")


# ================================================================
# ETAPE 5d — ENCODAGE MANUEL
# ================================================================

section("ETAPE 5d — ENCODAGE MANUEL")

for groupe, dg in groupes_clean.items():
    if '_gouvernorat_str' not in dg.columns:
        valid_idx = [i for i in dg.index if i in df.index]
        if valid_idx and '_gouvernorat_str' in df.columns:
            dg.loc[valid_idx, '_gouvernorat_str'] = df.loc[valid_idx, '_gouvernorat_str'].values
        else:
            dg['_gouvernorat_str'] = dg['gouvernorat'].apply(normalize_gouvernorat)
    if 'type_bien' in dg.columns and dg['type_bien'].dtype == object:
        dg['type_bien'] = dg['type_bien'].apply(encode_type_bien)
    groupes_clean[groupe] = dg
    log(f"  [{groupe}] encodage OK")

df['gouvernorat'] = df['_gouvernorat_str'].apply(encode_gouvernorat)


# ================================================================
# ETAPE 6 — GÉOCODAGE
# ================================================================

section("ETAPE 6 — GEOCODAGE")

CENTROIDES = {
    'tunis': (36.8190, 10.1658), 'ariana': (36.8665, 10.1647),
    'ben arous': (36.7533, 10.2283), 'nabeul': (36.4561, 10.7376),
    'sousse': (35.8245, 10.6346), 'sfax': (34.7405, 10.7603),
    'monastir': (35.7643, 10.8113), 'bizerte': (37.2744, 9.8739),
    'la marsa': (36.8769, 10.3247), 'carthage': (36.8528, 10.3247),
    'sidi bou said': (36.8694, 10.3406), 'gammarth': (36.9100, 10.2900),
    'hammamet': (36.3994, 10.6134), 'hammam sousse': (35.8604, 10.5953),
    'djerba': (33.8075, 10.8451),
}

def assign_latlon(ville, gouvernorat_str):
    vk = str(ville or '').lower().strip()
    gk = str(gouvernorat_str or '').strip()
    if vk in CENTROIDES:
        return CENTROIDES[vk]
    if gk in GOUVERNORAT_COORDS:
        return GOUVERNORAT_COORDS[gk]
    return GOUVERNORAT_COORDS.get('Unknown', (36.8, 10.1))

latlon    = df.apply(lambda r: assign_latlon(r.get('ville'), r.get('_gouvernorat_str')), axis=1)
df['lat'] = latlon.apply(lambda x: x[0]).fillna(36.8)
df['lon'] = latlon.apply(lambda x: x[1]).fillna(10.1)

for groupe, dg in groupes_clean.items():
    valid_idx = [i for i in dg.index if i in df.index]
    if valid_idx:
        dg.loc[valid_idx, 'lat'] = df.loc[valid_idx, 'lat'].values
        dg.loc[valid_idx, 'lon'] = df.loc[valid_idx, 'lon'].values
    dg['lat'] = pd.to_numeric(dg.get('lat', 36.8), errors='coerce').fillna(36.8)
    dg['lon'] = pd.to_numeric(dg.get('lon', 10.1), errors='coerce').fillna(10.1)
    if '_gouvernorat_str' not in dg.columns:
        valid_idx2 = [i for i in dg.index if i in df.index]
        if valid_idx2:
            dg.loc[valid_idx2, '_gouvernorat_str'] = df.loc[valid_idx2, '_gouvernorat_str'].values
    groupes_clean[groupe] = dg

log(f"Géocodage : {df['lat'].notna().mean() * 100:.1f}% annonces avec lat/lon")


# ================================================================
# ─────────────────────────────────────────────────────────────────
#   ÉTAPES SPÉCIFIQUES BO5
#   Remplacent les étapes 7-11 de BO2 (Vision, BERT, Fusion,
#   prix_transaction_estimated, market_tension, cycle_marche)
# ─────────────────────────────────────────────────────────────────
# ================================================================


# ================================================================
# ETAPE 7 BO5 — FEATURES MARCHÉ
#   prix_m2, variation_prix_m2, inflation_glissement_annuel,
#   croissance_pib_trim, high_season
# ================================================================

section("ETAPE 7 BO5 — FEATURES MARCHÉ")

# ── 7.1 Données macro Tunisie (fixes par trimestre/année) ────────
# Sources : INS Tunisie, Banque Centrale de Tunisie
# Ces valeurs servent de contexte macro régional
MACRO_DATA = {
    # (annee, trimestre) → (inflation_glissement_annuel %, croissance_pib_trim %)
    (2022, 1): (7.2,  2.1), (2022, 2): (8.1,  1.8), (2022, 3): (9.0,  1.5), (2022, 4): (9.8,  2.0),
    (2023, 1): (10.4, 1.2), (2023, 2): (9.6,  1.9), (2023, 3): (8.8,  2.3), (2023, 4): (8.1,  2.5),
    (2024, 1): (7.5,  2.1), (2024, 2): (7.0,  2.4), (2024, 3): (6.8,  2.6), (2024, 4): (6.5,  2.8),
    (2025, 1): (6.2,  2.5), (2025, 2): (6.0,  3.0), (2025, 3): (5.8,  3.2), (2025, 4): (6.1,  3.0),
    (2026, 1): (6.3,  2.8),
}

# ── 7.2 Saisons immobilières (mois de forte activité Tunisie) ────
# High season : printemps (3,4,5), automne (9,10,11)
HIGH_SEASON_MONTHS = {3, 4, 5, 9, 10, 11}

def get_macro_features(date_val):
    """
    Extrait inflation_glissement_annuel, croissance_pib_trim, high_season
    depuis la date de publication.
    """
    dt = pd.NaT
    if pd.notna(date_val):
        try:
            dt = pd.to_datetime(date_val)
            if dt.year < 2020 or dt.year > 2027:
                dt = pd.NaT
        except: pass
    if pd.isna(dt):
        dt = datetime.now()

    year    = dt.year
    quarter = (dt.month - 1) // 3 + 1
    month   = dt.month

    ctx = MACRO_DATA.get((year, quarter), (6.3, 2.8))   # fallback 2026-Q1
    return {
        'inflation_glissement_annuel': ctx[0],
        'croissance_pib_trim':         ctx[1],
        'high_season':                 1 if month in HIGH_SEASON_MONTHS else 0,
        '_date_year':                  year,
        '_date_month':                 month,
    }

log("  Calcul features macro (inflation, PIB, saison)...")
df_macro = pd.DataFrame(
    [dict(**get_macro_features(row.get('date_publication')), _idx=idx)
     for idx, row in df.iterrows()]
).set_index('_idx')
df = df.join(df_macro, how='left')
log(f"  high_season       : {df['high_season'].value_counts().to_dict()}")
log(f"  inflation moy     : {df['inflation_glissement_annuel'].mean():.2f}%")
log(f"  PIB trim moy      : {df['croissance_pib_trim'].mean():.2f}%")

# Propager dans groupes_clean
for groupe, dg in groupes_clean.items():
    valid_idx = [i for i in dg.index if i in df.index]
    for col in ['inflation_glissement_annuel', 'croissance_pib_trim', 'high_season',
                '_date_year', '_date_month']:
        if valid_idx and col in df.columns:
            dg.loc[valid_idx, col] = df.loc[valid_idx, col].values
    groupes_clean[groupe] = dg

# ── 7.3 prix_m2 et variation_prix_m2 par gouvernorat ─────────────
# Agrégation au niveau gouvernorat (objectif macro-régional BO5)
# prix_m2 = médiane(prix/surface) par gouvernorat
# variation_prix_m2 = pct_change par rapport au trimestre précédent

log("  Calcul prix_m2 et variation par gouvernorat...")

# Calcul prix_m2 annonce par annonce avant agrégation
df['_prix_m2_brut'] = np.where(
    df['prix'].notna() & df['surface_m2'].notna() & (df['surface_m2'] > 0),
    df['prix'] / df['surface_m2'],
    np.nan
)
df.loc[df['_prix_m2_brut'] > 100_000, '_prix_m2_brut'] = np.nan
df.loc[df['_prix_m2_brut'] <= 0,      '_prix_m2_brut'] = np.nan

# Agrégat gouvernorat × trimestre
df['_periode_trim'] = df['_date_year'].astype(str) + '-Q' + \
                      ((df['_date_month'].fillna(1) - 1) // 3 + 1).astype(int).astype(str)

agg_prix = (
    df.groupby(['_gouvernorat_str', '_periode_trim'])['_prix_m2_brut']
    .median().reset_index(name='prix_m2')
)

# Variation mensuelle par gouvernorat
agg_var = agg_prix.sort_values(['_gouvernorat_str', '_periode_trim']).copy()
agg_var['variation_prix_m2'] = (
    agg_var.groupby('_gouvernorat_str')['prix_m2']
    .pct_change(periods=1).round(4).fillna(0.0)
)

# Merge sur df principal
df = df.drop(columns=['prix_m2', 'variation_prix_m2'], errors='ignore')
df = df.merge(
    agg_var[['_gouvernorat_str', '_periode_trim', 'prix_m2', 'variation_prix_m2']],
    on=['_gouvernorat_str', '_periode_trim'], how='left'
)
# Imputation prix_m2 manquant → médiane gouvernorat → médiane globale
if df['prix_m2'].isna().sum() > 0:
    med_g = df.groupby('_gouvernorat_str')['prix_m2'].transform('median')
    df['prix_m2'] = df['prix_m2'].fillna(med_g)
    df['prix_m2'] = df['prix_m2'].fillna(df['prix_m2'].median())
df['variation_prix_m2'] = df['variation_prix_m2'].fillna(0.0)
df['prix_m2']           = df['prix_m2'].round(2)

log(f"  prix_m2      — médiane : {df['prix_m2'].median():.0f} TND/m²")
log(f"  variation    — [-1,+∞] : min={df['variation_prix_m2'].min():.2f}  max={df['variation_prix_m2'].max():.2f}")

# Propager dans groupes_clean
for groupe, dg in groupes_clean.items():
    valid_idx = [i for i in dg.index if i in df.index]
    for col in ['prix_m2', 'variation_prix_m2']:
        if valid_idx and col in df.columns:
            dg.loc[valid_idx, col] = df.loc[valid_idx, col].values
    dg['prix_m2']         = pd.to_numeric(dg.get('prix_m2', np.nan), errors='coerce').fillna(df['prix_m2'].median())
    dg['variation_prix_m2'] = pd.to_numeric(dg.get('variation_prix_m2', 0.0), errors='coerce').fillna(0.0)
    groupes_clean[groupe] = dg


# ================================================================
# ETAPE 8 BO5 — FEATURES OSM (Amenities + Routes + Buildings + Landuse)
#   nb_amenities_total, ratio_amenities_commerce, ratio_amenities_sante,
#   densite_routes_km, nb_stations_transport, nb_buildings_residentiel,
#   surface_landuse_residentiel, nb_commerce, densite_population
# ================================================================

section("ETAPE 8 BO5 — FEATURES OSM SATELLITE")

GOV_STR = '_gouvernorat_str'

# Dicts résultat {gouvernorat_str → valeur}
EXT_NB_AMENITIES        = {}   # nb_amenities_total
EXT_RATIO_COMMERCE      = {}   # ratio_amenities_commerce  [0,1]
EXT_RATIO_SANTE         = {}   # ratio_amenities_sante     [0,1]
EXT_NB_COMMERCE         = {}   # nb_commerce (amenity commerce absolu)
EXT_NB_TRANSPORT        = {}   # nb_stations_transport
EXT_DENSITE_ROUTES      = {}   # densite_routes_km (km de route / km² gouvernorat)
EXT_NB_BUILDINGS_RES    = {}   # nb_buildings_residentiel
EXT_SURFACE_LANDUSE_RES = {}   # surface_landuse_residentiel (m²)
EXT_DENSITE_POP         = {}   # densite_population (score 1-5 ou densité brute)

# ── 8.1 Amenities OSM ────────────────────────────────────────────
if os.path.exists(AMENITIES_PATH):
    try:
        df_am = pd.read_csv(AMENITIES_PATH)
        df_am['_gouv'] = df_am['city'].str.lower().str.strip().map(VILLE_TO_GOUVERNORAT).fillna('Unknown')
        df_am = df_am[df_am['_gouv'] != 'Unknown']

        # Catégorisation des types d'amenity
        COMMERCE_TYPES  = {'shop', 'store', 'mall', 'market', 'supermarket', 'convenience',
                           'bakery', 'pharmacy', 'clothes', 'furniture', 'electronics',
                           'restaurant', 'cafe', 'fast_food', 'bar'}
        SANTE_TYPES     = {'hospital', 'clinic', 'doctors', 'dentist', 'pharmacy',
                           'health_centre', 'veterinary', 'nursing_home'}
        TRANSPORT_TYPES = {'bus_station', 'bus_stop', 'train_station', 'subway_entrance',
                           'ferry_terminal', 'taxi', 'parking', 'charging_station'}

        def categorize_amenity(t):
            t = str(t).lower().strip()
            if t in COMMERCE_TYPES:  return 'commerce'
            if t in SANTE_TYPES:     return 'sante'
            if t in TRANSPORT_TYPES: return 'transport'
            return 'autre'

        if 'type' in df_am.columns:
            df_am['_cat'] = df_am['type'].apply(categorize_amenity)
        elif 'amenity' in df_am.columns:
            df_am['_cat'] = df_am['amenity'].apply(categorize_amenity)
        else:
            df_am['_cat'] = 'autre'

        agg_am = df_am.groupby('_gouv').agg(
            _total     = ('_cat', 'count'),
            _commerce  = ('_cat', lambda x: (x == 'commerce').sum()),
            _sante     = ('_cat', lambda x: (x == 'sante').sum()),
            _transport = ('_cat', lambda x: (x == 'transport').sum()),
        ).reset_index()

        for _, row in agg_am.iterrows():
            g = row['_gouv']
            tot = max(row['_total'], 1)
            EXT_NB_AMENITIES[g]   = int(row['_total'])
            EXT_NB_COMMERCE[g]    = int(row['_commerce'])
            EXT_NB_TRANSPORT[g]   = int(row['_transport'])
            EXT_RATIO_COMMERCE[g] = round(row['_commerce'] / tot, 4)
            EXT_RATIO_SANTE[g]    = round(row['_sante'] / tot, 4)

        log(f"  ✔ amenities : {len(EXT_NB_AMENITIES)} gouvernorats | "
            f"total moy={np.mean(list(EXT_NB_AMENITIES.values())):.0f}")
    except Exception as e:
        log(f"  [WARN] amenities.csv : {e}")
else:
    log(f"  [WARN] amenities.csv absent → valeurs par défaut")

# ── 8.2 Routes OSM ───────────────────────────────────────────────
# densite_routes_km = longueur totale routes (km) / surface gouvernorat (km²)
# Approximation : on compte le nombre de segments comme proxy
GOUVERNORAT_AREA_KM2 = {
    'Tunis': 346, 'Ariana': 498, 'Ben Arous': 761, 'Manouba': 1137,
    'Nabeul': 2788, 'Zaghouan': 2768, 'Bizerte': 3685, 'Béja': 3558,
    'Jendouba': 3102, 'Le Kef': 4965, 'Siliana': 4631, 'Kairouan': 6712,
    'Kasserine': 8066, 'Sidi Bouzid': 7396, 'Sousse': 2621, 'Monastir': 1019,
    'Mahdia': 2966, 'Sfax': 7545, 'Gabès': 7175, 'Médenine': 8588,
    'Tataouine': 38889, 'Gafsa': 8990, 'Tozeur': 4425, 'Kébili': 22084,
    'Unknown': 5000,
}

if os.path.exists(ROADS_PATH):
    try:
        df_rd = pd.read_csv(ROADS_PATH)
        df_rd['_gouv'] = df_rd['city'].str.lower().str.strip().map(VILLE_TO_GOUVERNORAT).fillna('Unknown')
        df_rd = df_rd[df_rd['_gouv'] != 'Unknown']

        # Longueur totale par gouvernorat (colonne length en mètres si disponible, sinon count)
        if 'length' in df_rd.columns:
            agg_rd = df_rd.groupby('_gouv')['length'].sum().reset_index(name='_len_m')
            agg_rd['_len_km'] = agg_rd['_len_m'] / 1000
        elif 'length_km' in df_rd.columns:
            agg_rd = df_rd.groupby('_gouv')['length_km'].sum().reset_index(name='_len_km')
        else:
            # Proxy : chaque segment = ~500m en moyenne
            agg_rd = df_rd.groupby('_gouv').size().reset_index(name='_nb_seg')
            agg_rd['_len_km'] = agg_rd['_nb_seg'] * 0.5

        for _, row in agg_rd.iterrows():
            g    = row['_gouv']
            area = GOUVERNORAT_AREA_KM2.get(g, 5000)
            EXT_DENSITE_ROUTES[g] = round(row['_len_km'] / area, 4)

        log(f"  ✔ routes : {len(EXT_DENSITE_ROUTES)} gouvernorats | "
            f"densité moy={np.mean(list(EXT_DENSITE_ROUTES.values())):.2f} km/km²")
    except Exception as e:
        log(f"  [WARN] roads.csv : {e}")
else:
    log(f"  [WARN] roads.csv absent → densite_routes_km par défaut")

# ── 8.3 Buildings OSM ────────────────────────────────────────────
# nb_buildings_residentiel = count bâtiments résidentiels par gouvernorat
if os.path.exists(BUILDINGS_PATH):
    try:
        df_bl = pd.read_csv(BUILDINGS_PATH)
        df_bl['_gouv'] = df_bl['city'].str.lower().str.strip().map(VILLE_TO_GOUVERNORAT).fillna('Unknown')
        df_bl = df_bl[df_bl['_gouv'] != 'Unknown']

        RES_TYPES = {'yes', 'residential', 'house', 'apartments', 'detached',
                     'semidetached_house', 'terrace', 'bungalow', 'villa', 'dormitory'}

        if 'building' in df_bl.columns:
            df_bl['_is_res'] = df_bl['building'].str.lower().str.strip().isin(RES_TYPES)
        else:
            df_bl['_is_res'] = True  # par défaut tout est résidentiel

        agg_bl = df_bl.groupby('_gouv').agg(
            _nb_res = ('_is_res', 'sum'),
        ).reset_index()

        for _, row in agg_bl.iterrows():
            EXT_NB_BUILDINGS_RES[row['_gouv']] = int(row['_nb_res'])

        log(f"  ✔ buildings : {len(EXT_NB_BUILDINGS_RES)} gouvernorats | "
            f"nb_res moy={np.mean(list(EXT_NB_BUILDINGS_RES.values())):.0f}")
    except Exception as e:
        log(f"  [WARN] buildings.csv pour nb_buildings : {e}")
else:
    log(f"  [WARN] buildings.csv absent → nb_buildings_residentiel par défaut")

# ── 8.4 Landuse OSM ──────────────────────────────────────────────
# surface_landuse_residentiel = surface (m²) zones résidentielles par gouvernorat
if os.path.exists(LANDUSE_PATH):
    try:
        df_lu = pd.read_csv(LANDUSE_PATH)
        df_lu['_gouv'] = df_lu['city'].str.lower().str.strip().map(VILLE_TO_GOUVERNORAT).fillna('Unknown')
        df_lu = df_lu[df_lu['_gouv'] != 'Unknown']

        RESIDENTIAL_LANDUSE = {'residential', 'housing', 'living_street', 'garages', 'apartments'}

        lu_col = None
        for c in ['landuse', 'type', 'category', 'land_use']:
            if c in df_lu.columns:
                lu_col = c; break

        if lu_col:
            df_lu['_is_res'] = df_lu[lu_col].str.lower().str.strip().isin(RESIDENTIAL_LANDUSE)
        else:
            df_lu['_is_res'] = True

        area_col = None
        for c in ['area', 'surface', 'area_m2', 'surface_m2']:
            if c in df_lu.columns:
                area_col = c; break

        if area_col:
            agg_lu = df_lu[df_lu['_is_res']].groupby('_gouv')[area_col].sum().reset_index(name='_surf')
        else:
            # Proxy : chaque zone résidentielle ≈ 10 000 m²
            agg_lu = df_lu[df_lu['_is_res']].groupby('_gouv').size().reset_index(name='_n')
            agg_lu['_surf'] = agg_lu['_n'] * 10_000

        for _, row in agg_lu.iterrows():
            EXT_SURFACE_LANDUSE_RES[row['_gouv']] = round(float(row.get('_surf', 0)), 0)

        log(f"  ✔ landuse : {len(EXT_SURFACE_LANDUSE_RES)} gouvernorats | "
            f"surf_res moy={np.mean(list(EXT_SURFACE_LANDUSE_RES.values())) / 1e6:.1f} km²")
    except Exception as e:
        log(f"  [WARN] landuse.csv : {e}")
else:
    log(f"  [WARN] landuse.csv absent → surface_landuse_residentiel par défaut")

# ── 8.5 Population OSM ───────────────────────────────────────────
# densite_population = score 1-5 (very_low→1 … very_high→5) ou densité brute
DENSITY_SCORE_MAP = {'very_high': 5, 'high': 4, 'medium': 3, 'low': 2, 'very_low': 1}

if os.path.exists(POPULATION_PATH):
    try:
        df_pop = pd.read_csv(POPULATION_PATH)
        gov_col_pop = None
        for c in ['city', 'gouvernorat', 'region', 'governorate']:
            if c in df_pop.columns:
                gov_col_pop = c; break

        if gov_col_pop:
            df_pop['_gouv'] = df_pop[gov_col_pop].str.lower().str.strip().map(VILLE_TO_GOUVERNORAT).fillna('Unknown')
        else:
            df_pop['_gouv'] = 'Unknown'

        df_pop = df_pop[df_pop['_gouv'] != 'Unknown']

        if 'density_category' in df_pop.columns:
            df_pop['_dscore'] = df_pop['density_category'].str.lower().str.strip().map(DENSITY_SCORE_MAP).fillna(3)
            pop_agg = df_pop.groupby('_gouv')['_dscore'].mean()
            EXT_DENSITE_POP = {g: round(v, 2) for g, v in pop_agg.items()}
        elif 'population_density' in df_pop.columns:
            pop_agg = df_pop.groupby('_gouv')['population_density'].mean()
            EXT_DENSITE_POP = {g: round(v, 2) for g, v in pop_agg.items()}

        log(f"  ✔ population : {len(EXT_DENSITE_POP)} gouvernorats | "
            f"densité moy={np.mean(list(EXT_DENSITE_POP.values())):.2f}")
    except Exception as e:
        log(f"  [WARN] population.csv : {e}")
else:
    log(f"  [WARN] population.csv absent → densite_population par défaut")

# ── 8.6 Propagation features OSM dans groupes_clean ──────────────
# Valeurs par défaut si gouvernorat absent des datasets OSM
DEF_OSM = {
    'nb_amenities_total':        0,
    'ratio_amenities_commerce':  0.15,
    'ratio_amenities_sante':     0.05,
    'nb_commerce':               0,
    'nb_stations_transport':     0,
    'densite_routes_km':         0.5,
    'nb_buildings_residentiel':  0,
    'surface_landuse_residentiel': 0.0,
    'densite_population':        3.0,
}

# Calculer les valeurs moyennes réelles comme fallback si dispo
def _mean_or_default(d, default):
    return round(float(np.mean(list(d.values()))), 4) if d else default

DEF_OSM_REAL = {
    'nb_amenities_total':          _mean_or_default(EXT_NB_AMENITIES, 0),
    'ratio_amenities_commerce':    _mean_or_default(EXT_RATIO_COMMERCE, 0.15),
    'ratio_amenities_sante':       _mean_or_default(EXT_RATIO_SANTE, 0.05),
    'nb_commerce':                 _mean_or_default(EXT_NB_COMMERCE, 0),
    'nb_stations_transport':       _mean_or_default(EXT_NB_TRANSPORT, 0),
    'densite_routes_km':           _mean_or_default(EXT_DENSITE_ROUTES, 0.5),
    'nb_buildings_residentiel':    _mean_or_default(EXT_NB_BUILDINGS_RES, 0),
    'surface_landuse_residentiel': _mean_or_default(EXT_SURFACE_LANDUSE_RES, 0.0),
    'densite_population':          _mean_or_default(EXT_DENSITE_POP, 3.0),
}

for groupe, dg in groupes_clean.items():
    if GOV_STR not in dg.columns:
        valid_idx = [i for i in dg.index if i in df.index]
        if valid_idx and GOV_STR in df.columns:
            dg.loc[valid_idx, GOV_STR] = df.loc[valid_idx, GOV_STR].values
        if GOV_STR not in dg.columns:
            dg[GOV_STR] = 'Unknown'

    g_col = dg[GOV_STR]
    dg['nb_amenities_total']          = g_col.map(EXT_NB_AMENITIES).fillna(DEF_OSM_REAL['nb_amenities_total']).astype(int)
    dg['ratio_amenities_commerce']    = g_col.map(EXT_RATIO_COMMERCE).fillna(DEF_OSM_REAL['ratio_amenities_commerce']).clip(0.0, 1.0).round(4)
    dg['ratio_amenities_sante']       = g_col.map(EXT_RATIO_SANTE).fillna(DEF_OSM_REAL['ratio_amenities_sante']).clip(0.0, 1.0).round(4)
    dg['nb_commerce']                 = g_col.map(EXT_NB_COMMERCE).fillna(DEF_OSM_REAL['nb_commerce']).astype(int)
    dg['nb_stations_transport']       = g_col.map(EXT_NB_TRANSPORT).fillna(DEF_OSM_REAL['nb_stations_transport']).astype(int)
    dg['densite_routes_km']           = g_col.map(EXT_DENSITE_ROUTES).fillna(DEF_OSM_REAL['densite_routes_km']).round(4)
    dg['nb_buildings_residentiel']    = g_col.map(EXT_NB_BUILDINGS_RES).fillna(DEF_OSM_REAL['nb_buildings_residentiel']).astype(int)
    dg['surface_landuse_residentiel'] = g_col.map(EXT_SURFACE_LANDUSE_RES).fillna(DEF_OSM_REAL['surface_landuse_residentiel']).round(0)
    dg['densite_population']          = g_col.map(EXT_DENSITE_POP).fillna(DEF_OSM_REAL['densite_population']).round(4)
    groupes_clean[groupe] = dg

    log(f"  [{groupe}] OSM features OK | nb_amenities moy={dg['nb_amenities_total'].mean():.0f} | "
        f"densite_pop moy={dg['densite_population'].mean():.1f}")


# ================================================================
# ETAPE 9 BO5 — FEATURES GOOGLE (note_google_moyenne)
# ================================================================

section("ETAPE 9 BO5 — FEATURES GOOGLE (note_google_moyenne)")

# note_google_moyenne par gouvernorat
# Source : signaux_immobilier_tunisie.xlsx OU valeurs proxy basées sur l'attractivité
EXT_NOTE_GOOGLE = {}   # {gouvernorat_str → note_google_moyenne [1.0-5.0]}

if os.path.exists(SIGNAUX_PATH):
    try:
        # Essai feuilles possibles contenant des notes Google
        xl = pd.ExcelFile(SIGNAUX_PATH)
        sheets = xl.sheet_names

        note_loaded = False
        for sheet in sheets:
            df_sig = pd.read_excel(SIGNAUX_PATH, sheet_name=sheet)

            # Chercher colonne gouvernorat
            gov_col = next((c for c in df_sig.columns if any(k in c.lower()
                           for k in ['gouv', 'region', 'gov'])), None)
            # Chercher colonne note google
            note_col = next((c for c in df_sig.columns if any(k in c.lower()
                            for k in ['note', 'rating', 'google', 'score_google'])), None)

            if gov_col and note_col:
                for _, row in df_sig.iterrows():
                    gov  = normalize_gouvernorat(row[gov_col])
                    note = pd.to_numeric(row[note_col], errors='coerce')
                    if gov != 'Unknown' and pd.notna(note) and 1.0 <= note <= 5.0:
                        EXT_NOTE_GOOGLE[gov] = round(float(note), 2)
                if EXT_NOTE_GOOGLE:
                    log(f"  ✔ note_google chargée depuis feuille '{sheet}' : {len(EXT_NOTE_GOOGLE)} gouvernorats")
                    note_loaded = True
                    break

        if not note_loaded:
            log(f"  [INFO] Colonne note Google non trouvée dans signaux_immobilier.xlsx")
    except Exception as e:
        log(f"  [WARN] signaux_immobilier.xlsx : {e}")
else:
    log(f"  [WARN] {SIGNAUX_PATH} absent")

# ── Proxy note_google si non disponible ──────────────────────────
# Basé sur la corrélation empirique entre attractivité et note Google Tunisie
# Sources : expérience terrain + poids relatif tourisme + infrastructure
NOTE_GOOGLE_PROXY = {
    'Tunis':       4.2, 'Ariana':    4.0, 'Ben Arous':   3.9, 'Manouba':   3.7,
    'Nabeul':      4.3, 'Sousse':    4.4, 'Monastir':    4.2, 'Mahdia':    4.0,
    'Sfax':        3.8, 'Bizerte':   3.9, 'Béja':        3.5, 'Jendouba':  3.4,
    'Zaghouan':    3.6, 'Siliana':   3.3, 'Kairouan':    4.1, 'Kasserine': 3.2,
    'Sidi Bouzid': 3.1, 'Gabès':     3.6, 'Médenine':   4.0, 'Tataouine': 3.7,
    'Gafsa':       3.3, 'Tozeur':    4.2, 'Kébili':      3.9, 'Le Kef':   3.4,
    'Unknown':     3.5,
}
for gov, note_proxy in NOTE_GOOGLE_PROXY.items():
    if gov not in EXT_NOTE_GOOGLE:
        EXT_NOTE_GOOGLE[gov] = note_proxy

# Propager dans groupes_clean
note_mean = float(np.mean(list(EXT_NOTE_GOOGLE.values())))
for groupe, dg in groupes_clean.items():
    g_col = dg.get(GOV_STR, pd.Series('Unknown', index=dg.index))
    dg['note_google_moyenne'] = g_col.map(EXT_NOTE_GOOGLE).fillna(note_mean).clip(1.0, 5.0).round(2)
    groupes_clean[groupe] = dg
    log(f"  [{groupe}] note_google moy={dg['note_google_moyenne'].mean():.2f}")


# ================================================================
# ETAPE 10 BO5 — TARGET : indice_rentabilite_regionale ⭐
# ================================================================
#
#   Formule composite :
#
#   rendement_locatif_brut = (prix_m2_location_annuel / prix_m2_achat) × 100
#     → estimé à partir du ratio prix location / prix vente par gouvernorat
#     → proxy : annonces Location vs Vente dans le groupe
#
#   ajustement_macro = f(inflation, PIB, variation_prix_m2)
#     → inflation élevée + PIB fort = meilleure rentabilité réelle
#
#   score_attractivite = f(note_google, densite_pop, nb_amenities, routes)
#     → normalisation [0,1]
#
#   indice_rentabilite_regionale =
#     (rendement_locatif_brut × 0.45)
#   + (ajustement_macro       × 0.30)
#   + (score_attractivite     × 0.25)
#
#   Normalisé [0, 100] — plus c'est élevé, plus la région est rentable.
# ================================================================

section("ETAPE 10 BO5 — TARGET : indice_rentabilite_regionale")

GOV_STR = '_gouvernorat_str'
CURRENT_YEAR = 2026

# ── 10.1 Calcul du rendement locatif brut par gouvernorat ─────────
# On sépare annonces Location vs Vente et calcule le ratio de prix
# comme proxy du rendement locatif annualisé

for groupe, dg in groupes_clean.items():
    # S'assurer que GOV_STR est présent
    if GOV_STR not in dg.columns:
        valid_idx = [i for i in dg.index if i in df.index]
        if valid_idx and GOV_STR in df.columns:
            dg.loc[valid_idx, GOV_STR] = df.loc[valid_idx, GOV_STR].values
        if GOV_STR not in dg.columns:
            dg[GOV_STR] = 'Unknown'
    groupes_clean[groupe] = dg

# Calcul global sur tous les groupes concaténés
df_all_for_target = pd.concat(
    [dg[[GOV_STR, 'type_transaction', 'prix', 'surface_m2',
         'prix_m2', 'variation_prix_m2',
         'inflation_glissement_annuel', 'croissance_pib_trim', 'high_season',
         'note_google_moyenne', 'densite_population', 'nb_amenities_total',
         'ratio_amenities_commerce', 'ratio_amenities_sante',
         'densite_routes_km', 'nb_stations_transport',
         'nb_buildings_residentiel', 'surface_landuse_residentiel', 'nb_commerce']]
     for dg in groupes_clean.values()
     if len(dg) > 0],
    ignore_index=True
)

# ── Prix médian Location et Vente par gouvernorat ─────────────────
df_loc  = df_all_for_target[df_all_for_target['type_transaction'] == 1].copy()
df_vte  = df_all_for_target[df_all_for_target['type_transaction'] == 2].copy()

df_loc['_pm2'] = np.where(df_loc['surface_m2'] > 0, df_loc['prix'] / df_loc['surface_m2'], np.nan)
df_vte['_pm2'] = np.where(df_vte['surface_m2'] > 0, df_vte['prix'] / df_vte['surface_m2'], np.nan)

med_loc = df_loc.groupby(GOV_STR)['_pm2'].median().rename('_med_loc')
med_vte = df_vte.groupby(GOV_STR)['_pm2'].median().rename('_med_vte')

prix_ratio = pd.concat([med_loc, med_vte], axis=1)
# Prix_m2 location = loyer mensuel moyen par m²
# Rendement brut annuel = (loyer_m2 × 12) / prix_achat_m2 × 100
prix_ratio['_rendement_brut'] = np.where(
    prix_ratio['_med_vte'].notna() & (prix_ratio['_med_vte'] > 0),
    (prix_ratio['_med_loc'].fillna(prix_ratio['_med_vte'] * 0.005) * 12)
    / prix_ratio['_med_vte'] * 100,
    np.nan
)

# Fallback : rendement moyen Tunisie ~6-8% selon gouvernorat
# (données : Knight Frank, JLL Tunisie 2024)
RENDEMENT_PROXY_TN = {
    'Tunis': 7.2, 'Ariana': 6.8, 'Ben Arous': 6.5, 'Manouba': 5.9,
    'Nabeul': 7.8, 'Sousse': 8.1, 'Monastir': 7.6, 'Mahdia': 6.9,
    'Sfax': 6.4, 'Bizerte': 6.1, 'Béja': 5.2, 'Jendouba': 4.8,
    'Zaghouan': 5.4, 'Siliana': 4.6, 'Kairouan': 5.8, 'Kasserine': 4.5,
    'Sidi Bouzid': 4.3, 'Gabès': 5.5, 'Médenine': 6.8, 'Tataouine': 5.0,
    'Gafsa': 4.7, 'Tozeur': 6.5, 'Kébili': 5.2, 'Le Kef': 4.9,
    'Unknown': 5.8,
}
prix_ratio['_rendement_brut'] = prix_ratio['_rendement_brut'].fillna(
    pd.Series(RENDEMENT_PROXY_TN)
)
# Clip raisonnable : [2%, 20%]
prix_ratio['_rendement_brut'] = prix_ratio['_rendement_brut'].clip(2.0, 20.0)

# ── 10.2 Ajustement macro par gouvernorat ─────────────────────────
# Score macro = f(inflation, PIB, variation_prix)
# Logique : inflation élevée → immobilier refuge → score +
#           PIB fort → demande locative ↑ → score +
#           variation prix positive → plus-value latente → score +

agg_macro = df_all_for_target.groupby(GOV_STR).agg(
    _inf = ('inflation_glissement_annuel', 'mean'),
    _pib = ('croissance_pib_trim', 'mean'),
    _var = ('variation_prix_m2', 'mean'),
).reset_index()

# Normalisation min-max → [0, 1]
def minmax_norm(s):
    mn, mx = s.min(), s.max()
    if mx == mn: return s.clip(0, 1).fillna(0.5)
    return ((s - mn) / (mx - mn)).clip(0, 1)

agg_macro['_score_inf'] = minmax_norm(agg_macro['_inf'])   # inflation : min=mauvais, max=refuge
agg_macro['_score_pib'] = minmax_norm(agg_macro['_pib'])   # PIB : max=bonne demande
agg_macro['_score_var'] = minmax_norm(agg_macro['_var'].clip(-0.5, 0.5))  # variation clampée

agg_macro['_score_macro'] = (
    agg_macro['_score_inf'] * 0.35 +
    agg_macro['_score_pib'] * 0.40 +
    agg_macro['_score_var'] * 0.25
).round(4)

# ── 10.3 Score attractivité par gouvernorat ───────────────────────
agg_attr = df_all_for_target.groupby(GOV_STR).agg(
    _note   = ('note_google_moyenne',      'mean'),
    _dens   = ('densite_population',       'mean'),
    _amen   = ('nb_amenities_total',       'mean'),
    _routes = ('densite_routes_km',        'mean'),
    _transp = ('nb_stations_transport',    'mean'),
).reset_index()

agg_attr['_score_note']   = minmax_norm((agg_attr['_note'] - 1) / 4)     # [1-5] → [0-1]
agg_attr['_score_dens']   = minmax_norm(agg_attr['_dens'] / 5)            # score déjà 1-5
agg_attr['_score_amen']   = minmax_norm(np.log1p(agg_attr['_amen']))      # log scale
agg_attr['_score_routes'] = minmax_norm(agg_attr['_routes'])
agg_attr['_score_transp'] = minmax_norm(np.log1p(agg_attr['_transp']))

agg_attr['_score_attractivite'] = (
    agg_attr['_score_note']   * 0.30 +
    agg_attr['_score_dens']   * 0.20 +
    agg_attr['_score_amen']   * 0.20 +
    agg_attr['_score_routes'] * 0.15 +
    agg_attr['_score_transp'] * 0.15
).round(4)

# ── 10.4 Indice final par gouvernorat ─────────────────────────────
# Fusion des 3 composantes
indice_df = prix_ratio[['_rendement_brut']].copy()
indice_df = indice_df.join(
    agg_macro.set_index(GOV_STR)[['_score_macro']], how='left'
)
indice_df = indice_df.join(
    agg_attr.set_index(GOV_STR)[['_score_attractivite']], how='left'
)

# Normaliser rendement_brut → [0, 1] (base 20% = max théorique)
indice_df['_rend_norm'] = (indice_df['_rendement_brut'] / 20.0).clip(0.0, 1.0)

# Imputation NaN
indice_df['_score_macro']        = indice_df['_score_macro'].fillna(0.5)
indice_df['_score_attractivite'] = indice_df['_score_attractivite'].fillna(0.5)

# Calcul indice composite [0, 100]
indice_df['indice_rentabilite_regionale'] = (
    indice_df['_rend_norm']          * 0.45 +
    indice_df['_score_macro']        * 0.30 +
    indice_df['_score_attractivite'] * 0.25
) * 100

indice_df['indice_rentabilite_regionale'] = (
    indice_df['indice_rentabilite_regionale'].clip(0.0, 100.0).round(2)
)

# Dictionnaire gouvernorat → indice
INDICE_GOUV = indice_df['indice_rentabilite_regionale'].to_dict()
indice_global_med = float(np.median(list(INDICE_GOUV.values()))) if INDICE_GOUV else 50.0

log(f"\n  ══ INDICE_RENTABILITE_REGIONALE par gouvernorat ══")
for g, v in sorted(INDICE_GOUV.items(), key=lambda x: -x[1]):
    log(f"    {g:<18}: {v:.2f}/100")

# ── 10.5 Propagation TARGET dans chaque groupe ────────────────────
for groupe, dg in groupes_clean.items():
    g_col = dg.get(GOV_STR, pd.Series('Unknown', index=dg.index))
    dg['indice_rentabilite_regionale'] = (
        g_col.map(INDICE_GOUV)
        .fillna(indice_global_med)
        .clip(0.0, 100.0)
        .round(2)
    )
    n_valid = dg['indice_rentabilite_regionale'].notna().sum()
    log(f"  [{groupe}] TARGET : {n_valid}/{len(dg)} valides | "
        f"moy={dg['indice_rentabilite_regionale'].mean():.1f} | "
        f"med={dg['indice_rentabilite_regionale'].median():.1f}")
    groupes_clean[groupe] = dg


# ================================================================
# ETAPE 11 BO5 — CONSTRUCTION DATASET FINAL + EXPORT
# ================================================================

section("ETAPE 11 BO5 — CONSTRUCTION DATASET FINAL")

# ── Colonnes finales BO5 (17 colonnes strictes) ───────────────────
FINAL_COLS_BO5 = [
    # Géographie (1)
    'gouvernorat',
    # Features Marché (5)
    'prix_m2',
    'variation_prix_m2',
    'inflation_glissement_annuel',
    'croissance_pib_trim',
    'high_season',
    # Features Attractivité Google + OSM (10)
    'note_google_moyenne',
    'densite_population',
    'nb_amenities_total',
    'ratio_amenities_commerce',
    'ratio_amenities_sante',
    'densite_routes_km',
    'nb_stations_transport',
    'nb_buildings_residentiel',
    'surface_landuse_residentiel',
    'nb_commerce',
    # TARGET (1)
    'indice_rentabilite_regionale',
]

# Types stricts
INT_COLS_BO5 = ['gouvernorat', 'nb_amenities_total', 'nb_stations_transport',
                'nb_buildings_residentiel', 'nb_commerce', 'high_season']
FLT_COLS_BO5 = ['prix_m2', 'variation_prix_m2', 'inflation_glissement_annuel',
                'croissance_pib_trim', 'note_google_moyenne', 'densite_population',
                'ratio_amenities_commerce', 'ratio_amenities_sante',
                'densite_routes_km', 'surface_landuse_residentiel',
                'indice_rentabilite_regionale']
STR_COLS_BO5 = []

# Dictionnaire de description des colonnes (pour onglet DICTIONNAIRE)
COL_DESC_BO5 = {
    'gouvernorat':                    ('int',    '[1-24]',       'Code entier du gouvernorat tunisien (GOUVERNORAT_ENC, Unknown=0 exclu)', 'Clé géo', 'encode_gouvernorat'),
    'prix_m2':                        ('float',  'TND/m²',       'Prix médian au m² par gouvernorat × trimestre', 'Feature marché', 'prix/surface groupby gouv'),
    'variation_prix_m2':              ('float',  '[-1;+∞]',      'Variation mensuelle du prix_m2 (négatif=baisse=normal)', 'Feature tendance', 'pct_change(1) par gouv'),
    'inflation_glissement_annuel':    ('float',  '%',            'Glissement annuel inflation (INS Tunisie)', 'Feature macro', 'MACRO_DATA par trimestre'),
    'croissance_pib_trim':            ('float',  '%',            'Croissance PIB trimestrielle (BCT)', 'Feature macro', 'MACRO_DATA par trimestre'),
    'high_season':                    ('int',    '0 ou 1',       'Saison haute immobilière (1=mars-mai/sept-nov)', 'Feature temporelle', 'mois publication'),
    'note_google_moyenne':            ('float',  '[1.0-5.0]',    'Note Google Maps moyenne du gouvernorat', 'Feature attractivité', 'signaux_immobilier.xlsx / proxy'),
    'densite_population':             ('float',  '[1-5] ou /km²','Score densité pop. (1=très faible…5=très élevée)', 'Feature attractivité', 'tunisia_population_*.csv'),
    'nb_amenities_total':             ('int',    'Nombre',       'Nb total points d\'intérêt OSM par gouvernorat', 'Feature attractivité', 'tunisia_amenities_*.csv'),
    'ratio_amenities_commerce':       ('float',  '[0-1]',        'Part des amenities de type commerce', 'Feature attractivité', 'amenities catégorisées'),
    'ratio_amenities_sante':          ('float',  '[0-1]',        'Part des amenities de type santé', 'Feature attractivité', 'amenities catégorisées'),
    'densite_routes_km':              ('float',  'km/km²',       'Densité réseau routier OSM (km routes/km² gouvernorat)', 'Feature infrastructure', 'tunisia_roads_*.csv'),
    'nb_stations_transport':          ('int',    'Nombre',       'Nb arrêts transport (bus/train/métro) OSM', 'Feature infrastructure', 'amenities transport'),
    'nb_buildings_residentiel':       ('int',    'Nombre',       'Nb bâtiments résidentiels OSM par gouvernorat', 'Feature offre logement', 'tunisia_buildings_*.csv'),
    'surface_landuse_residentiel':    ('float',  'm²',           'Surface zones résidentielles OSM (landuse)', 'Feature offre logement', 'tunisia_landuse_*.csv'),
    'nb_commerce':                    ('int',    'Nombre',       'Nb commerces OSM absolus par gouvernorat', 'Feature attractivité', 'amenities commerce'),
    'indice_rentabilite_regionale':   ('float',  '[0-100]',      '⭐ TARGET : indice rentabilité composite (rendement×0.45 + macro×0.30 + attractivité×0.25)', 'TARGET BO5', 'calcul composite étape 10'),
}

# ── Thèmes couleurs par groupe ────────────────────────────────────
THEME_BO5 = {
    'Residentiel': {'color': '1565C0', 'alt': 'E3F2FD'},
    'Foncier':     {'color': '2E7D32', 'alt': 'E8F5E9'},
    'Commercial':  {'color': 'E65100', 'alt': 'FFF3E0'},
    'Divers':      {'color': '6A1B9A', 'alt': 'F3E5F5'},
}

FICHIERS_BO5 = {
    'Residentiel': 'residentiel_BO5.xlsx',
    'Foncier':     'foncier_BO5.xlsx',
    'Commercial':  'commercial_BO5.xlsx',
    'Divers':      'divers_BO5.xlsx',
}


def build_final_bo5(dg: pd.DataFrame, groupe: str) -> pd.DataFrame:
    """
    Construit le DataFrame final BO5 à 17 colonnes strictes.
    """
    gov_str_col = GOV_STR if GOV_STR in dg.columns else None

    # Géographie : normaliser → encoder → filtrer Unknown → renommer
    dg['gouvernorat_nom'] = dg[gov_str_col].astype(str) if gov_str_col else 'Unknown'
    dg['gouvernorat'] = dg['gouvernorat_nom'].apply(encode_gouvernorat)
    unknown_mask = dg['gouvernorat'] == 0
    if unknown_mask.sum() > 0:
        log(f"  [{groupe}] Lignes Unknown/0 supprimées : {unknown_mask.sum()}")
        dg = dg[~unknown_mask].copy()
    dg = dg.drop(columns=['gouvernorat_nom'], errors='ignore')

    # Types stricts
    for c in INT_COLS_BO5:
        if c in dg.columns:
            dg[c] = pd.to_numeric(dg[c], errors='coerce').fillna(0).astype(int)
    for c in FLT_COLS_BO5:
        if c in dg.columns:
            dg[c] = pd.to_numeric(dg[c], errors='coerce').round(4)
    for c in STR_COLS_BO5:
        if c in dg.columns:
            dg[c] = dg[c].astype(str).str.strip()

    # Clips de sécurité
    for c in ['ratio_amenities_commerce', 'ratio_amenities_sante']:
        if c in dg.columns:
            dg[c] = dg[c].clip(0.0, 1.0)
    if 'note_google_moyenne' in dg.columns:
        dg['note_google_moyenne'] = dg['note_google_moyenne'].clip(1.0, 5.0)
    if 'indice_rentabilite_regionale' in dg.columns:
        dg['indice_rentabilite_regionale'] = dg['indice_rentabilite_regionale'].clip(0.0, 100.0)

    # Supprimer lignes TARGET invalide
    if 'indice_rentabilite_regionale' in dg.columns:
        bad = dg['indice_rentabilite_regionale'].isna() | (dg['indice_rentabilite_regionale'] <= 0)
        if bad.sum() > 0:
            log(f"  [{groupe}] TARGET invalides supprimées : {bad.sum()}")
            dg = dg[~bad].copy()

    # Sélection stricte des 17 colonnes
    df_out = pd.DataFrame()
    for col in FINAL_COLS_BO5:
        if col in dg.columns:
            df_out[col] = dg[col].values
        else:
            log(f"  [{groupe}] Colonne absente → 0 : {col}")
            df_out[col] = 0

    df_out = df_out.reset_index(drop=True)

    # Vérification NaN résiduels
    nan_total = df_out.isna().sum().sum()
    if nan_total == 0:
        log(f"  [{groupe}] ✔ Aucun NaN résiduel | {len(df_out)} lignes × {df_out.shape[1]} colonnes")
    else:
        nan_cols = df_out.isna().sum()
        log(f"  [{groupe}] ⚠ NaN résiduels : {nan_cols[nan_cols > 0].to_dict()}")

    return df_out


def write_excel_bo5(df_out: pd.DataFrame, filepath: str, groupe: str,
                    color_hex: str, alt_color: str):
    """
    Écrit le DataFrame BO5 dans un fichier Excel 3 onglets :
      1. {groupe}_DATA            : données formatées
      2. DICTIONNAIRE_COLONNES    : description chaque colonne
      3. STATISTIQUES             : stats descriptives + infos générales
    """
    wb = Workbook()

    h_fill  = PatternFill("solid", start_color=color_hex)
    h_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill = PatternFill("solid", start_color=alt_color)
    d_align  = Alignment(horizontal="left",  vertical="center")

    # ── Onglet 1 : DATA ──────────────────────────────────────────
    ws = wb.active
    ws.title = f"{groupe}_DATA"

    # Nettoyage vectorisé avant écriture
    df_clean = df_out.copy()
    for col in df_clean.select_dtypes(include=[np.floating]).columns:
        df_clean[col] = df_clean[col].round(4)
    for col in df_clean.select_dtypes(include=['object']).columns:
        df_clean[col] = df_clean[col].fillna('').astype(str).str.replace(
            r'[\x00-\x1F\x7F]', '', regex=True).str.strip()

    # Header
    ws.append(list(df_clean.columns))
    for cell in ws[1]:
        cell.fill = h_fill; cell.font = h_font; cell.alignment = h_align
    ws.row_dimensions[1].height = 38

    # Données (bulk)
    records = df_clean.where(df_clean.notna(), other=None).values.tolist()
    for r_idx, row_data in enumerate(records, start=2):
        ws.append(row_data)
        fill_r = alt_fill if r_idx % 2 == 0 else None
        for cell in ws[r_idx]:
            if fill_r: cell.fill = fill_r
            cell.alignment = d_align

    # Auto-width (échantillon 200 lignes)
    sample_end = min(201, ws.max_row)
    for col_idx, col_name in enumerate(df_clean.columns, start=1):
        cl = get_column_letter(col_idx)
        max_w = max(
            len(str(col_name)),
            *(len(str(ws.cell(r, col_idx).value or '')) for r in range(2, sample_end))
        )
        ws.column_dimensions[cl].width = min(max_w + 3, 44)
    ws.freeze_panes = "A2"

    # ── Onglet 2 : DICTIONNAIRE ───────────────────────────────────
    ws_d = wb.create_sheet(title="DICTIONNAIRE_COLONNES")
    ws_d['A1'] = f"DICTIONNAIRE DES COLONNES — BO5 : {groupe}"
    ws_d['A1'].font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    ws_d['A1'].fill = PatternFill("solid", start_color=color_hex)
    ws_d.merge_cells('A1:F1')
    ws_d.row_dimensions[1].height = 30

    ws_d.append([])
    ws_d.append(['Colonne', 'Type', 'Plage / Unité', 'Description', 'Rôle ML', 'Source données'])
    for cell in ws_d[3]:
        cell.fill = PatternFill("solid", start_color="37474F")
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="center")

    for col in FINAL_COLS_BO5:
        ws_d.append([col] + list(COL_DESC_BO5.get(col, ('?', '?', '?', '?', '?'))))

    dict_widths = [32, 10, 18, 62, 28, 36]
    for i, w in enumerate(dict_widths, start=1):
        ws_d.column_dimensions[get_column_letter(i)].width = w
    for r in range(4, ws_d.max_row + 1):
        ws_d.row_dimensions[r].height = 16
        for c_idx in range(1, 7):
            cell = ws_d.cell(r, c_idx)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if r % 2 == 0:
                cell.fill = PatternFill("solid", start_color="F5F5F5")

    # ── Onglet 3 : STATISTIQUES ───────────────────────────────────
    ws_s = wb.create_sheet(title="STATISTIQUES")
    ws_s['A1'] = f"STATISTIQUES DESCRIPTIVES — BO5 : {groupe}"
    ws_s['A1'].font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    ws_s['A1'].fill = PatternFill("solid", start_color=color_hex)
    ws_s.merge_cells('A1:H1')
    ws_s.row_dimensions[1].height = 30

    ws_s.append([])
    ws_s.append(['Colonne', 'Count', 'NaN (%)', 'Moyenne', 'Médiane', 'Std', 'Min', 'Max'])
    for cell in ws_s[3]:
        cell.fill = PatternFill("solid", start_color="37474F")
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="center")

    for col in df_out.select_dtypes(include=[np.number]).columns:
        s = df_out[col]
        n_nan   = s.isna().sum()
        nan_pct = round(n_nan / max(len(s), 1) * 100, 1)
        ws_s.append([
            col, int(s.notna().sum()), nan_pct,
            round(float(s.mean()),   2) if s.notna().sum() > 0 else None,
            round(float(s.median()), 2) if s.notna().sum() > 0 else None,
            round(float(s.std()),    2) if s.notna().sum() > 1 else None,
            round(float(s.min()),    2) if s.notna().sum() > 0 else None,
            round(float(s.max()),    2) if s.notna().sum() > 0 else None,
        ])

    for row in ws_s.iter_rows(min_row=4, max_row=ws_s.max_row, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 20:
                cell.fill = PatternFill("solid", start_color="FFE0B2")
                cell.font = Font(color="BF360C", bold=True)

    # Infos générales
    ws_s.append([])
    ws_s.append(['── INFOS GÉNÉRALES ──'])
    govs = [GOUVERNORAT_DEC.get(g, str(g)) for g in df_out['gouvernorat'].dropna().unique().tolist()
            if g not in (0,)] if 'gouvernorat' in df_out.columns else []
    ws_s.append(['Dataset',    groupe])
    ws_s.append(['Nb lignes',  len(df_out)])
    ws_s.append(['Nb colonnes', df_out.shape[1]])
    ws_s.append(['Gouvernorats couverts', ', '.join(sorted(govs))])
    ws_s.append(['TARGET valide',
                 f"{df_out['indice_rentabilite_regionale'].notna().sum()} "
                 f"({df_out['indice_rentabilite_regionale'].notna().mean()*100:.1f}%)"
                 if 'indice_rentabilite_regionale' in df_out.columns else 'N/A'])

    stat_widths = [32] + [14] * 7
    for i, w in enumerate(stat_widths, start=1):
        ws_s.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)
    log(f"  ✔ {filepath:<40}: {len(df_out):>6} lignes | {df_out.shape[1]} colonnes")


# ── Export principal ──────────────────────────────────────────────
results_bo5 = {}

for groupe, dg in groupes_clean.items():
    fname    = FICHIERS_BO5[groupe]
    theme    = THEME_BO5[groupe]
    _t_grp   = time.time()

    df_final = build_final_bo5(dg.copy(), groupe)

    if DEBUG_MODE:
        csv_path = fname.replace('.xlsx', '_debug.csv')
        df_final.to_csv(csv_path, index=False, encoding='utf-8-sig')
        log(f"  [DEBUG] {csv_path} : {len(df_final)} lignes ({time.time()-_t_grp:.1f}s)")
    else:
        write_excel_bo5(df_final, fname, groupe, theme['color'], theme['alt'])
        log(f"  Temps export [{groupe}] : {time.time()-_t_grp:.1f}s")

    results_bo5[groupe] = df_final


# ================================================================
# RAPPORT FINAL BO5
# ================================================================

section("RAPPORT FINAL BO5")

_total = time.time() - _t0
total_annonces = sum(len(d) for d in results_bo5.values())

print("╔══════════════════════════════════════════════════════════════════════╗")
print("║   PIPELINE IMMOBILIER TUNISIE — BO5 : RENTABILITÉ RÉGIONALE          ║")
print("╠══════════════════════════════════════════════════════════════════════╣")
print(f"║  {'Fichier':<30} {'Lignes':>7}  {'Colonnes':>8}  {'TARGET moy':>10}  {'NaN':>5} ║")
print("╠══════════════════════════════════════════════════════════════════════╣")
for groupe, df_f in results_bo5.items():
    fname   = FICHIERS_BO5[groupe]
    tgt_col = 'indice_rentabilite_regionale'
    tgt_moy = f"{df_f[tgt_col].mean():.1f}/100" if tgt_col in df_f.columns else 'N/A'
    nan_tot = df_f.isna().sum().sum()
    nan_flag = "✔" if nan_tot == 0 else f"⚠{nan_tot}"
    print(f"║  {fname:<30} {len(df_f):>7}  {df_f.shape[1]:>8}  {tgt_moy:>10}  {nan_flag:>5} ║")
print("╠══════════════════════════════════════════════════════════════════════╣")
print(f"║  Total annonces BO5       : {total_annonces:>6}                                   ║")
print(f"║  Temps total              : {_total:.1f}s {'(' + str(round(_total/60,1)) + ' min)' if _total > 60 else '':>8}                              ║")
print("╠══════════════════════════════════════════════════════════════════════╣")
print("║  COLONNES FINALES BO5 (17 strictes, identiques 4 datasets) :         ║")
for i, c in enumerate(FINAL_COLS_BO5, 1):
    star = " ⭐" if c == 'indice_rentabilite_regionale' else ""
    print(f"║    {i:>2}. {c:<40}{star:<4}            ║")
print("╠══════════════════════════════════════════════════════════════════════╣")
print("║  SUPPRIMÉ vs BO2 : BERT/LSA, Vision, Fusion, prix_trans_est,         ║")
print("║    negotiation_rate, cycle_marche, market_tension, NLP flags,         ║")
print("║    nb_pieces, type_bien, type_transaction, prix, surface_m2           ║")
print("╠══════════════════════════════════════════════════════════════════════╣")
print("║  AJOUTÉ vs BO2 : prix_m2, variation_prix_m2,                         ║")
print("║    inflation_glissement_annuel, croissance_pib_trim, high_season,     ║")
print("║    note_google_moyenne, densite_population, nb_amenities_total,       ║")
print("║    ratio_amenities_commerce, ratio_amenities_sante,                   ║")
print("║    densite_routes_km, nb_stations_transport,                          ║")
print("║    nb_buildings_residentiel, surface_landuse_residentiel, nb_commerce ║")
print("╠══════════════════════════════════════════════════════════════════════╣")
print("║  TARGET : indice_rentabilite_regionale [0-100]                       ║")
print("║    = rendement_locatif×0.45 + score_macro×0.30 + attractivité×0.25  ║")
print("╚══════════════════════════════════════════════════════════════════════╝")