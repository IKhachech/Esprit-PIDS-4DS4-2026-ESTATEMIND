"""
modeling_BO5.py — Monte Carlo et export pour l'Objectif 5 : Rentabilité Régionale.

Modèle Monte Carlo :
  - Simule N=1000 scénarios d'investissement par gouvernorat
  - Variables stochastiques : prix_m2, variation_prix_m2, inflation, croissance_pib
  - Calcule : rendement_espere, var_95, CVaR, sharpe_immo, probabilite_gain
  - Ajoute ces colonnes au dataset final

TARGET enrichi :
  indice_rentabilite_regionale (existant) +
  rendement_espere_mc, var_95_mc, cvar_mc, sharpe_mc, prob_gain_mc
"""

import os, re, json, warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

from mappings_BO5 import FICHIERS_BO5, FINAL_COLS_BO5, GOUVERNORAT_DEC


def _log(msg: str) -> None:
    print(f"  {msg}")


# ================================================================
# HELPER EXCEL (même style BO2/BO3)
# ================================================================

def _write_excel(df_out: pd.DataFrame, filename: str,
                 color_hex: str, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill = PatternFill("solid", start_color=color_hex)
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        clean = []
        for cell in r:
            if isinstance(cell, str):
                cell = cell.encode('utf-8', 'ignore').decode('utf-8')
                cell = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cell)
                cell = re.sub(r'\s+', ' ', cell).strip()
            elif isinstance(cell, bool):         cell = str(cell)
            elif isinstance(cell, np.integer):   cell = int(cell)
            elif isinstance(cell, np.floating):  cell = round(float(cell), 4) if not np.isnan(cell) else None
            clean.append(cell)
        ws.append(clean)

    for cell in ws[1]:
        cell.fill  = hfill
        cell.font  = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        w = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
    ws.freeze_panes = "A2"
    wb.save(filename)
    _log(f"  {filename:<45}: {len(df_out):>6} lignes | {len(df_out.columns)} colonnes")


# ================================================================
# MONTE CARLO — simulation par ligne
# ================================================================

def run_monte_carlo(groupes: dict, n_simulations: int = 1000,
                    seed: int = 42) -> dict:
    """
    Monte Carlo par gouvernorat :
    Pour chaque ligne, simule N scénarios de variation de prix
    et calcule les métriques de risque/rendement.

    Variables stochastiques (distribution normale) :
      - prix_m2         : ± 15% std autour de la valeur observée
      - variation_prix_m2 : ± 5% std
      - inflation       : ± 1% std
      - croissance_pib  : ± 0.5% std

    Métriques calculées :
      - rendement_espere_mc : rendement locatif espéré (%)
      - var_95_mc           : Value-at-Risk 95% (perte max probable)
      - cvar_mc             : CVaR/Expected Shortfall 95%
      - sharpe_mc           : ratio Sharpe immobilier simplifié
      - prob_gain_mc        : probabilité de rendement positif
    """
    print("\n" + "=" * 65)
    print("   ETAPE 6 — MONTE CARLO (N={:,} simulations)".format(n_simulations))
    print("=" * 65)

    rng = np.random.default_rng(seed)

    for nom, df in groupes.items():
        _log(f"Monte Carlo {nom}...")
        n = len(df)

        # Variables stochastiques — utiliser médiane si colonne absente
        prix_m2_obs   = df['prix_m2'].values
        variation_obs = df['variation_prix_m2'].values / 100
        # inflation et PIB : utiliser valeurs fixes si colonnes supprimées
        if 'inflation_glissement_annuel' in df.columns:
            inflation_obs = df['inflation_glissement_annuel'].values / 100
        else:
            inflation_obs = np.full(n, 0.063)   # 6.3% médiane INS 2020-2026
        if 'croissance_pib_trim' in df.columns:
            pib_obs = df['croissance_pib_trim'].values / 100
        else:
            pib_obs = np.full(n, 0.028)          # 2.8% médiane BCT 2020-2026
        attractivite  = df['indice_rentabilite_regionale'].values / 100

        # Simulation matricielle (n lignes × N_SIM)
        prix_sim   = prix_m2_obs[:, None] * (1 + rng.normal(0, 0.15, (n, n_simulations)))
        var_sim    = variation_obs[:, None] + rng.normal(0, 0.05, (n, n_simulations))
        infl_sim   = inflation_obs[:, None] + rng.normal(0, 0.01, (n, n_simulations))
        pib_sim    = pib_obs[:, None] + rng.normal(0, 0.005, (n, n_simulations))

        # Rendement locatif brut simulé (%)
        # rendement = (variation_prix + PIB) × attractivité - inflation
        rendement_sim = (var_sim + pib_sim) * attractivite[:, None] - infl_sim

        # Métriques
        rendement_espere = np.mean(rendement_sim, axis=1) * 100
        var_95           = np.percentile(rendement_sim, 5, axis=1) * 100   # perte 5%
        cvar             = rendement_sim[rendement_sim < np.percentile(
                               rendement_sim, 5, axis=1, keepdims=True)
                           ].reshape(n, -1).mean(axis=1) * 100
        std_sim          = np.std(rendement_sim, axis=1) * 100
        sharpe           = np.where(std_sim > 0, rendement_espere / std_sim, 0)
        prob_gain        = np.mean(rendement_sim > 0, axis=1) * 100

        df['rendement_espere_mc'] = np.round(rendement_espere, 4)
        df['var_95_mc']           = np.round(var_95, 4)
        df['cvar_mc']             = np.round(cvar, 4)
        df['sharpe_mc']           = np.round(sharpe, 4)
        df['prob_gain_mc']        = np.round(prob_gain, 2)

        groupes[nom] = df
        _log(f"  ✔ {nom:<15}: rendement_espere mean={rendement_espere.mean():.2f}% | "
             f"prob_gain mean={prob_gain.mean():.1f}% | VaR95 mean={var_95.mean():.2f}%")

    return groupes


# ================================================================
# EXPORT DATASETS (même style BO2/BO3)
# ================================================================

MC_COLS = ['rendement_espere_mc', 'var_95_mc', 'cvar_mc', 'sharpe_mc', 'prob_gain_mc']


def export_datasets(groupes: dict, output_dir: str = '.') -> None:
    """
    Exporte 2 sheets par fichier Excel :
      Sheet 'ML'  — 8 colonnes pour XGBoost / RandomForest / K-means
      Sheet 'MC'  — 8 + 5 colonnes Monte Carlo pour analyse risque/rendement

    Utilise un fichier temporaire pour éviter PermissionError
    si le fichier source est encore ouvert dans Excel.
    """
    print("\n" + "=" * 65)
    print("   ETAPE 7 — EXPORT FINAL (2 sheets par fichier)")
    print("=" * 65)

    os.makedirs(output_dir, exist_ok=True)

    for nom, df in groupes.items():
        fname, color = FICHIERS_BO5[nom]
        path     = os.path.join(output_dir, fname)
        path_tmp = path + '.tmp.xlsx'   # écriture temporaire

        # Colonnes ML = toutes sauf Monte Carlo
        cols_ml = [c for c in df.columns if c not in MC_COLS]
        # Colonnes MC = toutes
        cols_mc = df.columns.tolist()

        try:
            wb = Workbook()

            # ── Sheet 1 : ML ──────────────────────────────────────
            ws_ml = wb.active
            ws_ml.title = 'ML'
            hfill = PatternFill("solid", start_color=color)
            hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            df_ml = df[cols_ml]
            for r in dataframe_to_rows(df_ml, index=False, header=True):
                clean = []
                for cell in r:
                    if isinstance(cell, str):
                        cell = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cell)
                        cell = re.sub(r'\s+', ' ', cell).strip()
                    elif isinstance(cell, bool):        cell = str(cell)
                    elif isinstance(cell, np.integer):  cell = int(cell)
                    elif isinstance(cell, np.floating): cell = round(float(cell), 4) if not np.isnan(cell) else None
                    clean.append(cell)
                ws_ml.append(clean)
            for cell in ws_ml[1]:
                cell.fill  = hfill
                cell.font  = hfont
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws_ml.columns:
                w = max(len(str(c.value)) if c.value else 0 for c in col)
                ws_ml.column_dimensions[col[0].column_letter].width = min(w + 2, 35)
            ws_ml.freeze_panes = "A2"

            # ── Sheet 2 : Monte Carlo ──────────────────────────────
            ws_mc = wb.create_sheet(title='Monte_Carlo')
            mc_fill = PatternFill("solid", start_color="4A148C")
            df_mc = df[cols_mc]
            for r in dataframe_to_rows(df_mc, index=False, header=True):
                clean = []
                for cell in r:
                    if isinstance(cell, str):
                        cell = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cell)
                        cell = re.sub(r'\s+', ' ', cell).strip()
                    elif isinstance(cell, bool):        cell = str(cell)
                    elif isinstance(cell, np.integer):  cell = int(cell)
                    elif isinstance(cell, np.floating): cell = round(float(cell), 4) if not np.isnan(cell) else None
                    clean.append(cell)
                ws_mc.append(clean)
            for cell in ws_mc[1]:
                cell.fill  = mc_fill
                cell.font  = hfont
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws_mc.columns:
                w = max(len(str(c.value)) if c.value else 0 for c in col)
                ws_mc.column_dimensions[col[0].column_letter].width = min(w + 2, 35)
            ws_mc.freeze_panes = "A2"

            # ── Sauvegarde via fichier temporaire ──────────────────
            wb.save(path_tmp)

            # Supprimer l'ancien et renommer le tmp
            if os.path.exists(path):
                os.remove(path)
            os.rename(path_tmp, path)

            _log(f"  {fname:<40}: Sheet ML={len(cols_ml)} cols | Sheet MC={len(cols_mc)} cols | {len(df):,} lignes")

        except PermissionError:
            _log(f"  [ERREUR] {fname} est ouvert dans Excel — fermez le fichier et relancez.")
            if os.path.exists(path_tmp):
                os.remove(path_tmp)
            raise

        wb = Workbook()

        # ── Sheet 1 : ML (14 colonnes) ──────────────────────────
        ws_ml = wb.active
        ws_ml.title = 'ML'
        hfill = PatternFill("solid", start_color=color)
        hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        df_ml = df[cols_ml]
        for r in dataframe_to_rows(df_ml, index=False, header=True):
            clean = []
            for cell in r:
                if isinstance(cell, str):
                    cell = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cell)
                    cell = re.sub(r'\s+', ' ', cell).strip()
                elif isinstance(cell, bool):        cell = str(cell)
                elif isinstance(cell, np.integer):  cell = int(cell)
                elif isinstance(cell, np.floating): cell = round(float(cell), 4) if not np.isnan(cell) else None
                clean.append(cell)
            ws_ml.append(clean)
        for cell in ws_ml[1]:
            cell.fill  = hfill
            cell.font  = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws_ml.columns:
            w = max(len(str(c.value)) if c.value else 0 for c in col)
            ws_ml.column_dimensions[col[0].column_letter].width = min(w + 2, 35)
        ws_ml.freeze_panes = "A2"

        # Mappings JSON
    mappings = {
        'gouvernorat_decode': {str(k): v for k, v in GOUVERNORAT_DEC.items()},
        'colonnes_finales':   FINAL_COLS_BO5,
        'colonnes_mc':        ['rendement_espere_mc', 'var_95_mc',
                               'cvar_mc', 'sharpe_mc', 'prob_gain_mc'],
        'target':             'indice_rentabilite_regionale',
        'formule_target':     'rendement_locatif×0.45 + score_macro×0.30 + attractivité×0.25',
    }
    map_path = os.path.join(output_dir, 'encoding_mappings_BO5.json')
    with open(map_path, 'w', encoding='utf-8') as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)
    _log(f"  Mappings sauvegardés : encoding_mappings_BO5.json")


# ================================================================
# RAPPORT FINAL
# ================================================================

def print_final_report(groupes: dict) -> None:
    print("\n" + "=" * 65)
    print("   RAPPORT FINAL — OBJECTIF 5 : RENTABILITÉ RÉGIONALE")
    print("=" * 65)

    total = sum(len(df) for df in groupes.values())
    print(f"\n  {'Groupe':<15} {'Lignes':>8} {'Colonnes':>9} "
          f"{'TARGET mean':>12} {'Rendement MC':>13} {'Prob Gain':>10}")
    print(f"  {'-'*70}")

    for nom, df in groupes.items():
        tmean = df['indice_rentabilite_regionale'].mean()
        rmean = df['rendement_espere_mc'].mean() if 'rendement_espere_mc' in df.columns else 0
        pmean = df['prob_gain_mc'].mean() if 'prob_gain_mc' in df.columns else 0
        print(f"  {nom:<15} {len(df):>8,} {len(df.columns):>9} "
              f"{tmean:>12.2f} {rmean:>13.2f}% {pmean:>9.1f}%")

    print(f"  {'-'*70}")
    print(f"  {'TOTAL':<15} {total:>8,}")

    # Meilleurs gouvernorats
    print(f"\n  Meilleurs gouvernorats par rendement espéré :")
    all_df = pd.concat(groupes.values())
    if 'rendement_espere_mc' in all_df.columns:
        best = all_df.groupby('gouvernorat')['rendement_espere_mc'].mean().sort_values(ascending=False).head(5)
        for gov_code, rend in best.items():
            gov_nom = GOUVERNORAT_DEC.get(gov_code, str(gov_code))
            print(f"    {gov_nom:<20}: {rend:.2f}%")

    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   OBJECTIF 5 — DONNÉES PRÊTES POUR MODÉLISATION            ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  TARGET : indice_rentabilite_regionale [0-100]              ║")
    print("║  Monte Carlo : rendement_espere_mc, var_95_mc,             ║")
    print("║                cvar_mc, sharpe_mc, prob_gain_mc             ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  Modèles cibles :                                           ║")
    print("║    → XGBoost/RandomForest → prédire indice_rentabilite     ║")
    print("║    → Monte Carlo          → simulation risque/rendement    ║")
    print("║    → Clustering K-means   → profils régionaux              ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  Fichiers exportés :                                        ║")
    for nom, (fname, _) in FICHIERS_BO5.items():
        print(f"║    {fname:<56}║")
    print("╚══════════════════════════════════════════════════════════════╝")
