"""
INS Tunisia – Scraper LÉGAL 100% (Sources Officielles Uniquement)
==================================================================
Sources utilisées — TOUTES officielles et légalement libres :

  1. middleoffice-api.ins.tn  → API publique INS (découverte par diagnostic)
     ✅ Exposée volontairement par INS pour ses propres graphiques
     ✅ Aucune authentification requise = données publiques
     ✅ Couverture légale : décret-loi n°2011-41 (accès aux documents publics)

  2. Highcharts sur www.ins.tn → Données des graphiques interactifs
     ✅ Données publiques affichées sur le site officiel INS
     ✅ Même données que ce que n'importe quel visiteur voit

  3. dataportal.ins.tn        → Portail Open Data officiel INS + Banque Africaine
     ✅ Créé expressément pour la réutilisation des données

CITATION OBLIGATOIRE dans ton app :
  "Source : Institut National de la Statistique – www.ins.tn"

Usage :
    pip install playwright pandas openpyxl requests
    playwright install chromium
    python ins_officiel.py
"""

import json
import re
import requests
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, Page

# ══════════════════════════════════════════════════════════════
# CONFIG — SOURCES OFFICIELLES
# ══════════════════════════════════════════════════════════════

# Source 1 : API officielle INS (middleoffice = backend du site INS lui-même)
MIDDLEOFFICE_API = "https://middleoffice-api.ins.tn/data/find"

# IDs découverts par le diagnostic (tous publics, exposés par INS)
API_IDS = {
    "Taux d'Inflation":                1620305724,
    "Balance Commerciale (Imports)":   1675780959,
    "Indices Prix Immobilier":         1637313917,
    "Population au 1er Janvier":       1616749209,
    "Taux de Chômage par Genre":       1611322839,
    "PIB / Croissance Économique":     1684150984,
}

# Source 2 : Portail Open Data officiel INS
DATAPORTAL = "http://dataportal.ins.tn"

# Source 3 : Portail national open data
DATA_GOV_TN = "https://www.data.gov.tn"

# Couleurs Excel
C_DARK   = "1F4E79"
C_MID    = "2E75B6"
C_LIGHT  = "D6E4F0"
C_GREEN  = "E2EFDA"
C_AMBER  = "FFF2CC"
C_ORANGE = "FCE4D6"
C_PURPLE = "EAD1F5"
C_WHITE  = "FFFFFF"
C_GRAY   = "F5F5F5"

SOURCE_LABEL = "Institut National de la Statistique – www.ins.tn"

# ══════════════════════════════════════════════════════════════
# HELPERS EXCEL
# ══════════════════════════════════════════════════════════════

def _border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(c):  return PatternFill("solid", fgColor=c)
def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def title_row(ws, text, ncols, row=1, bg=C_DARK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, size=13, color="FFFFFF")
    c.fill = _fill(bg); c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 30

def subtitle_row(ws, text, ncols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 16

def header_row(ws, row, cols, bg=C_DARK):
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = _fill(bg); c.alignment = _align(h="center"); c.border = _border()
    ws.row_dimensions[row].height = 22

def write_df(ws, df, start_row, alt=C_LIGHT):
    for ri, row in enumerate(df.itertuples(index=False), start=start_row):
        bg = alt if ri % 2 == 0 else C_WHITE
        ws.row_dimensions[ri].height = 18
        for ci, val in enumerate(row, 1):
            col = df.columns[ci - 1]
            h   = "right" if pd.api.types.is_numeric_dtype(df[col]) else "left"
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font = _font(size=10); c.fill = _fill(bg)
            c.border = _border(); c.alignment = _align(h=h, v="center")
            if isinstance(val, float) and abs(val) < 10000:
                c.number_format = "0.00"
            elif isinstance(val, int) and abs(val) > 999:
                c.number_format = "#,##0"

def autowidth(ws, df):
    for ci, col in enumerate(df.columns, 1):
        w = max(len(str(col)),
                df[col].astype(str).str.len().max() if len(df) else 0) + 3
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 48)

def section_hdr(ws, row, ncols, text, bg=C_MID):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"  {text}")
    c.font = _font(bold=True, size=10, color="FFFFFF")
    c.fill = _fill(bg); c.alignment = _align(h="left"); c.border = _border()
    ws.row_dimensions[row].height = 20


# ══════════════════════════════════════════════════════════════
# SOURCE 1 — API middleoffice-api.ins.tn (via Playwright)
# ══════════════════════════════════════════════════════════════

def fetch_middleoffice_api(page: Page) -> dict[str, pd.DataFrame]:
    """
    Appelle chaque endpoint API depuis le contexte Playwright.
    Utilise fetch() JS pour éviter les problèmes CORS.
    Retourne un dict {nom_indicateur: DataFrame}.
    """
    print("\n[SOURCE 1] API officielle INS (middleoffice-api.ins.tn)")
    results = {}

    # Découvrir automatiquement les IDs appelés par la page
    live_ids = []
    def on_req(req):
        m = re.search(r'/data/find/(\d+)', req.url)
        if m:
            live_ids.append(int(m.group(1)))

    page.on("request", on_req)
    page.goto("https://www.ins.tn", wait_until="networkidle", timeout=30000)
    page.wait_for_timeout(4000)

    # Fusionner IDs connus + IDs découverts live
    all_ids = dict(API_IDS)
    for lid in live_ids:
        if lid not in all_ids.values():
            all_ids[f"Indicateur_{lid}"] = lid

    print(f"  → {len(all_ids)} endpoints API à appeler")

    for name, api_id in all_ids.items():
        url = f"{MIDDLEOFFICE_API}/{api_id}"
        try:
            raw = page.evaluate(f"""
            async () => {{
                const r = await fetch('{url}');
                if (!r.ok) throw new Error('HTTP ' + r.status);
                return await r.json();
            }}
            """)

            if not raw or not isinstance(raw, dict):
                print(f"  ✗ {name}: réponse invalide")
                continue

            # Chercher données dans toutes les langues
            vars_         = []
            official_name = name
            unit          = ""
            date_maj      = ""

            for lang_key in ["FR", "fr", "AR", "ar"]:
                lang_data = raw.get(lang_key)
                if not lang_data:
                    continue
                fix_data = lang_data.get("fix", {})
                ind_data = fix_data.get("Indicateurs", {})
                if not unit:
                    unit        = ind_data.get("IndicatorUnit", "")
                    date_maj    = ind_data.get("IndicatorTimeStimp", "")
                    ind_name    = ind_data.get("IndicatorName", "")
                    if lang_key in ("FR", "fr") and ind_name:
                        official_name = ind_name
                    elif ind_name and official_name == name:
                        official_name = ind_name
                v = lang_data.get("variable", [])
                if isinstance(v, list) and v:
                    vars_ = v
                    break
                elif isinstance(v, dict) and v:
                    vars_ = list(v.values())
                    break

            if not vars_:
                print(f"  ✗ {name}: données vides (clés JSON: {list(raw.keys())})")
                continue

            df = pd.DataFrame(vars_)
            df.columns = [str(c).strip() for c in df.columns]
            df["Unité"]        = unit
            df["MAJ_INS"]      = date_maj
            df["Source"]       = SOURCE_LABEL
            df["API_endpoint"] = url

            results[official_name] = df
            print(f"  ✓ \'{official_name}\' | {len(df)} lignes | {len(df.columns)} cols | MAJ: {date_maj}")

        except Exception as e:
            print(f"  ✗ {name} ({api_id}): {e}")

    return results


# ══════════════════════════════════════════════════════════════
# SOURCE 2 — Highcharts live (www.ins.tn)
# ══════════════════════════════════════════════════════════════

def fetch_highcharts(page: Page) -> dict[str, pd.DataFrame]:
    """
    Extrait les 6 graphiques Highcharts de la homepage INS.
    Page déjà chargée depuis fetch_middleoffice_api().
    """
    print("\n[SOURCE 2] Graphiques Highcharts live (www.ins.tn)")

    charts = page.evaluate("""
    () => {
        if (typeof Highcharts === 'undefined') return [];

        // Helper : extrait le texte de n'importe quel objet Highcharts title/subtitle
        function safeText(obj) {
            if (!obj) return '';
            if (typeof obj.textStr === 'string') return obj.textStr.trim();
            if (typeof obj.textStr !== 'undefined') return String(obj.textStr).trim();
            if (typeof obj.text === 'string') return obj.text.trim();
            return '';
        }

        const seen = {};
        return Highcharts.charts
            .filter(c => c && c.series && c.series.length > 0)
            .filter(c => {
                const t = safeText(c.title) || 'X';
                if (seen[t]) return false;
                seen[t] = true; return true;
            })
            .map(chart => {
                const cats = chart.xAxis[0] ? (chart.xAxis[0].categories || []) : [];
                const yAxis = chart.yAxis[0];
                const yLabel = yAxis && yAxis.axisTitle ? safeText(yAxis.axisTitle) : '';
                return {
                    title:      safeText(chart.title),
                    subtitle:   safeText(chart.subtitle),
                    y_label:    yLabel,
                    categories: cats,
                    series: chart.series.map(s => ({
                        name: s.name || '',
                        data: s.data.map((p, i) => ({
                            cat: cats[i] !== undefined ? cats[i] : i,
                            val: p.y
                        }))
                    }))
                };
            });
    }
    """) or []

    results = {}
    for chart in charts:
        title = chart["title"]
        rows  = []
        for serie in chart["series"]:
            for pt in serie["data"]:
                rows.append({
                    "Période":    pt["cat"],
                    "Série":      serie["name"],
                    "Valeur":     pt["val"],
                    "Unité":      chart.get("y_label", ""),
                    "Source":     SOURCE_LABEL,
                    "URL_source": "https://www.ins.tn",
                })
        if rows:
            df = pd.DataFrame(rows)
            # Pivot propre : une colonne par série
            try:
                pivot = df.pivot_table(
                    index=["Période", "Unité", "Source", "URL_source"],
                    columns="Série", values="Valeur", aggfunc="first"
                ).reset_index()
                pivot.columns.name = None
                results[title] = pivot
            except Exception:
                results[title] = df
            nb = len(rows)
            srs = len(chart["series"])
            print(f"  ✓ '{title}' | {srs} séries | {nb} points")

    return results


# ══════════════════════════════════════════════════════════════
# SOURCE 3 — dataportal.ins.tn (tentative Open Data officiel)
# ══════════════════════════════════════════════════════════════

def fetch_dataportal(page: Page) -> dict[str, pd.DataFrame]:
    """
    Tente de récupérer des données depuis dataportal.ins.tn.
    Ce portail utilise Prognoz — les données sont téléchargeables en JSON/CSV.
    """
    print("\n[SOURCE 3] dataportal.ins.tn (Open Data officiel INS)")
    results = {}

    endpoints = [
        ("http://dataportal.ins.tn/en/DataQuery", "DataQuery"),
        ("http://dataportal.ins.tn/fr/DataQuery", "DataQuery_FR"),
    ]

    for url, name in endpoints:
        try:
            page.goto(url, wait_until="networkidle", timeout=20000)
            page.wait_for_timeout(2000)
            html = page.content()
            dfs = pd.read_html(html, decimal=",", thousands=" ")
            for i, df in enumerate(dfs):
                if len(df) > 2:
                    df["Source"]     = SOURCE_LABEL
                    df["URL_source"] = url
                    results[f"{name}_table{i+1}"] = df
                    print(f"  ✓ {name} table {i+1}: {len(df)} lignes")
        except Exception as e:
            print(f"  ↩ {name}: {e}")

    if not results:
        print("  ↩ dataportal.ins.tn : aucune table HTML extraite (interface JS)")

    return results


# ══════════════════════════════════════════════════════════════
# GÉNÉRATION EXCEL — TOUTES LES DONNÉES VISIBLES
# ══════════════════════════════════════════════════════════════

def build_dataset(
    api_results: dict,
    hc_results:  dict,
) -> pd.DataFrame:
    """
    Fusionne toutes les données en un seul DataFrame normalisé.
    Colonnes : Indicateur | Série | Période | Valeur | Unité | Source | MAJ_INS | API_endpoint
    """
    all_rows = []

    # ── Données API middleoffice
    for ind_name, df in api_results.items():
        df = df.copy()
        # Trouver la colonne période et la colonne valeur
        period_col = next((c for c in df.columns if any(
            x in str(c).lower() for x in
            ["period", "mois", "trimest", "annee", "date", "time", "année"]
        )), None)
        val_cols = [c for c in df.columns if c not in
                    ["Unité", "MAJ_INS", "Source", "API_endpoint", period_col]
                    and pd.api.types.is_numeric_dtype(df[c])]

        unit     = df["Unité"].iloc[0]     if "Unité"        in df.columns else ""
        maj      = df["MAJ_INS"].iloc[0]   if "MAJ_INS"      in df.columns else ""
        endpoint = df["API_endpoint"].iloc[0] if "API_endpoint" in df.columns else ""

        for _, row in df.iterrows():
            period = str(row[period_col]) if period_col else ""
            for vc in val_cols:
                val = row[vc]
                if pd.isna(val):
                    continue
                all_rows.append({
                    "Indicateur":   ind_name,
                    "Série":        str(vc),
                    "Période":      period,
                    "Valeur":       val,
                    "Unité":        unit,
                    "Source":       SOURCE_LABEL,
                    "MAJ_INS":      maj,
                    "API_endpoint": endpoint,
                    "Type_source":  "API middleoffice-api.ins.tn",
                })

    # ── Données Highcharts
    for chart_name, df in hc_results.items():
        df = df.copy()
        period_col = "Période" if "Période" in df.columns else df.columns[0]
        skip_cols  = {"Période", "Unité", "Source", "URL_source"}
        val_cols   = [c for c in df.columns if c not in skip_cols
                      and pd.api.types.is_numeric_dtype(df[c])]
        unit       = df["Unité"].iloc[0] if "Unité" in df.columns else ""

        for _, row in df.iterrows():
            period = str(row[period_col])
            for vc in val_cols:
                val = row[vc]
                if pd.isna(val):
                    continue
                all_rows.append({
                    "Indicateur":   chart_name,
                    "Série":        str(vc),
                    "Période":      period,
                    "Valeur":       val,
                    "Unité":        unit,
                    "Source":       SOURCE_LABEL,
                    "MAJ_INS":      "",
                    "API_endpoint": "https://www.ins.tn (Highcharts)",
                    "Type_source":  "Highcharts live",
                })

    return pd.DataFrame(all_rows)


def build_excel(
    api_results: dict,
    hc_results:  dict,
    dataset_df:  pd.DataFrame,
    ts: str,
    output: str
):
    wb = Workbook()

    # ══════════════════════════════════════════════════════════
    # FEUILLE 1 : DATASET COMPLET (toutes données)
    # ══════════════════════════════════════════════════════════
    ws_data = wb.active
    ws_data.title = "📊 Dataset Complet"
    ws_data.sheet_view.showGridLines = False
    ncols = len(dataset_df.columns)

    title_row(ws_data, "INS TUNISIA — DATASET COMPLET (DONNÉES OFFICIELLES EN TEMPS RÉEL)", ncols)
    subtitle_row(ws_data, f"Source : middleoffice-api.ins.tn + Highcharts | Extrait le : {ts}", ncols)
    header_row(ws_data, 4, list(dataset_df.columns))
    write_df(ws_data, dataset_df, start_row=5, alt=C_LIGHT)
    autowidth(ws_data, dataset_df)
    print(f"  ✓ Dataset Complet : {len(dataset_df)} lignes × {ncols} colonnes")

    # ══════════════════════════════════════════════════════════
    # FEUILLE 2 : INFLATION
    # ══════════════════════════════════════════════════════════
    df_inf = dataset_df[dataset_df["Indicateur"].str.contains("INFLATION|Inflation|inflation", na=False)]
    if len(df_inf):
        ws = wb.create_sheet("💰 Inflation")
        ws.sheet_view.showGridLines = False
        title_row(ws, "TAUX D'INFLATION — INS TUNISIA", len(df_inf.columns), bg="B45309")
        subtitle_row(ws, f"Source : middleoffice-api.ins.tn | {ts}", len(df_inf.columns))
        header_row(ws, 4, list(df_inf.columns), bg="B45309")
        write_df(ws, df_inf.reset_index(drop=True), start_row=5, alt=C_AMBER)
        autowidth(ws, df_inf)
        print(f"  ✓ Inflation : {len(df_inf)} lignes")

    # ══════════════════════════════════════════════════════════
    # FEUILLE 3 : PIB / CROISSANCE
    # ══════════════════════════════════════════════════════════
    df_pib = dataset_df[dataset_df["Indicateur"].str.contains("PIB|CROISSANCE|Croissance|pib", na=False)]
    if len(df_pib):
        ws = wb.create_sheet("📈 PIB Croissance")
        ws.sheet_view.showGridLines = False
        title_row(ws, "PIB & CROISSANCE ÉCONOMIQUE — INS TUNISIA", len(df_pib.columns))
        subtitle_row(ws, f"Source : middleoffice-api.ins.tn | {ts}", len(df_pib.columns))
        header_row(ws, 4, list(df_pib.columns))
        write_df(ws, df_pib.reset_index(drop=True), start_row=5, alt=C_LIGHT)
        autowidth(ws, df_pib)
        print(f"  ✓ PIB/Croissance : {len(df_pib)} lignes")

    # ══════════════════════════════════════════════════════════
    # FEUILLE 4 : CHÔMAGE
    # ══════════════════════════════════════════════════════════
    df_cho = dataset_df[dataset_df["Indicateur"].str.contains("CH.MAGE|Chômage|chomage|EMPLOI", na=False)]
    if len(df_cho):
        ws = wb.create_sheet("👔 Chômage")
        ws.sheet_view.showGridLines = False
        title_row(ws, "TAUX DE CHÔMAGE — INS TUNISIA", len(df_cho.columns), bg="92400E")
        subtitle_row(ws, f"Source : middleoffice-api.ins.tn | {ts}", len(df_cho.columns))
        header_row(ws, 4, list(df_cho.columns), bg="92400E")
        write_df(ws, df_cho.reset_index(drop=True), start_row=5, alt=C_ORANGE)
        autowidth(ws, df_cho)
        print(f"  ✓ Chômage : {len(df_cho)} lignes")

    # ══════════════════════════════════════════════════════════
    # FEUILLE 5 : POPULATION
    # ══════════════════════════════════════════════════════════
    df_pop = dataset_df[dataset_df["Indicateur"].str.contains("POPULATION|Population|population", na=False)]
    if len(df_pop):
        ws = wb.create_sheet("👥 Population")
        ws.sheet_view.showGridLines = False
        title_row(ws, "POPULATION AU 1ER JANVIER — INS TUNISIA", len(df_pop.columns), bg="166534")
        subtitle_row(ws, f"Source : middleoffice-api.ins.tn | {ts}", len(df_pop.columns))
        header_row(ws, 4, list(df_pop.columns), bg="166534")
        write_df(ws, df_pop.reset_index(drop=True), start_row=5, alt=C_GREEN)
        autowidth(ws, df_pop)
        print(f"  ✓ Population : {len(df_pop)} lignes")

    # ══════════════════════════════════════════════════════════
    # FEUILLE 6 : IMMOBILIER
    # ══════════════════════════════════════════════════════════
    df_immo = dataset_df[dataset_df["Indicateur"].str.contains("IMMO|Immob|PRIX|Prix", na=False)]
    if len(df_immo):
        ws = wb.create_sheet("🏠 Immobilier")
        ws.sheet_view.showGridLines = False
        title_row(ws, "INDICES PRIX IMMOBILIER — INS TUNISIA", len(df_immo.columns), bg="5B21B6")
        subtitle_row(ws, f"Source : middleoffice-api.ins.tn | {ts}", len(df_immo.columns))
        header_row(ws, 4, list(df_immo.columns), bg="5B21B6")
        write_df(ws, df_immo.reset_index(drop=True), start_row=5, alt=C_PURPLE)
        autowidth(ws, df_immo)
        print(f"  ✓ Immobilier : {len(df_immo)} lignes")

    # ══════════════════════════════════════════════════════════
    # FEUILLE 7 : BALANCE COMMERCIALE
    # ══════════════════════════════════════════════════════════
    df_bal = dataset_df[dataset_df["Indicateur"].str.contains("BALANCE|Balance|balance|IMPORT|EXPORT", na=False)]
    if len(df_bal):
        ws = wb.create_sheet("🔄 Balance")
        ws.sheet_view.showGridLines = False
        title_row(ws, "BALANCE COMMERCIALE — INS TUNISIA", len(df_bal.columns), bg="1E40AF")
        subtitle_row(ws, f"Source : middleoffice-api.ins.tn | {ts}", len(df_bal.columns))
        header_row(ws, 4, list(df_bal.columns), bg="1E40AF")
        write_df(ws, df_bal.reset_index(drop=True), start_row=5, alt=C_PURPLE)
        autowidth(ws, df_bal)
        print(f"  ✓ Balance commerciale : {len(df_bal)} lignes")

    # ══════════════════════════════════════════════════════════
    # FEUILLE 8 : INDEX & BASE LÉGALE
    # ══════════════════════════════════════════════════════════
    ws_idx = wb.create_sheet("📋 Sources & Légalité")
    ws_idx.sheet_view.showGridLines = False
    title_row(ws_idx, "🇹🇳 INS TUNISIA — SOURCES OFFICIELLES & BASE LÉGALE", 4)
    subtitle_row(ws_idx, f"Extraction automatique | {ts}", 4)

    section_hdr(ws_idx, 4, 4, "RÉSUMÉ DU DATASET")
    info_rows = [
        ("Total lignes dataset",    len(dataset_df)),
        ("Indicateurs distincts",   dataset_df["Indicateur"].nunique()),
        ("Séries distinctes",       dataset_df["Série"].nunique()),
        ("Périodes couvertes",      dataset_df["Période"].nunique()),
        ("Date d'extraction",      ts),
        ("Source principale",       "middleoffice-api.ins.tn"),
        ("Source secondaire",       "www.ins.tn (Highcharts)"),
    ]
    header_row(ws_idx, 5, ["Métadonnée", "Valeur", "", ""])
    for ri, (label, val) in enumerate(info_rows, start=6):
        bg = C_LIGHT if ri % 2 == 0 else C_WHITE
        for ci, v in enumerate([label, val, "", ""], 1):
            c = ws_idx.cell(row=ri, column=ci, value=v)
            c.font = _font(size=10, bold=(ci==1))
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align(h="left")
        ws_idx.row_dimensions[ri].height = 18

    section_hdr(ws_idx, 14, 4, "BASE LÉGALE")
    legals = [
        "✅ Décret-loi n°2011-41 — Accès aux documents des organismes publics tunisiens",
        "✅ Article 32 Constitution tunisienne — Droit d'accès à l'information",
        "✅ API publique non protégée — exposée volontairement par INS",
        "✅ Données affichées publiquement sur www.ins.tn (Highcharts)",
        f"📌 Citation obligatoire : « Source : {SOURCE_LABEL} »",
    ]
    for ri, txt in enumerate(legals, start=15):
        ws_idx.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
        c = ws_idx.cell(row=ri, column=1, value=f"  {txt}")
        c.font = _font(size=10, bold=("📌" in txt))
        c.fill = _fill(C_GREEN if "✅" in txt else C_AMBER)
        c.border = _border()
        c.alignment = _align(h="left")
        ws_idx.row_dimensions[ri].height = 20

    for i, w in enumerate([40, 30, 5, 5], 1):
        ws_idx.column_dimensions[get_column_letter(i)].width = w

    wb.save(output)
    total = len(wb.sheetnames)
    print(f"\n{'='*65}")
    print(f"  ✅  {output}")
    print(f"  {total} feuilles | {len(dataset_df)} lignes de données réelles")
    print(f"  Citation : « Source : {SOURCE_LABEL} »")
    print(f"{'='*65}")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def main():
    ts     = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    output = f"ins_dataset_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    print("=" * 65)
    print("  INS Tunisia — Scraper Sources Officielles | Dataset Excel")
    print(f"  Base légale : décret-loi n°2011-41 (Open Data Tunisie)")
    print(f"  Démarré : {ts}")
    print("=" * 65)

    api_results = {}
    hc_results  = {}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )
        ctx  = browser.new_context(
            locale="fr-TN",
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 Chrome/122 Safari/537.36"
        )
        page = ctx.new_page()

        api_results = fetch_middleoffice_api(page)
        hc_results  = fetch_highcharts(page)

        browser.close()

    # Fusionner en dataset unique
    print("\n[DATASET] Fusion de toutes les données ...")
    dataset_df = build_dataset(api_results, hc_results)
    print(f"  ✓ {len(dataset_df)} lignes | {dataset_df['Indicateur'].nunique()} indicateurs")

    # Générer Excel
    print("\n[EXCEL] Génération des feuilles ...")
    build_excel(api_results, hc_results, dataset_df, ts, output)


if __name__ == "__main__":
    main()