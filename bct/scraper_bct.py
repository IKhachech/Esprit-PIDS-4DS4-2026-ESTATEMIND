"""
BCT (Banque Centrale de Tunisie) — Scraper 100% Légal
======================================================
Source officielle : https://www.bct.gov.tn/bct/siteprod/index.jsp

Données extraites :
  - Taux directeur, TM, TMM, TRE
  - Cours moyens des devises (USD, EUR, GBP, JPY, CAD, MAD, LYD)
  - Indicateurs monétaires (avoirs en devises, Trésor, billets, refinancement)
  - Actualités BCT
  - Statistiques détaillées (cours.jsp, indicateurs.jsp)

Base légale :
  ✅ Site officiel gouvernemental public
  ✅ Décret-loi n°2011-41 (accès aux documents publics tunisiens)
  ✅ Données affichées publiquement sans restriction

Citation obligatoire : « Source : Banque Centrale de Tunisie – www.bct.gov.tn »

Usage :
    pip install requests beautifulsoup4 pandas openpyxl lxml
    python bct_scraper.py
"""

import re
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════

BCT_BASE    = "https://www.bct.gov.tn/bct/siteprod"
BCT_HOME    = f"{BCT_BASE}/index.jsp"
BCT_COURS   = f"{BCT_BASE}/cours.jsp"
BCT_INDICAT = f"{BCT_BASE}/indicateurs.jsp"
BCT_STATS   = f"{BCT_BASE}/stat_index.jsp"

SOURCE_LABEL = "Banque Centrale de Tunisie – www.bct.gov.tn"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 Chrome/122 Safari/537.36",
    "Accept-Language": "fr-TN,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Couleurs Excel
C_DARK   = "1A3A5C"   # bleu BCT
C_MID    = "2E6DA4"
C_LIGHT  = "D6E8F7"
C_GREEN  = "E2EFDA"
C_AMBER  = "FFF2CC"
C_ORANGE = "FCE4D6"
C_RED    = "FADADD"
C_WHITE  = "FFFFFF"
C_GRAY   = "F5F5F5"

# ══════════════════════════════════════════════════════════════
# HELPERS EXCEL
# ══════════════════════════════════════════════════════════════

def _b():
    s = Side(style="thin", color="A0C4E8")
    return Border(left=s, right=s, top=s, bottom=s)

def _f(c):  return PatternFill("solid", fgColor=c)
def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def title_row(ws, text, ncols, row=1, bg=C_DARK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, size=13, color="FFFFFF")
    c.fill = _f(bg); c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 30

def subtitle_row(ws, text, ncols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 16

def header_row(ws, row, cols, bg=C_DARK):
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = _f(bg); c.alignment = _align(h="center"); c.border = _b()
    ws.row_dimensions[row].height = 22

def data_row(ws, row, values, bg=C_WHITE, bolds=None):
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        bold = bolds and ci in bolds
        c.font = _font(size=10, bold=bold)
        c.fill = _f(bg); c.border = _b()
        c.alignment = _align(h="right" if isinstance(v, (int, float)) else "left")
    ws.row_dimensions[row].height = 18

def write_df(ws, df, start_row, alt=C_LIGHT):
    for ri, row in enumerate(df.itertuples(index=False), start=start_row):
        bg = alt if ri % 2 == 0 else C_WHITE
        ws.row_dimensions[ri].height = 18
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _font(size=10); c.fill = _f(bg)
            c.border = _b()
            c.alignment = _align(h="right" if isinstance(val, (int, float)) else "left")

def autowidth(ws, df):
    for ci, col in enumerate(df.columns, 1):
        w = max(len(str(col)),
                df[col].astype(str).str.len().max() if len(df) else 0) + 3
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 50)

def section_hdr(ws, row, ncols, text, bg=C_MID):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"  {text}")
    c.font = _font(bold=True, size=10, color="FFFFFF")
    c.fill = _f(bg); c.alignment = _align(h="left"); c.border = _b()
    ws.row_dimensions[row].height = 20


# ══════════════════════════════════════════════════════════════
# SCRAPING BCT HOMEPAGE
# ══════════════════════════════════════════════════════════════

def get_page(url: str) -> BeautifulSoup | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
        r.encoding = "utf-8"
        return BeautifulSoup(r.text, "lxml")
    except Exception as e:
        print(f"  ✗ Erreur {url}: {e}")
        return None


def scrape_taux_homepage(soup: BeautifulSoup) -> list[dict]:
    """Extrait les taux directeur, TM, TMM, TRE depuis la homepage."""
    print("  → Taux monétaires ...")
    rows = []

    # Chercher le module "Principaux Taux"
    modules = soup.find_all("div", class_="module")
    for module in modules:
        h3 = module.find("h3")
        if h3 and "Principaux Taux" in h3.get_text():
            spans = module.find_all("span")
            for span in spans:
                text = span.get_text(strip=True)
                if not text or "TUNIBOR" in text or "Courbe" in text:
                    continue
                # Parser : "Taux du marché monétaire (TM) au 18/02/2026: 6,99000 %"
                # Extraire date
                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', text)
                date = date_match.group(1) if date_match else ""
                # Extraire valeur dans <b>
                b_tag = span.find("b")
                valeur = b_tag.get_text(strip=True) if b_tag else ""
                # Extraire nom (avant "au" ou "du")
                nom = re.split(r'\s+au\s+|\s+du\s+', text)[0].strip()

                if valeur and nom:
                    try:
                        val_float = float(valeur.replace(",", "."))
                    except:
                        val_float = valeur

                    rows.append({
                        "Indicateur": nom,
                        "Valeur":     val_float,
                        "Unité":      "%",
                        "Date":       date,
                        "Source":     SOURCE_LABEL,
                        "URL":        BCT_HOME,
                    })
                    print(f"    ✓ {nom}: {valeur}%  ({date})")

    return rows


def scrape_devises_homepage(soup: BeautifulSoup) -> list[dict]:
    """Extrait les cours des devises depuis la homepage."""
    print("  → Cours des devises ...")
    rows = []

    # Trouver le module "COURS MOYENS DES DEVISES"
    modules = soup.find_all("div", class_="module")
    date_cours = ""
    for module in modules:
        h3 = module.find("h3")
        if h3 and "COURS MOYENS" in h3.get_text().upper():
            # Extraire date du titre
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', h3.get_text())
            date_cours = date_match.group(1) if date_match else datetime.now().strftime("%d/%m/%Y")

            # Chercher les lignes de devises (chaque div avec img + spans)
            for div in module.find_all("div"):
                img = div.find("img")
                spans = div.find_all("span")
                if img and len(spans) >= 3:
                    # span[0] = code devise, span[1] = unité, span[2] = valeur
                    code  = spans[0].get_text(strip=True).replace(":", "").strip()
                    unite = spans[1].get_text(strip=True).strip()
                    valeur_b = spans[2].find("b")
                    valeur = valeur_b.get_text(strip=True) if valeur_b else spans[2].get_text(strip=True)

                    if code and valeur and len(code) <= 5:
                        try:
                            val_float = float(valeur.replace(",", "."))
                        except:
                            val_float = valeur

                        rows.append({
                            "Devise":   code,
                            "Unité":    unite,
                            "Valeur_TND": val_float,
                            "Date":     date_cours,
                            "Source":   SOURCE_LABEL,
                            "URL":      BCT_HOME,
                        })
                        print(f"    ✓ {code} ({unite}) = {valeur} TND")

    return rows


def scrape_indicateurs_homepage(soup: BeautifulSoup) -> list[dict]:
    """Extrait les indicateurs monétaires (avoirs devises, Trésor, billets, refinancement)."""
    print("  → Indicateurs monétaires ...")
    rows = []

    modules = soup.find_all("div", class_="module")
    for module in modules:
        h3 = module.find("h3")
        if h3 and "Principaux Indicateurs" in h3.get_text():
            spans = module.find_all("span")
            for span in spans:
                text = span.get_text(" ", strip=True)
                if not text or "(*)" in text and len(text) < 10:
                    continue

                b_tags = span.find_all("b")
                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', text)
                date = date_match.group(1) if date_match else ""

                # Chercher toutes les valeurs numériques dans les <b>
                for b in b_tags:
                    val_text = b.get_text(strip=True)
                    try:
                        val_float = float(val_text.replace(",", ".").replace(" ", ""))
                        # Extraire le nom (texte avant la valeur)
                        nom_raw = re.split(r':\s*$|au\s+\d', text)[0].strip()
                        nom = re.sub(r'\s+', ' ', nom_raw)

                        rows.append({
                            "Indicateur": nom[:80],
                            "Valeur":     val_float,
                            "Unité":      "MDT",
                            "Date":       date,
                            "Source":     SOURCE_LABEL,
                            "URL":        BCT_HOME,
                        })
                        print(f"    ✓ {nom[:60]}: {val_float} MDT ({date})")
                    except:
                        continue

    return rows


def scrape_actualites_homepage(soup: BeautifulSoup) -> list[dict]:
    """Extrait les actualités BCT du slider homepage."""
    print("  → Actualités BCT ...")
    rows = []
    seen = set()

    # Chercher tous les slider-items
    items = soup.find_all("div", class_="slider-item")
    for item in items:
        date_div = item.find("div", class_="date")
        desc_div = item.find("div", class_="description")
        link_tag = item.find("a", class_="button")

        date = date_div.get_text(strip=True) if date_div else ""
        desc = desc_div.get_text(strip=True) if desc_div else ""
        link = link_tag.get("href", "") if link_tag else ""
        if link and not link.startswith("http"):
            link = f"{BCT_BASE}/{link}"

        key = (date, desc[:30])
        if desc and key not in seen:
            seen.add(key)
            rows.append({
                "Date":        date,
                "Titre":       desc,
                "Lien":        link,
                "Type":        "Actualité",
                "Source":      SOURCE_LABEL,
            })
            print(f"    ✓ [{date}] {desc[:60]}")

    return rows


# ══════════════════════════════════════════════════════════════
# SCRAPING PAGES DÉTAILLÉES BCT
# ══════════════════════════════════════════════════════════════

def flatten_columns(df):
    """Aplatit les colonnes multi-niveaux en colonnes simples."""
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            " ".join(str(v) for v in col if "Unnamed" not in str(v)).strip() or f"Col_{i}"
            for i, col in enumerate(df.columns)
        ]
    else:
        df.columns = [
            str(c) if "Unnamed" not in str(c) else f"Col_{i}"
            for i, c in enumerate(df.columns)
        ]
    df = df.dropna(how="all", axis=1)
    df = df.dropna(how="all", axis=0)
    return df


def scrape_cours_devises_detail():
    """Scrape la page cours.jsp pour l historique complet des cours."""
    print("  → Page cours détaillée (cours.jsp) ...")
    soup = get_page(BCT_COURS)
    if not soup:
        return []
    try:
        from io import StringIO
        dfs = pd.read_html(StringIO(str(soup)), decimal=",", thousands=" ")
        result = []
        for df in dfs:
            if len(df) > 2 and len(df.columns) >= 3:
                df = flatten_columns(df)
                df["Source"] = SOURCE_LABEL
                df["URL"]    = BCT_COURS
                result.append(df)
                print(f"    ✓ Table cours: {len(df)} lignes × {len(df.columns)} cols")
        return result
    except Exception as e:
        print(f"    ✗ {e}")
        return []


def scrape_indicateurs_detail():
    """Scrape la page indicateurs.jsp pour les indicateurs détaillés."""
    print("  → Page indicateurs détaillée (indicateurs.jsp) ...")
    soup = get_page(BCT_INDICAT)
    if not soup:
        return []
    try:
        from io import StringIO
        dfs = pd.read_html(StringIO(str(soup)), decimal=",", thousands=" ")
        result = []
        for df in dfs:
            if len(df) > 2 and len(df.columns) >= 2:
                df = flatten_columns(df)
                df["Source"] = SOURCE_LABEL
                df["URL"]    = BCT_INDICAT
                result.append(df)
                print(f"    ✓ Table indicateurs: {len(df)} lignes × {len(df.columns)} cols")
        return result
    except Exception as e:
        print(f"    ✗ {e}")
        return []


# ══════════════════════════════════════════════════════════════
# GÉNÉRATION EXCEL
# ══════════════════════════════════════════════════════════════

def build_excel(
    taux:       list[dict],
    devises:    list[dict],
    indicateurs: list[dict],
    actualites: list[dict],
    cours_dfs:  list[pd.DataFrame],
    indicat_dfs: list[pd.DataFrame],
    ts: str,
    output: str
):
    wb = Workbook()

    # ══ FEUILLE 1 : DASHBOARD (tout en un) ════════════════════
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard BCT"
    ws_dash.sheet_view.showGridLines = False

    title_row(ws_dash, "🏦 BANQUE CENTRALE DE TUNISIE — DONNÉES OFFICIELLES EN TEMPS RÉEL", 5)
    subtitle_row(ws_dash, f"Source : www.bct.gov.tn | Extrait le : {ts}", 5)

    row = 4

    # ── Taux monétaires
    section_hdr(ws_dash, row, 5, "PRINCIPAUX TAUX MONÉTAIRES")
    row += 1
    header_row(ws_dash, row, ["Indicateur", "Valeur (%)", "Date", "Source", "URL"], bg=C_DARK)
    row += 1
    for i, t in enumerate(taux):
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        data_row(ws_dash, row, [
            t["Indicateur"], t["Valeur"], t["Date"], t["Source"], t["URL"]
        ], bg=bg, bolds={2})
        row += 1

    row += 1

    # ── Cours devises
    section_hdr(ws_dash, row, 5, f"COURS MOYENS DES DEVISES — {devises[0]['Date'] if devises else ts}")
    row += 1
    header_row(ws_dash, row, ["Devise", "Unité", "Valeur (TND)", "Date", "Source"], bg=C_MID)
    row += 1
    for i, d in enumerate(devises):
        bg = C_AMBER if i % 2 == 0 else C_WHITE
        data_row(ws_dash, row, [
            d["Devise"], d["Unité"], d["Valeur_TND"], d["Date"], d["Source"]
        ], bg=bg, bolds={3})
        row += 1

    row += 1

    # ── Indicateurs monétaires
    section_hdr(ws_dash, row, 5, "INDICATEURS MONÉTAIRES & FINANCIERS")
    row += 1
    header_row(ws_dash, row, ["Indicateur", "Valeur (MDT)", "Date", "Unité", "Source"], bg=C_DARK)
    row += 1
    for i, ind in enumerate(indicateurs):
        bg = C_GREEN if i % 2 == 0 else C_WHITE
        data_row(ws_dash, row, [
            ind["Indicateur"], ind["Valeur"], ind["Date"], ind["Unité"], ind["Source"]
        ], bg=bg, bolds={2})
        row += 1

    for i, w in enumerate([55, 15, 14, 12, 40], 1):
        ws_dash.column_dimensions[get_column_letter(i)].width = w

    print(f"  ✓ Dashboard : {row} lignes")

    # ══ FEUILLE 2 : TAUX ══════════════════════════════════════
    if taux:
        df_taux = pd.DataFrame(taux)
        ws = wb.create_sheet("💰 Taux Monétaires")
        ws.sheet_view.showGridLines = False
        title_row(ws, "TAUX MONÉTAIRES BCT — DONNÉES OFFICIELLES", len(df_taux.columns))
        subtitle_row(ws, f"www.bct.gov.tn | {ts}", len(df_taux.columns))
        header_row(ws, 4, list(df_taux.columns))
        write_df(ws, df_taux, start_row=5, alt=C_LIGHT)
        autowidth(ws, df_taux)
        print(f"  ✓ Feuille Taux : {len(df_taux)} lignes")

    # ══ FEUILLE 3 : DEVISES ═══════════════════════════════════
    if devises:
        df_dev = pd.DataFrame(devises)
        ws = wb.create_sheet("💱 Cours Devises")
        ws.sheet_view.showGridLines = False
        title_row(ws, "COURS MOYENS DES DEVISES — BCT", len(df_dev.columns), bg=C_MID)
        subtitle_row(ws, f"www.bct.gov.tn | {ts}", len(df_dev.columns))
        header_row(ws, 4, list(df_dev.columns), bg=C_MID)
        write_df(ws, df_dev, start_row=5, alt=C_AMBER)
        autowidth(ws, df_dev)
        print(f"  ✓ Feuille Devises : {len(df_dev)} lignes")

    # ══ FEUILLE 4 : INDICATEURS ═══════════════════════════════
    if indicateurs:
        df_ind = pd.DataFrame(indicateurs)
        ws = wb.create_sheet("📈 Indicateurs")
        ws.sheet_view.showGridLines = False
        title_row(ws, "INDICATEURS MONÉTAIRES & FINANCIERS — BCT", len(df_ind.columns))
        subtitle_row(ws, f"www.bct.gov.tn | {ts}", len(df_ind.columns))
        header_row(ws, 4, list(df_ind.columns))
        write_df(ws, df_ind, start_row=5, alt=C_GREEN)
        autowidth(ws, df_ind)
        print(f"  ✓ Feuille Indicateurs : {len(df_ind)} lignes")

    # ══ FEUILLE 5 : COURS DÉTAILLÉ ════════════════════════════
    for i, df in enumerate(cours_dfs, 1):
        ws = wb.create_sheet(f"💱 Cours_Detail_{i}")
        ws.sheet_view.showGridLines = False
        title_row(ws, f"COURS DÉTAILLÉS DEVISES — BCT (Table {i})", len(df.columns))
        subtitle_row(ws, f"cours.jsp | {ts}", len(df.columns))
        header_row(ws, 4, list(df.columns))
        write_df(ws, df, start_row=5, alt=C_AMBER)
        autowidth(ws, df)
        print(f"  ✓ Cours détaillé table {i}: {len(df)} lignes")

    # ══ FEUILLE 6 : INDICATEURS DÉTAILLÉ ══════════════════════
    for i, df in enumerate(indicat_dfs, 1):
        ws = wb.create_sheet(f"📊 Indicat_Detail_{i}")
        ws.sheet_view.showGridLines = False
        title_row(ws, f"INDICATEURS DÉTAILLÉS — BCT (Table {i})", len(df.columns))
        subtitle_row(ws, f"indicateurs.jsp | {ts}", len(df.columns))
        header_row(ws, 4, list(df.columns))
        write_df(ws, df, start_row=5, alt=C_GREEN)
        autowidth(ws, df)
        print(f"  ✓ Indicateurs détaillé table {i}: {len(df)} lignes")

    # ══ FEUILLE 7 : ACTUALITÉS ════════════════════════════════
    if actualites:
        df_act = pd.DataFrame(actualites)
        ws = wb.create_sheet("📰 Actualités")
        ws.sheet_view.showGridLines = False
        title_row(ws, "ACTUALITÉS BCT — DERNIÈRES NOUVELLES", len(df_act.columns))
        subtitle_row(ws, f"www.bct.gov.tn | {ts}", len(df_act.columns))
        header_row(ws, 4, list(df_act.columns))
        write_df(ws, df_act, start_row=5, alt=C_LIGHT)
        autowidth(ws, df_act)
        print(f"  ✓ Actualités : {len(df_act)} lignes")

    # ══ FEUILLE 8 : DATASET COMPLET ═══════════════════════════
    all_rows = []
    for t in taux:
        all_rows.append({
            "Catégorie":  "Taux Monétaire",
            "Indicateur": t["Indicateur"],
            "Valeur":     t["Valeur"],
            "Unité":      "%",
            "Date":       t["Date"],
            "Source":     SOURCE_LABEL,
            "URL":        BCT_HOME,
        })
    for d in devises:
        all_rows.append({
            "Catégorie":  "Cours Devise",
            "Indicateur": f"1 {d['Unité']} {d['Devise']} = X TND",
            "Valeur":     d["Valeur_TND"],
            "Unité":      "TND",
            "Date":       d["Date"],
            "Source":     SOURCE_LABEL,
            "URL":        BCT_HOME,
        })
    for ind in indicateurs:
        all_rows.append({
            "Catégorie":  "Indicateur Monétaire",
            "Indicateur": ind["Indicateur"],
            "Valeur":     ind["Valeur"],
            "Unité":      "MDT",
            "Date":       ind["Date"],
            "Source":     SOURCE_LABEL,
            "URL":        BCT_HOME,
        })

    if all_rows:
        df_all = pd.DataFrame(all_rows)
        ws = wb.create_sheet("🗂️ Dataset Complet")
        ws.sheet_view.showGridLines = False
        title_row(ws, "BCT TUNISIA — DATASET COMPLET (TOUTES DONNÉES)", len(df_all.columns))
        subtitle_row(ws, f"Source : www.bct.gov.tn | {ts}", len(df_all.columns))
        header_row(ws, 4, list(df_all.columns))
        write_df(ws, df_all, start_row=5, alt=C_LIGHT)
        autowidth(ws, df_all)
        print(f"  ✓ Dataset complet : {len(df_all)} lignes")

    # ══ FEUILLE 9 : BASE LÉGALE ═══════════════════════════════
    ws_leg = wb.create_sheet("📋 Sources & Légalité")
    ws_leg.sheet_view.showGridLines = False
    title_row(ws_leg, "🏦 BCT TUNISIA — BASE LÉGALE & SOURCES", 4)
    subtitle_row(ws_leg, f"Extraction automatique | {ts}", 4)

    section_hdr(ws_leg, 4, 4, "SOURCES UTILISÉES")
    sources = [
        ("Homepage",    BCT_HOME,    "Taux, devises, indicateurs, actualités"),
        ("Cours",       BCT_COURS,   "Historique cours des devises"),
        ("Indicateurs", BCT_INDICAT, "Indicateurs monétaires détaillés"),
        ("Statistiques",BCT_STATS,   "Index des statistiques BCT"),
    ]
    header_row(ws_leg, 5, ["Section", "URL", "Données extraites", ""])
    for ri, (sec, url, data) in enumerate(sources, start=6):
        bg = C_LIGHT if ri % 2 == 0 else C_WHITE
        for ci, v in enumerate([sec, url, data, ""], 1):
            c = ws_leg.cell(row=ri, column=ci, value=v)
            c.font = _font(size=10); c.fill = _f(bg)
            c.border = _b(); c.alignment = _align(h="left")
        ws_leg.row_dimensions[ri].height = 18

    section_hdr(ws_leg, 11, 4, "BASE LÉGALE")
    legals = [
        "✅ Site officiel gouvernemental public de la BCT (banque centrale tunisienne)",
        "✅ Décret-loi n°2011-41 du 26 mai 2011 — Accès aux documents administratifs publics",
        "✅ Article 32 de la Constitution tunisienne — Droit d'accès à l'information",
        "✅ Données affichées publiquement, sans authentification, sans restriction",
        "📌 Citation obligatoire : « Source : Banque Centrale de Tunisie – www.bct.gov.tn »",
    ]
    for ri, txt in enumerate(legals, start=12):
        ws_leg.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
        c = ws_leg.cell(row=ri, column=1, value=f"  {txt}")
        c.font = _font(size=10, bold="📌" in txt)
        c.fill = _f(C_GREEN if "✅" in txt else C_AMBER)
        c.border = _b(); c.alignment = _align(h="left")
        ws_leg.row_dimensions[ri].height = 20

    for i, w in enumerate([18, 55, 45, 5], 1):
        ws_leg.column_dimensions[get_column_letter(i)].width = w

    wb.save(output)
    total = len(wb.sheetnames)
    print(f"\n{'='*65}")
    print(f"  ✅  {output}")
    print(f"  {total} feuilles | données officielles BCT Tunisia")
    print(f"  Citation : « Source : {SOURCE_LABEL} »")
    print(f"{'='*65}")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def main():
    ts     = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    output = f"bct_dataset_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    print("=" * 65)
    print("  BCT Tunisia — Scraper Sources Officielles")
    print(f"  Source : www.bct.gov.tn")
    print(f"  Base légale : décret-loi n°2011-41")
    print(f"  Démarré : {ts}")
    print("=" * 65)

    # ── Charger la homepage
    print("\n[1/5] Chargement homepage BCT ...")
    soup = get_page(BCT_HOME)
    if not soup:
        print("❌ Impossible de charger le site BCT. Vérifier la connexion.")
        return

    # ── Extraire depuis la homepage
    print("\n[2/5] Extraction données homepage ...")
    taux        = scrape_taux_homepage(soup)
    devises     = scrape_devises_homepage(soup)
    indicateurs = scrape_indicateurs_homepage(soup)
    actualites  = scrape_actualites_homepage(soup)

    # ── Pages détaillées
    print("\n[3/5] Pages détaillées ...")
    cours_dfs   = scrape_cours_devises_detail()
    indicat_dfs = scrape_indicateurs_detail()

    # ── Résumé
    print(f"\n[4/5] Résumé des données collectées :")
    print(f"  Taux monétaires   : {len(taux)} indicateurs")
    print(f"  Cours devises     : {len(devises)} devises")
    print(f"  Indicateurs BCT   : {len(indicateurs)} indicateurs")
    print(f"  Actualités        : {len(actualites)} articles")
    print(f"  Tables cours.jsp  : {len(cours_dfs)} tableaux")
    print(f"  Tables indicat.   : {len(indicat_dfs)} tableaux")

    # ── Générer Excel
    print(f"\n[5/5] Génération Excel ...")
    build_excel(taux, devises, indicateurs, actualites,
                cours_dfs, indicat_dfs, ts, output)


if __name__ == "__main__":
    main()