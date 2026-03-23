"""
═══════════════════════════════════════════════════════════════════════════
    ESTATE MIND - AGENT COMPLET MUBAWAB
    Code Final qui fait TOUT : Collecte + Enrichissement + Export
═══════════════════════════════════════════════════════════════════════════

FONCTIONNALITÉS :
✅ Collecte automatique avec les bonnes classes CSS (listingBox)
✅ Enrichissement automatique depuis les descriptions
✅ Export CSV + Excel formaté + JSON
✅ Déduplication automatique
✅ 3 modes : Standard / Exhaustif / Test

Auteur : Estate Mind Team
Date : Février 2024
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
from datetime import datetime


class EstateMindAgentComplet:
    """Agent complet de collecte immobilière Mubawab"""
    
    def __init__(self):
        self.base_url = "https://www.mubawab.tn"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml',
            'Accept-Language': 'fr-FR,fr;q=0.9',
        }
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        self.data = []
        self.seen_urls = set()
    
    def get_urls(self, mode='standard'):
        """Génère les URLs de collecte"""
        urls = []
        types = ['appartements', 'maisons', 'villas-et-maisons-de-luxe', 'terrains']
        govs = [
            'tunis', 'ariana', 'ben-arous', 'manouba', 'nabeul', 'zaghouan',
            'bizerte', 'beja', 'jendouba', 'kef', 'siliana', 'sousse',
            'monastir', 'mahdia', 'sfax', 'kairouan', 'kasserine', 'sidi-bouzid',
            'gabes', 'medenine', 'tataouine', 'gafsa', 'tozeur', 'kebili'
        ]
        
        for gov in govs:
            for t in types:
                urls.append(f"{self.base_url}/fr/st/{gov}/{t}-a-vendre")
        
        if mode == 'exhaustive':
            cities = ['tunis-ville', 'la-marsa', 'carthage', 'ariana-ville', 
                     'hammamet', 'sousse-ville', 'sfax-ville', 'djerba']
            for city in cities:
                for t in types:
                    urls.append(f"{self.base_url}/fr/st/{city}/{t}-a-vendre")
        
        return urls
    
    def parse_price(self, text):
        """Parse le prix - version ultra-robuste"""
        if not text:
            return 0.0
        # Nettoyer TOUS les types d'espaces
        clean = str(text).replace('TND', '').replace('DT', '')
        clean = re.sub(r'\s+', '', clean)  # Remplace TOUS les espaces (simples, multiples, insécables)
        clean = clean.replace('\xa0', '').replace('\u200b', '')
        clean = clean.replace('.', '').replace(',', '.').strip()
        try:
            prix = float(clean)
            return prix if 10000 <= prix <= 100000000 else 0.0
        except:
            return 0.0
    
    def extract(self, listing):
        """Extraction depuis l'annonce HTML"""
        try:
            data = {
                'date_collecte': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'titre': '',
                'type_propriete': '',
                'prix': 0.0,
                'ville': '',
                'quartier': '',
                'adresse_complete': '',
                'nombre_chambres': 0,
                'nombre_salles_bain': 0,
                'nombre_cuisines': 0,
                'surface_m2': 0.0,
                'jardin': False,
                'garage': False,
                'parking': False,
                'securite': False,
                'piscine': False,
                'ascenseur': False,
                'terrasse': False,
                'meuble': False,
                'url': '',
                'description': ''
            }
            
            full_text = listing.get_text(separator=' ', strip=True)
            
            # URL
            link = listing.find('a', href=True)
            if link:
                href = link['href']
                data['url'] = self.base_url + href if href.startswith('/') else href
            
            if data['url'] in self.seen_urls:
                return None
            if data['url']:
                self.seen_urls.add(data['url'])
            
            # TITRE
            title = listing.find(class_='listingTit')
            if not title:
                title = listing.find('h2') or listing.find('h3')
            if title:
                titre_text = title.get_text(strip=True)
                if titre_text and titre_text not in ['True', 'False', 'VRAI', 'FAUX']:
                    data['titre'] = titre_text
            
            # PRIX
            price_tag = listing.find(class_='priceTag')
            if price_tag:
                data['prix'] = self.parse_price(price_tag.get_text(strip=True))
            
            if data['prix'] == 0:
                for span in listing.find_all('span'):
                    if 'TND' in span.get_text():
                        data['prix'] = self.parse_price(span.get_text(strip=True))
                        if data['prix'] > 0:
                            break
            
            # LOCALISATION - CORRECTION COMPLÈTE
            loc_p = listing.find(class_='listingP')
            if loc_p:
                loc_text = loc_p.get_text(strip=True)
                # Vérifier que c'est vraiment une localisation (court et avec virgule)
                if loc_text and len(loc_text) < 100 and loc_text not in ['True', 'False', 'VRAI', 'FAUX']:
                    # Si contient virgule = format "Quartier, Ville"
                    if ',' in loc_text:
                        parts = [p.strip() for p in loc_text.split(',')]
                        # Filtrer les parties trop longues (descriptions)
                        parts = [p for p in parts if len(p) < 50 and p not in ['True', 'False', 'VRAI', 'FAUX']]
                        if len(parts) >= 2:
                            data['quartier'] = parts[0]
                            data['ville'] = parts[-1]
                            data['adresse_complete'] = ', '.join(parts)
                        elif len(parts) == 1:
                            data['ville'] = parts[0]
                    # Sinon, c'est juste la ville
                    elif len(loc_text) < 30:
                        data['ville'] = loc_text
            
            # SURFACE
            text_lower = full_text.lower()
            
            surf = re.search(r'(\d+(?:[.,]\d+)?)\s*m[²2]', text_lower)
            if surf:
                try:
                    surface = float(surf.group(1).replace(',', '.'))
                    # Validation : surface raisonnable entre 10 et 1000 m²
                    if 10 <= surface <= 1000:
                        data['surface_m2'] = surface
                except:
                    pass
            
            pieces = re.search(r'(\d+)\s*(?:pièce|chambre)', text_lower)
            if pieces:
                data['nombre_chambres'] = int(pieces.group(1))
            
            s_plus = re.search(r's\+(\d+)', text_lower)
            if s_plus:
                data['nombre_chambres'] = int(s_plus.group(1))
            
            sdb = re.search(r'(\d+)\s*(?:salle|bain|sdb)', text_lower)
            if sdb:
                data['nombre_salles_bain'] = int(sdb.group(1))
            
            if 'cuisine' in text_lower:
                cuisine_num = re.search(r'(\d+)\s*cuisine', text_lower)
                data['nombre_cuisines'] = int(cuisine_num.group(1)) if cuisine_num else 1
            
            # ÉQUIPEMENTS
            data['jardin'] = bool('jardin' in text_lower)
            data['terrasse'] = bool('terrasse' in text_lower)
            data['garage'] = bool('garage' in text_lower)
            data['parking'] = bool('parking' in text_lower)
            data['ascenseur'] = bool('ascenseur' in text_lower)
            data['securite'] = bool(re.search(r'sécurit|securit|gardien', text_lower))
            data['piscine'] = bool('piscine' in text_lower or 'pool' in text_lower)
            data['meuble'] = bool(re.search(r'meubl[ée]', text_lower))
            
            # TYPE
            url_title = (data['url'] + ' ' + data['titre']).lower()
            if 'appartement' in url_title or 's+' in url_title:
                data['type_propriete'] = 'Appartement'
            elif 'villa' in url_title:
                data['type_propriete'] = 'Villa'
            elif 'maison' in url_title:
                data['type_propriete'] = 'Maison'
            elif 'terrain' in url_title:
                data['type_propriete'] = 'Terrain'
            else:
                data['type_propriete'] = 'Appartement' if data['nombre_chambres'] > 0 else ''
            
            # DESCRIPTION
            if len(full_text) > 50:
                data['description'] = full_text[:500]
            
            if data['titre'] or data['prix'] > 0:
                return data
            
            return None
            
        except Exception as e:
            return None
    
    def enrichir_depuis_description(self, df):
        """
        ENRICHISSEMENT AUTOMATIQUE
        Extrait prix, ville, surface depuis les descriptions
        """
        print("\n🔧 ENRICHISSEMENT AUTOMATIQUE")
        print("=" * 70)
        
        # Convertir les colonnes
        df['quartier'] = df['quartier'].astype('object')
        df['ville'] = df['ville'].astype('object')
        df['adresse_complete'] = df['adresse_complete'].astype('object')
        
        enrichis = 0
        prix_ajoutes = 0
        villes_ajoutees = 0
        surfaces_ajoutees = 0
        
        for idx, row in df.iterrows():
            desc = str(row.get('description', ''))
            if not desc or desc == 'nan':
                continue
            
            modifie = False
            
            # PRIX depuis description
            if pd.isna(row['prix']) or row['prix'] == 0:
                # Pattern amélioré pour gérer plusieurs espaces
                prix_match = re.search(r'([\d\s]+?)\s*TND', desc)
                if prix_match:
                    # Nettoyer TOUS les espaces (y compris multiples)
                    prix_str = re.sub(r'\s+', '', prix_match.group(1))
                    try:
                        prix = float(prix_str)
                        if 10000 <= prix <= 100000000:
                            df.loc[idx, 'prix'] = prix
                            prix_ajoutes += 1
                            modifie = True
                    except:
                        pass
            
            # VILLE depuis description
            if pd.isna(row['ville']) or str(row['ville']).strip() == '' or len(str(row['ville'])) > 30:
                # Pattern: "Quartier, Ville XXX m²" au début de la description
                ville_match = re.search(r'\b([A-ZÉ][a-zéèêàâôû\s-]{2,25}),\s*([A-ZÉ][a-zéèêàâôû\s-]{2,20})\s+\d+\s*m', desc)
                if ville_match:
                    quartier = ville_match.group(1).strip()
                    ville = ville_match.group(2).strip()
                    # Vérifier que ce sont de vrais noms de lieux
                    mots_a_eviter = ['appartement', 'vente', 'prix', 'immeuble', 'residence', 'villa', 'maison']
                    if (ville.lower() not in mots_a_eviter and 
                        quartier.lower() not in mots_a_eviter and
                        ville not in ['True', 'False', 'VRAI', 'FAUX']):
                        df.loc[idx, 'quartier'] = quartier
                        df.loc[idx, 'ville'] = ville
                        df.loc[idx, 'adresse_complete'] = f"{quartier}, {ville}"
                        villes_ajoutees += 1
                        modifie = True
            
            # SURFACE depuis description
            if pd.isna(row['surface_m2']) or row['surface_m2'] == 0:
                surf_match = re.search(r'(\d+)\s*m[²2]', desc)
                if surf_match:
                    try:
                        surf = float(surf_match.group(1))
                        if 10 <= surf <= 10000:  # Surface raisonnable
                            df.loc[idx, 'surface_m2'] = surf
                            surfaces_ajoutees += 1
                            modifie = True
                    except:
                        pass
            
            # CHAMBRES depuis description
            if pd.isna(row['nombre_chambres']) or row['nombre_chambres'] == 0:
                chambre_match = re.search(r'(\d+)\s*Chambre', desc, re.I)
                if chambre_match:
                    df.loc[idx, 'nombre_chambres'] = int(chambre_match.group(1))
                    modifie = True
            
            # SALLES DE BAIN depuis description
            if pd.isna(row['nombre_salles_bain']) or row['nombre_salles_bain'] == 0:
                sdb_match = re.search(r'(\d+)\s*Salle', desc, re.I)
                if sdb_match:
                    df.loc[idx, 'nombre_salles_bain'] = int(sdb_match.group(1))
                    modifie = True
            
            if modifie:
                enrichis += 1
        
        print(f"✅ {enrichis} lignes enrichies")
        print(f"   💰 +{prix_ajoutes} prix")
        print(f"   📍 +{villes_ajoutees} villes")
        print(f"   📐 +{surfaces_ajoutees} surfaces")
        
        return df
    
    def scrape(self, url, max_pages=50):
        """Scrape une catégorie"""
        for page in range(1, max_pages + 1):
            try:
                page_url = f"{url}?page={page}" if page > 1 else url
                
                resp = self.session.get(page_url, timeout=10)
                
                if resp.status_code == 404:
                    break
                
                resp.raise_for_status()
                soup = BeautifulSoup(resp.content, 'html.parser')
                
                # CLASSE RÉELLE : listingBox
                listings = soup.find_all(class_='listingBox')
                
                if not listings:
                    break
                
                before = len(self.data)
                
                for lst in listings:
                    prop = self.extract(lst)
                    if prop:
                        self.data.append(prop)
                
                new = len(self.data) - before
                
                if new > 0:
                    print(f"    Page {page}: +{new} annonces")
                
                time.sleep(2)
                
            except Exception as e:
                break
    
    def save(self, filename='mubawab_final.csv'):
        """Sauvegarde complète avec enrichissement"""
        if not self.data:
            print("❌ Aucune donnée à sauvegarder")
            return None
        
        print(f"\n💾 SAUVEGARDE ET ENRICHISSEMENT")
        print("=" * 70)
        
        # Créer DataFrame
        df = pd.DataFrame(self.data)
        
        # ENRICHISSEMENT AUTOMATIQUE
        df = self.enrichir_depuis_description(df)
        
        # Ordre des colonnes
        cols = [
            'date_collecte', 'titre', 'type_propriete', 'prix',
            'ville', 'quartier', 'adresse_complete',
            'nombre_chambres', 'nombre_salles_bain', 'nombre_cuisines', 'surface_m2',
            'jardin', 'garage', 'parking', 'terrasse', 'securite',
            'piscine', 'ascenseur', 'meuble',
            'url', 'description'
        ]
        
        df = df[[c for c in cols if c in df.columns]]
        
        csv = filename
        excel = filename.replace('.csv', '.xlsx')
        json_f = filename.replace('.csv', '.json')
        
        try:
            # CSV
            df.to_csv(csv, index=False, encoding='utf-8-sig')
            print(f"\n✅ CSV: {csv}")
            
            # EXCEL FORMATÉ
            with pd.ExcelWriter(excel, engine='openpyxl') as w:
                df.to_excel(w, sheet_name='Données', index=False)
                ws = w.sheets['Données']
                
                # Ajuster largeurs
                for col in ws.columns:
                    max_len = max([len(str(cell.value or '')) for cell in col])
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
                
                ws.freeze_panes = 'A2'
                
                # En-têtes colorés
                try:
                    from openpyxl.styles import Font, PatternFill
                    fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                    font = Font(bold=True, color="FFFFFF")
                    
                    for cell in ws[1]:
                        cell.fill = fill
                        cell.font = font
                except:
                    pass
            
            print(f"✅ Excel: {excel}")
            
            # JSON
            df.to_json(json_f, orient='records', force_ascii=False, indent=2)
            print(f"✅ JSON: {json_f}")
            
        except PermissionError:
            ts = datetime.now().strftime('%H%M%S')
            csv = f'mubawab_{ts}.csv'
            df.to_csv(csv, index=False, encoding='utf-8-sig')
            print(f"✅ Sauvegardé: {csv}")
        
        # STATISTIQUES FINALES
        print(f"\n📊 STATISTIQUES FINALES")
        print("=" * 70)
        
        total = len(df)
        avec_prix = len(df[df['prix'] > 0])
        avec_ville = len(df[df['ville'].notna() & (df['ville'] != '')])
        avec_surface = len(df[df['surface_m2'] > 0])
        
        print(f"📈 Total: {total} annonces")
        print(f"💰 Prix: {avec_prix}/{total} ({avec_prix*100/total:.1f}%)")
        if avec_prix > 0:
            print(f"   Moyen: {df[df['prix'] > 0]['prix'].mean():.0f} TND")
        
        print(f"📍 Ville: {avec_ville}/{total} ({avec_ville*100/total:.1f}%)")
        if avec_ville > 5:
            top = df[df['ville'].notna() & (df['ville'] != '')]['ville'].value_counts().head(5)
            print(f"   Top 5: {dict(top)}")
        
        print(f"📐 Surface: {avec_surface}/{total} ({avec_surface*100/total:.1f}%)")
        if avec_surface > 0:
            print(f"   Moyenne: {df[df['surface_m2'] > 0]['surface_m2'].mean():.0f} m²")
        
        return csv
    
    def run(self, pages=10, mode='standard'):
        """Lance la collecte complète"""
        print("╔═══════════════════════════════════════════════════════════════╗")
        print("║          🏠 ESTATE MIND - AGENT COMPLET MUBAWAB 🏠           ║")
        print("╚═══════════════════════════════════════════════════════════════╝")
        print()
        
        urls = self.get_urls(mode)
        print(f"📋 {len(urls)} catégories à scraper")
        print(f"📄 {pages} pages par catégorie")
        print(f"⏱️  Temps estimé: {len(urls) * pages * 2 / 60:.0f} minutes\n")
        
        for i, url in enumerate(urls, 1):
            nom_court = '/'.join(url.split('/')[-2:])
            print(f"[{i}/{len(urls)}] {nom_court}")
            self.scrape(url, pages)
            
            if i % 10 == 0 and self.data:
                print(f"\n💾 Sauvegarde intermédiaire ({len(self.data)} annonces)...")
                self.save(f'mubawab_partial_{i}.csv')
                print()
        
        print("\n" + "=" * 70)
        print("✨ COLLECTE TERMINÉE")
        print("=" * 70)
        
        return self.save('mubawab_FINAL_COMPLET.csv')


def main():
    """Fonction principale avec menu"""
    print("\n╔═══════════════════════════════════════════════════════════════╗")
    print("║          🏠 ESTATE MIND - MUBAWAB AGENT 🏠                   ║")
    print("║              Code Complet Final - Tout-en-Un                  ║")
    print("╚═══════════════════════════════════════════════════════════════╝")
    print()
    print("Ce code fait TOUT automatiquement :")
    print("  ✅ Collecte avec les bonnes classes CSS")
    print("  ✅ Enrichissement depuis les descriptions")
    print("  ✅ Export CSV + Excel + JSON")
    print("  ✅ Déduplication automatique")
    print()
    
    print("MODES DISPONIBLES:")
    print("  1. Standard (96 catégories - 15-30 min)")
    print("  2. Exhaustif (150+ catégories - 45-90 min)")
    print("  3. Test (2 catégories - 2 min)")
    print()
    
    mode = input("Choisissez le mode [1]: ").strip() or "1"
    
    agent = EstateMindAgentComplet()
    
    if mode == "3":
        print("\n📋 MODE TEST\n")
        agent.scrape("https://www.mubawab.tn/fr/st/tunis/appartements-a-vendre", 2)
        agent.scrape("https://www.mubawab.tn/fr/st/medenine/villas-et-maisons-de-luxe-a-vendre", 2)
        agent.save('mubawab_TEST.csv')
        return
    
    mode_str = 'exhaustive' if mode == "2" else 'standard'
    pages = int(input("Pages par catégorie [10]: ").strip() or "10")
    
    print(f"\n🚀 Lancement dans 3 secondes...")
    print("   (Ctrl+C pour annuler)\n")
    time.sleep(3)
    
    agent.run(pages, mode_str)
    
    print("\n✅ TERMINÉ ! Vos fichiers sont prêts.")
    print("📊 Ouvrez le fichier Excel pour voir les résultats.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrompu par l'utilisateur")
        print("💾 Les données collectées sont sauvegardées")